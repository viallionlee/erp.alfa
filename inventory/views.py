from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, FileResponse, HttpResponseRedirect
from .models import Stock, HistoryImportStock, Inbound, InboundItem, Supplier, OpnameQueue, OpnameHistory, StockCardEntry, RakOpnameSession, RakOpnameItem, RakCapacity
from products.models import Product, ProductExtraBarcode
import pandas as pd
import io
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone
from django.db.models import Q, F, Subquery, OuterRef, Value, Count, Sum
from django.db.models.functions import Coalesce
from django.db import models
from django.views.decorators.http import require_GET, require_POST
from django.db import transaction
import xlsxwriter
import tempfile
import logging
from django.urls import reverse
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
from django.utils import timezone as dj_timezone
from inventory.models import Rak # Pastikan ini sudah diimpor
from django.contrib.contenttypes.models import ContentType # Pastikan ini sudah diimpor
from django.contrib.contenttypes.fields import GenericForeignKey # Pastikan ini sudah diimpor
from inventory.models import InventoryRakStockLog # Pastikan ini sudah diimpor
import json
from inventory.models import InventoryRakStock # Pastikan ini sudah diimpor
from inventory.models import PutawaySlottingLog

@login_required
def mobile_inventory(request):
    """
    Mobile view untuk inventory - auto redirect jika diakses dari mobile
    """
    return render(request, 'inventory/mobile_inventory.html')
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.template.loader import render_to_string


logger = logging.getLogger("import_massal")
logging.basicConfig(level=logging.INFO)

TIPE_PERGERAKAN = [
    ('inbound', 'Inbound ke Putaway (+)'),  # Ubah label
    ('reverse_inbound', 'Reverse Inbound (-)'),
    ('outbound', 'Outbound (-)'),
    ('reverse_outbound', 'Reverse Outbound (+)'),
    ('opname_in', 'Opname In (+)'),
    ('opname_out', 'Opname Out (-)'),
    ('close_batch', 'Close Batch (-)'),
    ('reverse_batch', 'Reverse Batch (+)'),
    ('return_stock', 'Return Stock (+)'),
    ('reject_stock', 'Reject Stock (-)'),
    ('pindah_rak', 'Pindah Rak'),
    ('pindah_gudang', 'Pindah Gudang'),
    ('koreksi_stok', 'Koreksi Stok (+/-)'),
    ('putaway_complete', 'Putaway Selesai (-)'),  # Tambah baru
]

def create_stock_card_entry(product, tipe_pergerakan, qty, qty_awal, qty_akhir, notes, user=None, reference=None):
    """
    Helper function untuk membuat entry kartu stock
    """
    entry = StockCardEntry.objects.create(
        product=product,
        tipe_pergerakan=tipe_pergerakan,
        qty=qty,
        qty_awal=qty_awal,
        qty_akhir=qty_akhir,
        notes=notes,
        user=user
    )
    
    # Set reference jika ada
    if reference:
        entry.reference = reference
        entry.save(update_fields=['content_type', 'object_id'])
    
    return entry

@login_required
@permission_required('inventory.view_stock', raise_exception=True)
def index(request):
    # Auto redirect ke mobile jika diakses dari device mobile
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = any(device in user_agent for device in ['mobile', 'android', 'iphone', 'ipad', 'tablet'])
    
    if is_mobile:
        # Render mobile version of index (stock overview)
        # Calculate putaway summary for mobile
        from django.db.models import Sum
        from inventory.models import Stock
        
        putaway_summary = Stock.objects.filter(quantity_putaway__gt=0).aggregate(
            total_items=Count('product'),
            total_quantity=Sum('quantity_putaway')
        )
        
        context = {
            'product_list': Product.objects.all(),
            'total_putaway_items': putaway_summary['total_items'] or 0,
            'total_putaway_quantity': putaway_summary['total_quantity'] or 0,
        }
        return render(request, 'inventory/mobile_index.html', context)
    
    # Mengambil semua produk
    # product_list = Product.objects.all().order_by('nama_produk')

    # Mengambil product_list untuk DataTables via AJAX di stock_data
    # Jadi di sini hanya perlu render template
    
    # NEW: Calculate total putaway items and quantity
    from django.db.models import Sum
    from inventory.models import Stock # Import the Stock model

    putaway_summary = Stock.objects.filter(quantity_putaway__gt=0).aggregate(
        total_items=Count('product'),
        total_quantity=Sum('quantity_putaway')
    )

    total_putaway_items = putaway_summary['total_items'] or 0
    total_putaway_quantity = putaway_summary['total_quantity'] or 0

    context = {
        'product_list': Product.objects.all(), # product_list can be an empty QS here
        'total_putaway_items': total_putaway_items,
        'total_putaway_quantity': total_putaway_quantity,
    }
    return render(request, 'inventory/index.html', context)

@login_required
@permission_required('inventory.view_stock', raise_exception=True)
def export_stock(request):
    # Export all products with stock info to Excel
    products = Product.objects.all().select_related()
    stock_map = {s.product_id: s for s in Stock.objects.all()}
    data = []
    for p in products:
        stock = stock_map.get(p.id)
        # Perbaiki: gunakan property quantity_ready_virtual atau hitung manual
        quantity_ready = 0
        if stock:
            quantity_ready = stock.quantity - stock.quantity_locked
        
        data.append({
            'SKU': p.sku,
            'Barcode': p.barcode,
            'Nama Produk': p.nama_produk,
            'Variant': p.variant_produk,
            'Brand': p.brand,
            'Quantity': stock.quantity if stock else 0,
            'Quantity Locked': stock.quantity_locked if stock else 0,
            'Quantity Ready': quantity_ready,  # Perbaiki: gunakan variabel yang sudah dihitung
        })
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Stock')
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=stock_export.xlsx'  # Perbaiki nama file
    return response

@csrf_exempt
@login_required
@permission_required('inventory.add_stock', raise_exception=True)
def import_stock(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        ext = file.name.split('.')[-1].lower()
        try:
            if ext in ['xls', 'xlsx']:
                df = pd.read_excel(file)
            elif ext == 'csv':
                df = pd.read_csv(file)
            else:
                messages.error(request, 'Format file tidak didukung.')
                return redirect('inventory:index')
            required_cols = ['SKU', 'Barcode', 'Nama Produk', 'Variant', 'Brand', 'Quantity', 'Quantity Locked']
            for col in required_cols:
                if col not in df.columns:
                    messages.error(request, f'Kolom {col} wajib ada di file.')
                    return redirect('inventory:index')
            success, failed = 0, 0
            notes = []
            for _, row in df.iterrows():
                sku = str(row['SKU']).strip()
                try:
                    product = Product.objects.get(sku=sku)
                    stock, _ = Stock.objects.get_or_create(product=product)
                    stock.quantity = int(row['Quantity'])
                    stock.quantity_locked = int(row['Quantity Locked'])
                    stock.save()
                    success += 1
                except Exception as e:
                    failed += 1
                    notes.append(f"SKU {sku}: {e}")
            HistoryImportStock.objects.create(
                imported_by=str(request.user),
                file_name=file.name,
                notes='\n'.join(notes) if notes else 'Berhasil',
                success_count=success,
                failed_count=failed
            )
            messages.success(request, f'Import selesai. Sukses: {success}, Gagal: {failed}')
        except Exception as e:
            messages.error(request, f'Gagal import: {e}')
        return redirect('inventory:index')
    return redirect('inventory:index')

@login_required
@permission_required('inventory.view_inbound', raise_exception=True)
def inbound_list(request):
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    
    template_name = 'inventory/inbound_mobile.html' if is_mobile else 'inventory/inbound.html'
    
    # Ambil semua data untuk desktop, atau halaman pertama untuk mobile
    inbounds_qs = Inbound.objects.all().order_by('-tanggal', '-id')
    
    if is_mobile:
        paginator = Paginator(inbounds_qs, 15) # 15 item per halaman
        inbounds = paginator.get_page(1)
    else:
        inbounds = inbounds_qs

    context = {
        'inbound_list': inbounds
    }
    return render(request, template_name, context)

@login_required
@permission_required('inventory.view_inbound', raise_exception=True)
def inbound_list_api(request):
    page = request.GET.get('page', 1)
    query = request.GET.get('query', '')

    inbounds_qs = Inbound.objects.select_related('supplier').all().order_by('-tanggal', '-id')

    if query:
        inbounds_qs = inbounds_qs.filter(
            Q(nomor_inbound__icontains=query) |
            Q(supplier__nama_supplier__icontains=query)
        )

    paginator = Paginator(inbounds_qs, 15) # 15 item per halaman

    try:
        inbounds_page = paginator.page(page)
    except (EmptyPage, PageNotAnInteger):
        return JsonResponse({'html': '', 'has_more': False})

    html = render_to_string(
        'inventory/partials/inbound_cards.html',
        {'inbound_list': inbounds_page},
        request=request
    )

    return JsonResponse({'html': html, 'has_more': inbounds_page.has_next()})

def get_next_inbound_id():
    last = Inbound.objects.order_by('-id').first()
    return (last.id + 1) if last else 1

@login_required
@permission_required('inventory.add_inbound', raise_exception=True)
def inbound_tambah(request):
    """
    Menangani pembuatan data Inbound baru, item-itemnya,
    dan melakukan update ke tabel stok.
    Mengikuti pola yang mirip dengan `orders.views.add_order`.
    """
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    template_name = 'inventory/inbound_tambah_mobile.html' if is_mobile else 'inventory/inbound_tambah.html'

    if request.method == 'POST':
        try:
            # Menggunakan transaction.atomic untuk memastikan semua operasi 
            # berhasil atau tidak sama sekali.
            with transaction.atomic():
                # 1. Ambil data dari form
                nomor_inbound = request.POST.get('nomor_inbound')
                tanggal = request.POST.get('tanggal')
                keterangan = request.POST.get('keterangan')
                from_warehouse = request.POST.get('from_warehouse')
                to_warehouse = request.POST.get('to_warehouse')
                produk_sku = request.POST.getlist('produk_sku[]')
                produk_qty = request.POST.getlist('produk_qty[]')

                if not (nomor_inbound and tanggal and from_warehouse and to_warehouse and produk_sku):
                    messages.error(request, "Nomor Inbound, Tanggal, Dari Gudang, Ke Gudang, dan minimal 1 produk wajib diisi.")
                    # Kembalikan ke form dengan data yang sudah diisi
                    context = {
                        'default_nomor_inbound': nomor_inbound,
                        'default_tanggal': tanggal,
                        'keterangan': keterangan,
                        'from_warehouse': from_warehouse,
                        'to_warehouse': to_warehouse,
                    }
                    return render(request, template_name, context)

                # 2. Buat objek Inbound (Transfer Antar Gudang)
                inbound_obj, created = Inbound.objects.update_or_create(
                    nomor_inbound=nomor_inbound,
                    defaults={
                        'tanggal': tanggal,
                        'keterangan': keterangan or '',
                        'from_warehouse': from_warehouse,
                        'to_warehouse': to_warehouse,
                        'created_by': request.user,
                        'received_by': request.user,
                        'received_at': timezone.now(),
                    }
                )

                # 3. Hapus item lama jika ini adalah proses update (pola dari add_order)
                if not created:
                    # Ambil item lama untuk mengkoreksi stok
                    old_items = InboundItem.objects.filter(inbound=inbound_obj)
                    for item in old_items:
                        try:
                            # Kunci baris stok untuk update
                            stock = Stock.objects.select_for_update().get(product_id=item.product_id)
                            qty_awal = stock.quantity
                            stock.quantity -= item.quantity # Operasi di Python
                            stock.quantity_putaway -= item.quantity # Koreksi stok belum di-putaway
                            stock.save(update_fields=['quantity', 'quantity_putaway'])
                            
                            # Catat kartu stock untuk koreksi
                            create_stock_card_entry(
                                product=item.product,
                                tipe_pergerakan='reverse_inbound', # Tipe yang benar untuk koreksi
                                qty=-item.quantity,  # Negatif karena koreksi
                                qty_awal=qty_awal,
                                qty_akhir=stock.quantity,
                                notes=f'Koreksi Inbound {inbound_obj.nomor_inbound}',
                                user=request.user,
                                reference=inbound_obj
                            )
                        except Stock.DoesNotExist:
                            # Jika stok tidak ada, tidak perlu melakukan apa-apa untuk koreksi
                            pass
                    old_items.delete()

                # 4. Loop untuk membuat InboundItem dan update stok (seperti Order di add_order)
                for sku, qty_str in zip(produk_sku, produk_qty):
                    if not (sku and qty_str): continue # Lewati baris kosong
                    
                    qty = int(qty_str)
                    product = Product.objects.filter(sku=sku).first()

                    if not product:
                        messages.warning(request, f"SKU '{sku}' tidak ditemukan di master produk dan dilewati.")
                        continue
                    
                    # Buat detail item
                    InboundItem.objects.create(
                        inbound=inbound_obj,
                        product=product,
                        quantity=qty
                    )

                    # --- LOGIKA UPDATE STOK DENGAN ROW LOCKING ---
                    stock, stock_created = Stock.objects.select_for_update().get_or_create(
                        product=product,
                        defaults={'sku': sku, 'quantity': 0, 'quantity_locked': 0, 'quantity_putaway': 0}
                    )

                    qty_awal = stock.quantity
                    stock.quantity += qty # Langsung ke inventory stock
                    stock.quantity_putaway += qty # Stok belum di-putaway
                    stock.save(update_fields=['quantity', 'quantity_putaway'])

                    create_stock_card_entry(
                        product=product,
                        tipe_pergerakan='inbound',
                        qty=qty,
                        qty_awal=qty_awal,
                        qty_akhir=stock.quantity,
                        notes=f'Inbound {inbound_obj.nomor_inbound}' + (' (Stok Baru)' if stock_created else ''),
                        user=request.user,
                        reference=inbound_obj
                    )

                messages.success(request, f"Data Inbound '{nomor_inbound}' berhasil disimpan.")
                return redirect('inventory:inbound')

        except Exception as e:
            messages.error(request, f"Terjadi kesalahan fatal: {e}")
            # Kembali ke halaman tambah, jangan redirect ke list
            return redirect('inventory:inbound_tambah')

    # Untuk request GET
    else:
        next_id = get_next_inbound_id()
        now = timezone.localtime()
        nomor_inbound = f"INV/{now.strftime('%d-%m-%Y')}/{next_id}"
        tanggal = now.strftime('%Y-%m-%dT%H:%M')
        context = {
            'default_nomor_inbound': nomor_inbound,
            'default_tanggal': tanggal,
        }
        return render(request, template_name, context)

@login_required
@permission_required('inventory.view_inbound', raise_exception=True)
def produk_lookup(request):
    q = request.GET.get('q', '').strip()
    mode = request.GET.get('mode', '')
    if mode == 'manual':
        # Search by SKU, Barcode, Nama Produk, Variant, Brand (case-insensitive, contains)
        produk_qs = Product.objects.filter(
            Q(sku__icontains=q) |
            Q(barcode__icontains=q) |
            Q(nama_produk__icontains=q) |
            Q(variant_produk__icontains=q) |
            Q(brand__icontains=q)
        ).order_by('nama_produk')
        # Hilangkan paginasi, ambil semua data
        data = []
        for p in produk_qs:
            photo_url = ''
            if p.photo:
                try:
                    # Check if file exists
                    if p.photo.storage.exists(p.photo.name):
                        photo_url = p.photo.url
                except:
                    pass

            # Get stock quantity from inventory.stock
            qty_fisik = 0
            try:
                stock = p.stock
                if stock:
                    # Gunakan quantity_ready_virtual (stok siap jual)
                    qty_fisik = stock.quantity_ready_virtual
            except:
                qty_fisik = 0

            data.append({
                'id': p.id,
                'sku': p.sku,
                'barcode': p.barcode,
                'nama': p.nama_produk,
                'variant': p.variant_produk or '',
                'brand': p.brand or '',
                'photo_url': photo_url,
                'qty_fisik': qty_fisik,
                'location': '-'  # Default location
            })
        return JsonResponse(data, safe=False)
    # Default: scan barcode/sku exact
    produk = Product.objects.filter(Q(barcode=q) | Q(sku=q)).first()
    if produk:
        photo_url = ''
        try:
            if produk.photo:
                photo_url = produk.photo.url
        except ValueError:
            photo_url = '' # Handle missing file safely

        # Get stock quantity from inventory.stock
        qty_fisik = 0
        try:
            stock = produk.stock
            if stock:
                # Gunakan quantity_ready_virtual (stok siap jual)
                qty_fisik = stock.quantity_ready_virtual
        except:
            qty_fisik = 0

        return JsonResponse({
            'id': produk.id,
            'sku': produk.sku,
            'barcode': produk.barcode,
            'nama': produk.nama_produk,
            'variant': produk.variant_produk or '',
            'brand': produk.brand or '',
            'photo_url': photo_url,
            'qty_fisik': qty_fisik,
            'location': '-'  # Default location
        })
    return JsonResponse({'error': 'Produk tidak ditemukan'}, status=404)

@login_required
@permission_required('inventory.view_inbound', raise_exception=True)
def inbound_detail(request, pk):
    inbound = get_object_or_404(Inbound, pk=pk)
    items = InboundItem.objects.filter(inbound=inbound).select_related('product')

    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    template_name = 'inventory/inbound_detail_mobile.html' if is_mobile else 'inventory/inbound_detail.html'

    context = {
        'inbound': inbound,
        'items': items
    }
    return render(request, template_name, context)

@login_required
@permission_required('inventory.delete_inbound', raise_exception=True)
def inbound_delete(request, pk):
    inbound = get_object_or_404(Inbound, pk=pk)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                for item in inbound.items.all():
                    stock = Stock.objects.select_for_update().filter(product=item.product).first()
                    if stock:
                        qty_awal = stock.quantity
                        jumlah_hapus = item.quantity

                        stock.quantity -= jumlah_hapus
                        stock.quantity_putaway -= jumlah_hapus
                        stock.save(update_fields=['quantity', 'quantity_putaway'])

                        create_stock_card_entry(
                            product=item.product,
                            tipe_pergerakan='reverse_inbound',
                            qty=-jumlah_hapus,
                            qty_awal=qty_awal,
                            qty_akhir=stock.quantity,
                            notes=f"Hapus Inbound: {inbound.nomor_inbound}",
                            user=request.user,
                            reference=inbound
                        )
                
                inbound.delete()
                messages.success(request, f'Inbound {inbound.nomor_inbound} berhasil dihapus.')
                return redirect('inventory:inbound')
        except Exception as e:
            messages.error(request, f"Gagal menghapus inbound: {str(e)}")
            return redirect('inventory:inbound_detail', pk=pk)

    # Prepare data for template with calculated stock after delete
    items_data = []
    for item in inbound.items.all():
        current_stock = item.product.stock
        stock_after_delete = {
            'quantity': (current_stock.quantity if current_stock else 0) - item.quantity,
            'quantity_putaway': (current_stock.quantity_putaway if current_stock else 0) - item.quantity,
            'quantity_locked': current_stock.quantity_locked if current_stock else 0,
            'quantity_ready': (current_stock.quantity_ready_virtual if current_stock else 0) - item.quantity,
        }
        items_data.append({
            'item': item,
            'stock_after_delete': stock_after_delete
        })
    
    context = {
        'inbound': inbound,
        'items_data': items_data
    }
    return render(request, 'inventory/inbound_confirm_delete.html', context)

@csrf_exempt
@login_required
@permission_required('inventory.add_inbound', raise_exception=True)
def inbound_import_massal(request):
    import pandas as pd
    from django.utils import timezone
    from django.db import transaction
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        ext = file.name.split('.')[-1].lower()

        try:
            if ext in ['xls', 'xlsx']:
                df = pd.read_excel(file)
            elif ext == 'csv':
                df = pd.read_csv(file)
            else:
                messages.error(request, 'Format file tidak didukung.')
                return redirect('inventory:inbound')

            # --- LOGIKA BARU DENGAN HEADER SEDERHANA & KONSISTEN ---
            
            # 1. Definisikan header sederhana yang wajib ada
            sku_header = 'sku'
            quantity_header = 'quantity'

            # 2. Cek keberadaan header
            if sku_header not in df.columns or quantity_header not in df.columns:
                messages.error(request, f"File yang diupload tidak valid. Pastikan kolom '{sku_header}' dan '{quantity_header}' ada.")
                return redirect('inventory:inbound')
            
            # 3. Filter baris yang relevan (hanya yang quantity-nya diisi)
            df[quantity_header] = pd.to_numeric(df[quantity_header], errors='coerce')
            df.dropna(subset=[quantity_header], inplace=True)
            df = df[df[quantity_header] > 0]
            
            if df.empty:
                messages.warning(request, "Tidak ada data quantity yang valid untuk diimpor.")
                return redirect('inventory:inbound')

            # 4. Proses data
            now = timezone.localtime()
            nomor_inbound = f"IMP-MASSAL/{now.strftime('%d%m%Y-%H%M%S')}"
            
            supplier_id = request.POST.get('import_supplier_id')
            supplier = None
            if supplier_id:
                try:
                    supplier = Supplier.objects.get(id=supplier_id)
                except Supplier.DoesNotExist:
                    supplier = None

            with transaction.atomic():
                inbound = Inbound.objects.create(
                    nomor_inbound=nomor_inbound,
                    tanggal=now,
                    keterangan=f'Import massal dari file: {file.name}',
                    supplier=supplier
                )

                success_count = 0
                failed_count = 0
                notes = []
                
                # Ambil semua SKU yang unik untuk efisiensi query
                skus = df[sku_header].astype(str).str.strip().unique()
                products = Product.objects.filter(sku__in=skus).in_bulk(field_name='sku')

                for _, row in df.iterrows():
                    sku = str(row[sku_header]).strip()
                    quantity = int(row[quantity_header])
                    
                    product = products.get(sku)

                    if not product:
                        failed_count += 1
                        notes.append(f"SKU '{sku}' tidak ditemukan dan dilewati.")
                        continue
                    
                    try:
                        InboundItem.objects.create(
                            inbound=inbound,
                            product=product,
                            quantity=quantity
                        )

                        stock, _ = Stock.objects.select_for_update().get_or_create(
                            product=product,
                            defaults={'sku': product.sku, 'quantity': 0, 'quantity_locked': 0, 'quantity_putaway': 0}
                        )
                        
                        qty_awal = stock.quantity
                        stock.quantity += quantity
                        stock.save(update_fields=['quantity'])

                        create_stock_card_entry(
                            product=product,
                            tipe_pergerakan='inbound',
                            qty=quantity,
                            qty_awal=qty_awal,
                            qty_akhir=stock.quantity,
                            notes=f'Inbound Massal: {inbound.nomor_inbound}',
                            user=request.user,
                            reference=inbound
                        )
                        success_count += 1

                    except Exception as e:
                        failed_count += 1
                        notes.append(f"Gagal memproses SKU '{sku}': {e}")

            messages.success(request, f'Impor Inbound selesai. Sukses: {success_count}, Gagal: {failed_count}.')
            if notes:
                messages.warning(request, '\n'.join(notes))
        
        except Exception as e:
            messages.error(request, f'Gagal memproses file: {e}')
        
        return redirect('inventory:inbound')

    return redirect('inventory:inbound')


@require_GET
@login_required
@permission_required('inventory.view_stock', raise_exception=True)
def stock_data(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        order_column_index = int(request.GET.get('order[0][column]', 0))
        order_dir = request.GET.get('order[0][dir]', 'asc')
        columns = ['sku', 'barcode', 'nama_produk', 'variant_produk', 'brand', 'quantity', 'quantity_putaway', 'quantity_locked', 'quantity_ready']
        order_column_name = columns[order_column_index] if order_column_index < len(columns) else 'sku'

        # Annotate with stock quantities for sorting and filtering
        stock_qs = Stock.objects.filter(product_id=OuterRef('pk'))
        queryset = Product.objects.annotate(
            quantity=Coalesce(Subquery(stock_qs.values('quantity')[:1]), 0, output_field=models.IntegerField()),
            quantity_putaway=Coalesce(Subquery(stock_qs.values('quantity_putaway')[:1]), 0, output_field=models.IntegerField()),
            quantity_locked=Coalesce(Subquery(stock_qs.values('quantity_locked')[:1]), 0, output_field=models.IntegerField())
        ).annotate(
            quantity_ready=F('quantity') - F('quantity_locked')
        )
        
        total_count = queryset.count()

        # --- LOGIKA FILTER & SORT ---

        # Global search
        if search_value:
            global_q = (
                Q(sku__icontains=search_value) |
                Q(barcode__icontains=search_value) |
                Q(nama_produk__icontains=search_value) |
                Q(variant_produk__icontains=search_value) |
                Q(brand__icontains=search_value)
            )
            # Add numeric fields to global search if search_value is a number
            if search_value.isdigit():
                global_q |= Q(quantity=search_value)
                global_q |= Q(quantity_putaway=search_value)
                global_q |= Q(quantity_locked=search_value)
                global_q |= Q(quantity_ready=search_value)
            
            queryset = queryset.filter(global_q)

        # Per-column search
        for idx, col_name in enumerate(columns):
            col_search = request.GET.get(f'columns[{idx}][search][value]', '').strip()
            if col_search:
                if col_name in ['sku', 'barcode', 'nama_produk', 'variant_produk', 'brand']:
                    queryset = queryset.filter(**{f"{col_name}__icontains": col_search})
                else: # Kolom dari tabel Stock (now annotated)
                    try:
                        queryset = queryset.filter(**{f"{col_name}": int(col_search)})
                    except (ValueError, TypeError):
                        # Abaikan jika input untuk kolom angka tidak valid
                        pass

        filtered_count = queryset.count()

        # Sorting
        sort_field = order_column_name
        if order_dir == 'desc':
            sort_field = f'-{sort_field}'

        queryset = queryset.order_by(sort_field).distinct()

        # Ambil data setelah di-filter dan di-sort
        paginated_qs = queryset[start:start+length]
        
        data = []
        for p in paginated_qs:
            data.append({
                'sku': p.sku,
                'barcode': p.barcode,
                'nama_produk': p.nama_produk,
                'variant_produk': p.variant_produk or '',
                'brand': p.brand or '',
                'quantity': p.quantity,
                'quantity_putaway': p.quantity_putaway,
                'quantity_locked': p.quantity_locked,
                'quantity_ready': p.quantity_ready,
            })
            
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_count,
            'recordsFiltered': filtered_count,
            'data': data,
        })
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in stock_data view: {e}", exc_info=True)
        return JsonResponse({'error': 'Terjadi kesalahan pada server.', 'details': str(e)}, status=500)

@login_required
@permission_required('inventory.view_inbound', raise_exception=True)
@require_GET
def inbound_data(request):
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    order_column_index = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'asc')
    columns = ['id', 'nomor_inbound', 'tanggal', 'from_warehouse', 'to_warehouse', 'jumlah_sku', 'received_by', 'received_at', 'keterangan']
    order_column = columns[order_column_index] if order_column_index < len(columns) else 'id'
    if order_dir == 'desc':
        order_column = '-' + order_column

    queryset = Inbound.objects.select_related('received_by').all()
    total_count = queryset.count()
    
    # Search
    if search_value:
        queryset = queryset.filter(
            Q(id__icontains=search_value) |
            Q(nomor_inbound__icontains=search_value) |
            Q(from_warehouse__icontains=search_value) |
            Q(to_warehouse__icontains=search_value) |
            Q(keterangan__icontains=search_value)
        )
    
    filtered_count = queryset.count()

    queryset = queryset.order_by(order_column)[start:start+length]

    data = []
    for inbound in queryset:
        data.append({
            'id': inbound.id,
            'nomor_inbound': inbound.nomor_inbound,
            'tanggal': timezone.localtime(inbound.tanggal).strftime('%d-%m-%Y %H:%M'),
            'from_warehouse': inbound.from_warehouse or '-',
            'to_warehouse': inbound.to_warehouse or '-',
            'jumlah_sku': inbound.items.count(),
            'received_by': inbound.received_by.username if inbound.received_by else '-',
            'received_at': timezone.localtime(inbound.received_at).strftime('%d-%m-%Y %H:%M') if inbound.received_at else '-',
            'keterangan': inbound.keterangan or '',
            'aksi': f'<a href="/inventory/inbound/{inbound.id}/" class="btn btn-sm btn-info"><i class="bi bi-eye"></i> Detail</a> '
                    f'<a href="/inventory/inbound/{inbound.id}/edit/" class="btn btn-sm btn-warning"><i class="bi bi-pencil"></i> Edit</a> '
                    f'<a href="/inventory/inbound/{inbound.id}/delete/" class="btn btn-sm btn-danger"><i class="bi bi-trash"></i> Delete</a>'
        })
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_count,
        'recordsFiltered': filtered_count,
        'data': data,
    })

def download_template_import_stock(request):
    """
    Menghasilkan file Excel untuk template import stok.
    Formatnya sama dengan ekspor produk, ditambah kolom 'quantity' kosong.
    """
    # Ambil semua data produk yang relevan
    products_qs = Product.objects.all().values(
        'id', 'sku', 'barcode', 'nama_produk', 'variant_produk', 'brand', 'rak'
    )
    
    # Konversi ke DataFrame
    df = pd.DataFrame(list(products_qs))
    
    # Tambahkan kolom 'quantity' yang kosong
    df['quantity'] = ''
    
    # Atur urutan kolom sesuai permintaan
    df = df[['id', 'sku', 'barcode', 'nama_produk', 'variant_produk', 'brand', 'rak', 'quantity']]
    
    # Ganti nama header untuk kejelasan
    df.rename(columns={
        'id': 'product_id (JANGAN DIUBAH)',
        'sku': 'SKU (Hanya Tampilan)',
        'barcode': 'Barcode (Hanya Tampilan)',
        'nama_produk': 'Nama Produk (Hanya Tampilan)',
        'variant_produk': 'Variant (Hanya Tampilan)',
        'brand': 'Brand (Hanya Tampilan)',
        'rak': 'Rak (Hanya Tampilan)',
        'quantity': 'Isi Jumlah Stok Di Sini'
    }, inplace=True)

    # Buat file Excel di memori
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Template Import Stok')
        # Opsional: Atur lebar kolom agar lebih mudah dibaca
        worksheet = writer.sheets['Template Import Stok']
        worksheet.set_column('A:A', 25) # product_id
        worksheet.set_column('B:B', 20) # sku
        worksheet.set_column('C:C', 20) # barcode
        worksheet.set_column('D:D', 40) # nama_produk
        worksheet.set_column('E:H', 25) # sisa kolom
        
    output.seek(0)
    
    # Kirim sebagai response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=template_import_stock.xlsx'
    return response

def download_template_inbound(request):
    """
    Menghasilkan file Excel template untuk import inbound.
    Hanya berisi header kolom yang diperlukan, tanpa data produk.
    """
    # 1. Buat workbook dan response
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Template Import Inbound'
    
    # Siapkan response untuk file Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=template_import_inbound.xlsx'

    # 2. Definisikan header kolom yang diperlukan untuk import inbound
    headers = [
        'sku', 'quantity'
    ]
    worksheet.append(headers)
    
    # 3. Tambahkan satu baris contoh kosong untuk menunjukkan format
    example_row = [
        '',  # sku - kosong, akan diisi user  
        ''   # quantity - kosong, akan diisi user
    ]
    worksheet.append(example_row)
        
    # 4. Membuat style
    bold_font = Font(bold=True)
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Light Blue
    gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid") # Light Gray
    
    # Terapkan bold ke semua header
    for cell in worksheet["1:1"]:
        cell.font = bold_font

    # Terapkan highlight biru ke kolom B (quantity) untuk header dan contoh
    for row in range(1, 3):  # Hanya baris 1 (header) dan 2 (contoh)
        worksheet[f'B{row}'].fill = blue_fill
    
    # Terapkan highlight abu-abu ke kolom A (sku) untuk baris contoh
    worksheet['A2'].fill = gray_fill
        
    # 5. Atur lebar kolom agar mudah dibaca
    worksheet.column_dimensions['A'].width = 25  # sku
    worksheet.column_dimensions['B'].width = 20  # quantity

    # 6. Simpan workbook LANGSUNG ke dalam 'response'
    workbook.save(response)
    
    return response

@require_POST
@csrf_exempt
@login_required
@permission_required('inventory.add_inbound', raise_exception=True)
def add_supplier(request):
    nama = request.POST.get('nama_supplier', '').strip()
    nomor_telepon = request.POST.get('nomor_telepon', '').strip()
    alamat = request.POST.get('alamat_supplier', '').strip()
    kota = request.POST.get('kota_supplier', '').strip()
    brand = request.POST.get('brand', '').strip()
    if not nama:
        return JsonResponse({'error': 'Nama supplier wajib diisi!'}, status=400)
    supplier, created = Supplier.objects.get_or_create(
        nama_supplier=nama,
        defaults={
            'nomor_telepon': nomor_telepon,
            'alamat': alamat,
            'kota': kota,
            'brand': brand,
        }
    )
    return JsonResponse({
        'nama_supplier': supplier.nama_supplier,
        'id': supplier.id,
        'created': created,
    })

@require_GET
@login_required
@permission_required('inventory.view_inbound', raise_exception=True)
def search_supplier(request):
    q = request.GET.get('q', '').strip()
    
    qs = Supplier.objects.all()
    if q:
        qs = qs.filter(
            Q(nama_supplier__icontains=q) |
            Q(alamat__icontains=q) |
            Q(kota__icontains=q) |
            Q(brand__icontains=q)
        )
    
    suppliers_data = [
        {
            'id': s.id,
            'nama_supplier': s.nama_supplier,
            'nomor_telepon': s.nomor_telepon or '',
            'alamat_supplier': s.alamat or '',
            'kota_supplier': s.kota or '',
            'brand': s.brand or ''
        }
        for s in qs.order_by('nama_supplier')[:20]
    ]
    
    return JsonResponse(suppliers_data, safe=False)

def brand_list(request):
    brands = Product.objects.values_list('brand', flat=True).distinct().order_by('brand')
    return JsonResponse(list(filter(None, brands)), safe=False)

def opnamequeue_list(request):
    opnamequeue = OpnameQueue.objects.filter(status='pending').order_by('prioritas', 'terakhir_opname', 'ditambahkan_pada')
    
    # Auto redirect to mobile template if accessed from mobile device
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = any(mobile in user_agent for mobile in ['mobile', 'android', 'iphone', 'ipad', 'tablet'])
    
    template_name = 'inventory/opnamequeue_mobile.html' if is_mobile else 'inventory/opnamequeue.html'
    
    return render(request, template_name, {'opnamequeue': opnamequeue})

def opnamehistory_list(request):
    opnamehistory = OpnameHistory.objects.all().order_by('-tanggal_opname')[:200]
    return render(request, 'inventory/opnamehistory.html', {'opnamehistory': opnamehistory})

def opname_input(request, queue_id):
    queue = get_object_or_404(OpnameQueue, id=queue_id, status='pending')
    stock = Stock.objects.filter(product=queue.product).first()
    if request.method == 'POST':
        qty_fisik = request.POST.get('qty_fisik')
        catatan = request.POST.get('catatan', '')
        try:
            qty_fisik = int(qty_fisik)
        except Exception:
            messages.error(request, 'Qty fisik harus angka!')
            # Auto redirect to mobile template if accessed from mobile device
            user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
            is_mobile = any(mobile in user_agent for mobile in ['mobile', 'android', 'iphone', 'ipad', 'tablet'])
            template_name = 'inventory/opname_input_mobile.html' if is_mobile else 'inventory/opname_input.html'
            return render(request, template_name, {'queue': queue, 'stock': stock})
        qty_sistem = stock.quantity if stock else 0
        selisih = qty_fisik - qty_sistem
        # Update stock
        if stock:
            qty_awal = stock.quantity
            stock.quantity = qty_fisik
            stock.save(update_fields=['quantity'])
            # Catat kartu stock untuk opname
            create_stock_card_entry(
                product=queue.product,
                tipe_pergerakan='koreksi_stok',  # Perbaiki tipe pergerakan
                qty=selisih,
                qty_awal=qty_awal,
                qty_akhir=qty_fisik,
                notes=f'Opname - {catatan}' if catatan else 'Opname',
                user=request.user,
                reference=queue  # Tambah referensi ke queue
            )
        # Catat ke history
        OpnameHistory.objects.create(
            product=queue.product,
            qty_fisik=qty_fisik,
            qty_sistem=qty_sistem,
            selisih=selisih,
            petugas_opname=request.user,
            catatan=catatan
        )
        # Update queue
        queue.status = 'done'
        queue.terakhir_opname = timezone.now()
        queue.catatan = catatan
        queue.save(update_fields=['status', 'terakhir_opname', 'catatan'])
        messages.success(request, f'Opname untuk {queue.product} berhasil!')
        return HttpResponseRedirect(reverse('inventory:opnamequeue_list'))
    
    # Auto redirect to mobile template if accessed from mobile device
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = any(mobile in user_agent for mobile in ['mobile', 'android', 'iphone', 'ipad', 'tablet'])
    template_name = 'inventory/opname_input_mobile.html' if is_mobile else 'inventory/opname_input.html'
    
    return render(request, template_name, {'queue': queue, 'stock': stock})

def stock_card_view(request):
    """
    Menampilkan halaman kartu stok.
    """
    # Mengambil choices dari model untuk dropdown filter
    tipe_pergerakan_list = StockCardEntry.TIPE_PERGERAKAN
    context = {
        'tipe_pergerakan_list': tipe_pergerakan_list
    }
    return render(request, 'inventory/stock_card.html', context)

@require_GET
def stock_card_data(request):
    """
    Provides server-side data for the DataTables on the stock card page.
    """
    try:
        # Base queryset
        queryset = StockCardEntry.objects.select_related('product', 'user').all()

        # Total records
        total_records = queryset.count()

        # Search logic (global search)
        search_value = request.GET.get('search[value]', '')
        if search_value:
            queryset = queryset.filter(
                Q(product__sku__icontains=search_value) |
                Q(product__barcode__icontains=search_value) |
                Q(product__nama_produk__icontains=search_value) |
                Q(tipe_pergerakan__icontains=search_value) |
                Q(user__username__icontains=search_value) |
                Q(notes__icontains=search_value)
            )
        
        # Per-column filtering
        # Waktu (kolom 0)
        waktu_search = request.GET.get('columns[0][search][value]', '')
        if waktu_search:
            try:
                # Handle date search
                queryset = queryset.filter(waktu__date=waktu_search)
            except:
                pass  # Ignore invalid date format

        # SKU (kolom 1)
        sku_search = request.GET.get('columns[1][search][value]', '')
        if sku_search:
            queryset = queryset.filter(product__sku__icontains=sku_search)

        # Barcode (kolom 2) - BARU
        barcode_search = request.GET.get('columns[2][search][value]', '')
        if barcode_search:
            queryset = queryset.filter(product__barcode__icontains=barcode_search)

        # Nama Produk (kolom 3)
        nama_search = request.GET.get('columns[3][search][value]', '')
        if nama_search:
            queryset = queryset.filter(product__nama_produk__icontains=nama_search)

        # Tipe Pergerakan (kolom 4)
        tipe_search = request.GET.get('columns[4][search][value]', '')
        if tipe_search:
            queryset = queryset.filter(tipe_pergerakan=tipe_search)

        # User (kolom 5)
        user_search = request.GET.get('columns[5][search][value]', '')
        if user_search:
            queryset = queryset.filter(user__username__icontains=user_search)
            
        # Status (kolom 6)
        status_search = request.GET.get('columns[6][search][value]', '')
        if status_search:
            queryset = queryset.filter(status=status_search)
            
        # Notes (kolom 10)
        notes_search = request.GET.get('columns[10][search][value]', '')
        if notes_search:
            queryset = queryset.filter(notes__icontains=notes_search)

        # Total filtered records
        total_filtered = queryset.count()

        # Ordering - PERBAIKAN: Gunakan mapping yang benar
        order_column_index = request.GET.get('order[0][column]', '0')
        order_dir = request.GET.get('order[0][dir]', 'desc')
        
        # Mapping kolom untuk ordering
        column_mapping = {
            '0': 'waktu',
            '1': 'product__sku', 
            '2': 'product__barcode',  # BARU
            '3': 'product__nama_produk',
            '4': 'tipe_pergerakan',
            '5': 'user__username',
            '6': 'status',
            '7': 'qty',
            '8': 'qty_awal',
            '9': 'qty_akhir',
            '10': 'notes'
        }
        
        order_column_name = column_mapping.get(order_column_index, 'waktu')
        if order_dir == 'desc':
            order_column_name = f'-{order_column_name}'
        
        # Apply ordering
        if order_column_name in ['waktu', '-waktu', 'product__sku', '-product__sku', 
                               'product__barcode', '-product__barcode',  # BARU
                               'product__nama_produk', '-product__nama_produk', 
                               'tipe_pergerakan', '-tipe_pergerakan', 
                               'user__username', '-user__username', 'status', '-status',
                               'qty', '-qty', 'qty_awal', '-qty_awal', 'qty_akhir', '-qty_akhir',
                               'notes', '-notes']:
            queryset = queryset.order_by(order_column_name)

        # Pagination
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 50))
        queryset = queryset[start:start + length]

        # Prepare data for JSON response
        jakarta_tz = dj_timezone.get_current_timezone()

        data = []
        for entry in queryset:
            try:
                data.append({
                    "waktu": dj_timezone.localtime(entry.waktu, jakarta_tz).strftime('%Y-%m-%d %H:%M:%S') if entry.waktu else '',
                    "product__sku": entry.product.sku if entry.product else '',
                    "product__barcode": entry.product.barcode if entry.product else '',  # BARU
                    "product__nama_produk": entry.product.nama_produk if entry.product else '',
                    "product__variant_produk": entry.product.variant_produk if entry.product else '',
                    "tipe_pergerakan": entry.get_tipe_pergerakan_display(),
                    "user": entry.user.username if entry.user else '-',
                    "qty": entry.qty,
                    "qty_awal": entry.qty_awal,
                    "qty_akhir": entry.qty_akhir,
                    "notes": entry.notes or '',
                    "status": entry.status,
                    "status_display": entry.get_status_display(),
                    "reference_url": entry.reference.get_absolute_url() if hasattr(entry.reference, 'get_absolute_url') else None,
                })
            except Exception as e:
                # Log error but continue processing other entries
                print(f"Error processing entry {entry.id}: {e}")
                continue

        return JsonResponse({
            "draw": int(request.GET.get('draw', 0)),
            "recordsTotal": total_records,
            "recordsFiltered": total_filtered,
            "data": data,
        })
        
    except Exception as e:
        # Return error response instead of crashing
        return JsonResponse({
            "draw": int(request.GET.get('draw', 0)),
            "recordsTotal": 0,
            "recordsFiltered": 0,
            "data": [],
            "error": str(e)
        }, status=500)

@login_required
@permission_required('inventory.change_inbound', raise_exception=True)
def inbound_edit(request, pk):
    inbound = get_object_or_404(Inbound, pk=pk)
    
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    template_name = 'inventory/inbound_edit_mobile.html' if is_mobile else 'inventory/inbound_edit.html'

    if request.method == 'POST':
        try:
            with transaction.atomic():
                from_warehouse = request.POST.get('from_warehouse')
                to_warehouse = request.POST.get('to_warehouse')
                keterangan = request.POST.get('keterangan', '')
                produk_sku = request.POST.getlist('produk_sku[]')
                produk_qty = request.POST.getlist('produk_qty[]')

                # 1. Update Inbound header (Transfer Antar Gudang)
                inbound.from_warehouse = from_warehouse
                inbound.to_warehouse = to_warehouse
                inbound.keterangan = keterangan
                inbound.save(update_fields=['from_warehouse', 'to_warehouse', 'keterangan'])

                # 2. Proses item (menghitung selisih dan update stok)
                old_items = {item.product.sku: item for item in inbound.items.all()}
                current_skus = set()

                for sku, qty_str in zip(produk_sku, produk_qty):
                    if not (sku and qty_str): continue
                    
                    qty = int(qty_str)
                    current_skus.add(sku)
                    product = get_object_or_404(Product, sku=sku)
                    stock, _ = Stock.objects.get_or_create(product=product, defaults={'sku': product.sku, 'quantity': 0, 'quantity_locked': 0, 'quantity_putaway': 0})
                    
                    if sku in old_items:
                        # Produk sudah ada, hitung selisih
                        old_item = old_items[sku]
                        selisih = qty - old_item.quantity
                        
                        if selisih != 0:
                            qty_awal = stock.quantity
                            stock.quantity += selisih
                            stock.quantity_putaway += selisih
                            stock.save(update_fields=['quantity', 'quantity_putaway'])
                            
                            create_stock_card_entry(
                                product=product,
                                tipe_pergerakan='koreksi_stok',
                                qty=selisih,
                                qty_awal=qty_awal,
                                qty_akhir=stock.quantity,
                                notes=f"Edit Inbound {inbound.nomor_inbound}",
                                user=request.user,
                                reference=inbound
                            )
                        
                        # Update item yang ada
                        old_item.quantity = qty
                        old_item.save(update_fields=['quantity'])
                    else:
                        # Produk baru ditambahkan
                        InboundItem.objects.create(inbound=inbound, product=product, quantity=qty)
                        
                        qty_awal = stock.quantity
                        stock.quantity += qty
                        stock.quantity_putaway += qty
                        stock.save(update_fields=['quantity', 'quantity_putaway'])

                        create_stock_card_entry(
                            product=product,
                            tipe_pergerakan='inbound',
                            qty=qty,
                            qty_awal=qty_awal,
                            qty_akhir=stock.quantity,
                            notes=f"Edit Inbound {inbound.nomor_inbound}",
                            user=request.user,
                            reference=inbound
                        )
                
                # Hapus produk yang dihilangkan dari form
                skus_to_delete = set(old_items.keys()) - current_skus
                for sku in skus_to_delete:
                    item_to_delete = old_items[sku]
                    stock = Stock.objects.get(product=item_to_delete.product)
                    
                    qty_awal = stock.quantity
                    stock.quantity -= item_to_delete.quantity # Kurangi stok
                    stock.save(update_fields=['quantity'])
                    
                    create_stock_card_entry(
                        product=item_to_delete.product,
                        tipe_pergerakan='reverse_inbound',
                        qty=-item_to_delete.quantity,
                        qty_awal=qty_awal,
                        qty_akhir=stock.quantity,
                        notes=f"Hapus item dari Inbound {inbound.nomor_inbound}",
                        user=request.user,
                        reference=inbound
                    )
                    
                    item_to_delete.delete()

                messages.success(request, f"Inbound {inbound.nomor_inbound} berhasil diperbarui.")
                return redirect('inventory:inbound_detail', pk=inbound.pk)

        except Exception as e:
            messages.error(request, f"Gagal memperbarui inbound: {e}")
            # Fall through to render the form again with an error

    items = inbound.items.select_related('product', 'product__stock').all()
    context = {
        'inbound': inbound,
        'items': items,
    }
    return render(request, template_name, context)

@login_required
@permission_required('inventory.view_supplier', raise_exception=True)
def daftar_supplier(request):
    """Halaman daftar supplier"""
    return render(request, 'inventory/daftar_supplier.html')

@require_GET
@login_required
@permission_required('inventory.view_supplier', raise_exception=True)
def supplier_data(request):
    """Data supplier untuk DataTables"""
    from django.db.models import Q
    
    # DataTables parameters
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '')
    
    # Get all suppliers
    suppliers = Supplier.objects.all()
    
    # Apply search
    if search_value:
        suppliers = suppliers.filter(
            Q(nama_supplier__icontains=search_value) |
            Q(nomor_telepon__icontains=search_value) |
            Q(alamat__icontains=search_value) |
            Q(kota__icontains=search_value) |
            Q(brand__icontains=search_value)
        )
    
    # Apply column-specific search
    for i in range(6):  # 6 columns
        column_search = request.GET.get(f'columns[{i}][search][value]', '')
        if column_search:
            if i == 0:  # ID
                suppliers = suppliers.filter(id__icontains=column_search)
            elif i == 1:  # Nama Supplier
                suppliers = suppliers.filter(nama_supplier__icontains=column_search)
            elif i == 2:  # Nomor Telepon
                suppliers = suppliers.filter(nomor_telepon__icontains=column_search)
            elif i == 3:  # Alamat
                suppliers = suppliers.filter(alamat__icontains=column_search)
            elif i == 4:  # Kota
                suppliers = suppliers.filter(kota__icontains=column_search)
            elif i == 5:  # Brand
                suppliers = suppliers.filter(brand__icontains=column_search)
    
    # Get total count before pagination
    total_records = suppliers.count()
    
    # Apply ordering
    order_column = int(request.GET.get('order[0][column]', 1))
    order_dir = request.GET.get('order[0][dir]', 'asc')
    
    if order_dir == 'asc':
        if order_column == 0:
            suppliers = suppliers.order_by('id')
        elif order_column == 1:
            suppliers = suppliers.order_by('nama_supplier')
        elif order_column == 2:
            suppliers = suppliers.order_by('nomor_telepon')
        elif order_column == 3:
            suppliers = suppliers.order_by('alamat')
        elif order_column == 4:
            suppliers = suppliers.order_by('kota')
        elif order_column == 5:
            suppliers = suppliers.order_by('brand')
    else:
        if order_column == 0:
            suppliers = suppliers.order_by('-id')
        elif order_column == 1:
            suppliers = suppliers.order_by('-nama_supplier')
        elif order_column == 2:
            suppliers = suppliers.order_by('-nomor_telepon')
        elif order_column == 3:
            suppliers = suppliers.order_by('-alamat')
        elif order_column == 4:
            suppliers = suppliers.order_by('-kota')
        elif order_column == 5:
            suppliers = suppliers.order_by('-brand')
    
    # Apply pagination
    suppliers = suppliers[start:start + length]
    
    # Prepare data
    data = []
    for supplier in suppliers:
        data.append({
            'id': supplier.id,
            'nama_supplier': supplier.nama_supplier,
            'nomor_telepon': supplier.nomor_telepon or '',
            'alamat': supplier.alamat or '',
            'kota': supplier.kota or '',
            'brand': supplier.brand or '',
        })
    
    return JsonResponse({
        'draw': int(request.GET.get('draw', 1)),
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })

@require_POST
@csrf_exempt
@login_required
@permission_required('inventory.add_supplier', raise_exception=True)
def supplier_add(request):
    """Tambah supplier baru"""
    try:
        nama_supplier = request.POST.get('nama_supplier', '').strip()
        nomor_telepon = request.POST.get('nomor_telepon', '').strip()
        alamat = request.POST.get('alamat', '').strip()
        kota = request.POST.get('kota', '').strip()
        brand = request.POST.get('brand', '').strip()
        
        if not nama_supplier:
            return JsonResponse({'success': False, 'error': 'Nama supplier wajib diisi'})
        
        # Check if supplier already exists
        if Supplier.objects.filter(nama_supplier__iexact=nama_supplier).exists():
            return JsonResponse({'success': False, 'error': 'Supplier dengan nama tersebut sudah ada'})
        
        # Create new supplier
        supplier = Supplier.objects.create(
            nama_supplier=nama_supplier,
            nomor_telepon=nomor_telepon or None,
            alamat=alamat or None,
            kota=kota or None,
            brand=brand or None
        )
        
        return JsonResponse({'success': True, 'id': supplier.id})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_GET
@login_required
@permission_required('inventory.view_supplier', raise_exception=True)
def supplier_detail(request, pk):
    """Get supplier detail for editing"""
    try:
        supplier = get_object_or_404(Supplier, pk=pk)
        return JsonResponse({
            'id': supplier.id,
            'nama_supplier': supplier.nama_supplier,
            'nomor_telepon': supplier.nomor_telepon or '',
            'alamat': supplier.alamat or '',
            'kota': supplier.kota or '',
            'brand': supplier.brand or '',
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

@require_POST
@csrf_exempt
@login_required
@permission_required('inventory.change_supplier', raise_exception=True)
def supplier_edit(request, pk):
    """Edit supplier"""
    try:
        supplier = get_object_or_404(Supplier, pk=pk)
        
        nama_supplier = request.POST.get('nama_supplier', '').strip()
        nomor_telepon = request.POST.get('nomor_telepon', '').strip()
        alamat = request.POST.get('alamat', '').strip()
        kota = request.POST.get('kota', '').strip()
        brand = request.POST.get('brand', '').strip()
        
        if not nama_supplier:
            return JsonResponse({'success': False, 'error': 'Nama supplier wajib diisi'})
        
        # Check if supplier name already exists (excluding current supplier)
        if Supplier.objects.filter(nama_supplier__iexact=nama_supplier).exclude(id=pk).exists():
            return JsonResponse({'success': False, 'error': 'Supplier dengan nama tersebut sudah ada'})
        
        # Update supplier
        supplier.nama_supplier = nama_supplier
        supplier.nomor_telepon = nomor_telepon or None
        supplier.alamat = alamat or None
        supplier.kota = kota or None
        supplier.brand = brand or None
        supplier.save()
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_POST
@csrf_exempt
@login_required
@permission_required('inventory.delete_supplier', raise_exception=True)
def supplier_delete(request, pk):
    """Delete supplier"""
    try:
        supplier = get_object_or_404(Supplier, pk=pk)
        
        # Check if supplier is used in any inbound
        from .models import Inbound
        if Inbound.objects.filter(supplier=supplier).exists():
            return JsonResponse({
                'success': False, 
                'error': 'Supplier tidak dapat dihapus karena masih digunakan dalam inbound'
            })
        
        supplier.delete()
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_POST
def close_batch(request, batch_id):
    """
    Menutup batch, mengubah status, dan yang paling penting:
    Memindahkan `quantity_locked` menjadi pengurangan permanen dari `quantity`.
    Ini menandakan barang secara fisik sudah tidak ada di rak.
    """
    try:
        with transaction.atomic():
            batch = get_object_or_404(BatchList, pk=batch_id)
            if batch.status_batch != 'completed':
                messages.error(request, "Hanya batch yang berstatus 'completed' yang bisa ditutup.")
                return redirect('fullfilment:index')

            # 1. Ubah status batch menjadi 'closed'
            batch.status_batch = 'closed'
            batch.completed_at = timezone.now()
            batch.save(update_fields=['status_batch', 'completed_at'])

            # 2. Kurangi stok fisik berdasarkan item yang sudah di-pick
            batch_items = BatchItem.objects.filter(batchlist=batch, is_picked=True)
            for batch_item in batch_items:
                try:
                    stock = Stock.objects.select_for_update().get(product=batch_item.product)
                    
                    qty_awal_fisik = stock.quantity
                    qty_awal_locked = stock.quantity_locked

                    stock.quantity -= batch_item.picked_quantity
                    stock.quantity_locked -= batch_item.picked_quantity
                    stock.save(update_fields=['quantity', 'quantity_locked'])

                    # 3. Buat entri StockCardEntry
                    StockCardEntry.objects.create(
                        product=batch_item.product,
                        tipe_pergerakan='close_batch',
                        qty=-batch_item.picked_quantity,
                        qty_awal=qty_awal_fisik,
                        qty_akhir=stock.quantity,
                        notes=f'Stok dikunci untuk Batch {batch.nama_batch}',
                        user=request.user,
                        reference=batch_item
                    )
                except Stock.DoesNotExist:
                    messages.warning(request, f"Stok untuk produk {batch_item.product.sku} tidak ditemukan saat Close Batch.")
                    continue
            
            messages.success(request, f"Batch '{batch.nama_batch}' berhasil ditutup (Closed).")
            return redirect('fullfilment:index')
    
    except Exception as e:
        logging.error(f"Error closing batch {batch_id}: {e}")
        messages.error(request, f"Terjadi kesalahan saat menutup batch: {e}")
        return redirect('fullfilment:index')

@require_POST
def reopen_batch(request, batch_id):
    try:
        with transaction.atomic():
            batch = get_object_or_404(BatchList, pk=batch_id)
            if batch.status_batch != 'closed':
                messages.error(request, "Hanya batch yang berstatus 'closed' yang bisa di Re-Open.")
                return redirect('fullfilment:index')

            batch.status_batch = 'completed'
            batch.save(update_fields=['status_batch'])

            batch_items = BatchItem.objects.filter(batchlist=batch, is_picked=True)
            for batch_item in batch_items:
                try:
                    stock = Stock.objects.select_for_update().get(product=batch_item.product)
                    
                    qty_awal_locked = stock.quantity_locked

                    stock.quantity_locked += batch_item.picked_quantity
                    stock.save(update_fields=['quantity_locked'])
                    
                    StockCardEntry.objects.create(
                        product=batch_item.product,
                        tipe_pergerakan='reopen_batch',
                        qty=batch_item.picked_quantity,
                        qty_awal=qty_awal_locked,
                        qty_akhir=stock.quantity_locked,
                        notes=f'Stok dikembalikan dari Re-Open Batch {batch.nama_batch}',
                        user=request.user,
                        reference=batch_item
                    )
                except Stock.DoesNotExist:
                    messages.warning(request, f"Stok untuk produk {batch_item.product.sku} tidak ditemukan saat Re-Open Batch.")
                    continue

            messages.success(request, f"Batch '{batch.nama_batch}' berhasil dibuka kembali (Re-Opened).")
            return redirect('fullfilment:index')
    
    except Exception as e:
        logging.error(f"Error reopening batch {batch_id}: {e}")
        messages.error(request, f"Terjadi kesalahan saat membuka kembali batch: {e}")
        return redirect('fullfilment:index')

@login_required
@permission_required('inventory.view_putaway', raise_exception=True)
def putaway_list(request):
    """
    Menampilkan daftar produk yang perlu di-putaway (quantity_putaway > 0)
    """
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    
    template_name = 'inventory/putaway_mobile.html' if is_mobile else 'inventory/putaway.html'
    
    # Ambil produk yang memiliki quantity_putaway > 0
    stock_qs = Stock.objects.filter(product=OuterRef('pk'))
    # Hanya ambil slotting log yang belum di-putaway (putaway_by is null)
    slotting_qs = PutawaySlottingLog.objects.filter(
        product=OuterRef('pk'),
        putaway_by__isnull=True  # Hanya slotting yang belum di-putaway
    ).order_by('-created_at')
    
    # Ambil timestamp terbaru dari InventoryRakStockLog untuk putaway_masuk
    from .models import InventoryRakStockLog
    putaway_log_qs = InventoryRakStockLog.objects.filter(
        produk=OuterRef('pk'),
        tipe_pergerakan='putaway_masuk'
    ).order_by('-waktu_buat')

    products = Product.objects.annotate(
        quantity=Coalesce(Subquery(stock_qs.values('quantity')[:1]), Value(0)),
        quantity_locked=Coalesce(Subquery(stock_qs.values('quantity_locked')[:1]), Value(0)),
        quantity_putaway=Coalesce(Subquery(stock_qs.values('quantity_putaway')[:1]), Value(0)),
        quantity_ready_v2=F('quantity') - F('quantity_locked'),
        # Data slotting dari PutawaySlottingLog (hanya yang belum di-putaway)
        suggested_rak_id=Coalesce(Subquery(slotting_qs.values('suggested_rak')[:1]), Value(None)),
        putaway_by_id=Coalesce(Subquery(slotting_qs.values('putaway_by')[:1]), Value(None)),
        putaway_time=Coalesce(Subquery(slotting_qs.values('putaway_time')[:1]), Value(None)),
        slotting_created_at=Coalesce(Subquery(slotting_qs.values('created_at')[:1]), Value(None)),
        # Timestamp terbaru dari InventoryRakStockLog untuk putaway
        putaway_created_at=Coalesce(Subquery(putaway_log_qs.values('waktu_buat')[:1]), Value(None))
    ).filter(
        quantity_putaway__gt=0  # Hanya produk dengan quantity_putaway > 0
    ).values(
        'id', 'sku', 'barcode', 'nama_produk', 'variant_produk', 'brand',
        'lebar_cm', 'panjang_cm', 'tinggi_cm',
        'quantity', 'quantity_locked', 'quantity_putaway', 'quantity_ready_v2',
        'suggested_rak_id', 'putaway_by_id', 'putaway_time', 'slotting_created_at', 'putaway_created_at'
    ).order_by('-putaway_created_at', '-quantity_putaway', 'nama_produk')  # Urutkan berdasarkan yang terbaru
    
    # Ambil sesi transfer yang siap putaway
    from .models import RakTransferSession
    transfer_sessions = RakTransferSession.objects.select_related('rak_asal', 'rak_tujuan').filter(
        status='ready_for_putaway', mode='transfer_putaway'
    ).order_by('-tanggal_transfer').values(
        'id', 'session_code', 'rak_asal__kode_rak', 'rak_tujuan__kode_rak', 'tanggal_transfer'
    )

    # Get unique brands for filter
    brands = list(set(p['brand'] for p in products if p['brand']))
    brands.sort()
    
    # Enrich products with rak and user data
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    enriched_products = []
    for product in products:
        # Get full product object for photo
        try:
            product_obj = Product.objects.get(id=product['id'])
            photo_url = product_obj.photo.url if product_obj.photo else ''
        except Product.DoesNotExist:
            photo_url = ''
        
        # Get rak data
        suggested_rak = None
        if product['suggested_rak_id']:
            try:
                suggested_rak = Rak.objects.get(id=product['suggested_rak_id'])
            except Rak.DoesNotExist:
                pass
        
        # Get user data
        putaway_by = None
        if product['putaway_by_id']:
            try:
                putaway_by = User.objects.get(id=product['putaway_by_id'])
            except User.DoesNotExist:
                pass
        
        enriched_products.append({
            **product,
            'photo': photo_url,  # Add photo URL as string
            'suggested_rak': suggested_rak,
            'putaway_by': putaway_by
        })
    
    context = {
        'putaway_list': enriched_products,
        'total_items': len(enriched_products),
        'total_quantity': sum(p['quantity_putaway'] for p in enriched_products),
        'transfer_sessions': list(transfer_sessions),
        'brands': brands
    }
    return render(request, template_name, context)

# ===== PUTAWAY SCAN VIEWS =====

@login_required
@permission_required('inventory.change_putaway', raise_exception=True)
def putaway_scan(request):
    """
    Interface untuk scan putaway - mobile friendly
    """
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    
    template_name = 'inventory/putaway_scan_mobile.html' if is_mobile else 'inventory/putaway_scan.html'
    
    context = {
        'is_mobile': is_mobile,
    }
    return render(request, template_name, context)

@require_GET
@login_required
@permission_required('inventory.change_putaway', raise_exception=True)
def putaway_scan_rak(request):
    """
    AJAX endpoint untuk scan rak: bisa pakai kode_rak atau barcode_rak
    """
    kode_rak = request.GET.get('kode_rak', '').strip()
    barcode_rak = request.GET.get('barcode_rak', '').strip()

    if not kode_rak and not barcode_rak:
        return JsonResponse({'error': 'Kode rak atau barcode rak harus diisi'}, status=400)

    try:
        if barcode_rak:
            # Case-insensitive search for barcode_rak
            rak = Rak.objects.filter(barcode_rak__iexact=barcode_rak).first()
        else:
            # Case-insensitive search for kode_rak
            rak = Rak.objects.filter(kode_rak__iexact=kode_rak).first()
        
        if not rak:
            return JsonResponse({'error': 'Rak tidak ditemukan'}, status=404)
        return JsonResponse({
            'success': True,
            'rak': {
                'id': rak.id,
                'kode_rak': rak.kode_rak,
                'nama_rak': rak.nama_rak,
                'lokasi': rak.lokasi or '',
                'barcode_rak': rak.barcode_rak or ''
            }
        })
    except Rak.DoesNotExist:
        return JsonResponse({'error': 'Rak tidak ditemukan'}, status=404)

@require_GET
@login_required
@permission_required('inventory.change_putaway', raise_exception=True)
def putaway_scan_product(request):
    """
    AJAX endpoint untuk scan barcode produk
    """
    barcode = request.GET.get('barcode', '').strip()
    rak_id = request.GET.get('rak_id')
    
    if not barcode:
        return JsonResponse({'error': 'Barcode tidak boleh kosong'}, status=400)
    
    if not rak_id:
        return JsonResponse({'error': 'Rak ID tidak boleh kosong'}, status=400)
    
    try:
        product = Product.objects.get(barcode=barcode)
        stock = Stock.objects.filter(product=product).first()
        
        if not stock or stock.quantity_putaway <= 0:
            return JsonResponse({'error': f'Produk {product.nama_produk} tidak memiliki stok yang perlu di-putaway (Qty Putaway: 0).'}, status=400)
        
        # Get suggested_rak from PutawaySlottingLog
        from .models import PutawaySlottingLog
        slotting_log = PutawaySlottingLog.objects.filter(
            product=product,
            putaway_by__isnull=True  # Belum di-scan putaway
        ).order_by('-created_at').first()
        
        suggested_rak = None
        if slotting_log and slotting_log.suggested_rak:
            suggested_rak = slotting_log.suggested_rak.kode_rak
        
        return JsonResponse({
            'success': True,
            'product': {
                'id': product.id,
                'sku': product.sku,
                'barcode': product.barcode,
                'nama_produk': product.nama_produk,
                'variant_produk': product.variant_produk or '',
                'brand': product.brand or '',
                'quantity_putaway': stock.quantity_putaway,
                'quantity_ready': stock.quantity - stock.quantity_locked,
                'suggested_rak': suggested_rak,  # Add suggested_rak data
                'photo': product.photo.url if product.photo else '',
            }
        })
    except Product.DoesNotExist:
        return JsonResponse({'error': f'Produk dengan barcode "{barcode}" tidak ditemukan'}, status=404)

@require_POST
@csrf_exempt
@login_required
@permission_required('inventory.change_putaway', raise_exception=True)
def putaway_save(request):
    """
    Menyimpan data putaway menggunakan centralized PutawayService
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Putaway save request received: {request.body}")
        data = json.loads(request.body)
        rak_id = data.get('rak_id')
        items_to_putaway = data.get('items', [])
        is_transfer_putaway = bool(data.get('is_transfer_putaway'))
        transfer_session_id = data.get('transfer_session_id')

        logger.info(f"Parsed data: rak_id={rak_id}, items_count={len(items_to_putaway)}, is_transfer_putaway={is_transfer_putaway}")

        if not rak_id:
            return JsonResponse({'success': False, 'error': 'Rak tujuan wajib diisi.'}, status=400)

        # Jika mode transfer putaway dan items kosong, ambil dari RakTransferSession
        if is_transfer_putaway and transfer_session_id and not items_to_putaway:
            try:
                from .models import RakTransferSession
                transfer_session = RakTransferSession.objects.select_related('rak_tujuan').get(id=transfer_session_id)
                items_to_putaway = [
                    { 'product_id': it.product_id, 'quantity': it.qty_transfer }
                    for it in transfer_session.items.all()
                    if it.qty_transfer and it.qty_transfer > 0
                ]
                if str(transfer_session.rak_tujuan_id) != str(rak_id):
                    return JsonResponse({'success': False, 'error': 'Rak tujuan tidak sesuai dengan sesi transfer.'}, status=400)
            except Exception as e:
                logger.error(f"Error getting transfer session: {e}")
                return JsonResponse({'success': False, 'error': 'Sesi transfer tidak ditemukan.'}, status=404)

        if not items_to_putaway:
            return JsonResponse({'success': False, 'error': 'Daftar produk kosong.'}, status=400)

        # Use centralized PutawayService
        from .putaway import PutawayService
        
        putaway_type = 'transfer' if is_transfer_putaway else 'regular'
        kwargs = {'session_id': transfer_session_id} if transfer_session_id else {}
        
        result = PutawayService.process_putaway(
            request=request,
            rak_id=rak_id,
            items_to_putaway=items_to_putaway,
            putaway_type=putaway_type,
            **kwargs
        )
        
        if result['success']:
            return JsonResponse({'success': True, 'message': result['message']})
        else:
            return JsonResponse({'success': False, 'error': result['error']}, status=400)

    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error: {e}")
        return JsonResponse({'success': False, 'error': 'Format data JSON tidak valid.'}, status=400)
    except Exception as e:
        logger.error(f"Unexpected error in putaway_save: {e}")
        return JsonResponse({'success': False, 'error': 'Terjadi kesalahan sistem.'}, status=500)

# ===== PICKING PROCESS VIEWS =====

def picking_scan_view(request):
    """
    Menampilkan interface untuk proses picking barang dari rak.
    """
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    
    template_name = 'inventory/pickingrak_mobile.html' if is_mobile else 'inventory/pickingrak.html'
    
    context = {
        'is_mobile': is_mobile,
    }
    return render(request, template_name, context)

@require_GET
def api_scan_rak_for_picking(request):
    """
    AJAX endpoint untuk scan kode rak di proses picking.
    """
    kode_rak = request.GET.get('kode_rak', '').strip()
    rak_id_param = request.GET.get('rak_id', '').strip()

    if not kode_rak and not rak_id_param:
        return JsonResponse({'success': False, 'error': 'Kode rak atau ID rak tidak boleh kosong.'}, status=400)
    
    try:
        if rak_id_param:
            rak = Rak.objects.get(id=rak_id_param)
        else:
            rak = Rak.objects.get(kode_rak=kode_rak)
        
        return JsonResponse({
            'success': True,
            'rak': {
                'id': rak.id,
                'kode_rak': rak.kode_rak,
                'nama_rak': rak.nama_rak,
                'lokasi': rak.lokasi or '',
            }
        })
    except Rak.DoesNotExist:
        return JsonResponse({'success': False, 'error': f'Rak dengan kode/ID "{kode_rak or rak_id_param}" tidak ditemukan.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Terjadi kesalahan: {str(e)}'}, status=500)

@require_GET
def api_scan_product_for_picking(request):
    """
    AJAX endpoint untuk scan barcode produk di proses picking.
    Mengembalikan info produk dan stoknya di rak yang discan.
    """
    barcode = request.GET.get('barcode', '').strip()
    rak_id = request.GET.get('rak_id')
    
    if not barcode:
        return JsonResponse({'success': False, 'error': 'Barcode produk tidak boleh kosong.'}, status=400)
    if not rak_id:
        return JsonResponse({'success': False, 'error': 'Rak asal belum dipilih.'}, status=400)
    
    try:
        rak = get_object_or_404(Rak, id=rak_id)

        # Cari produk berdasarkan barcode utama atau barcode tambahan
        product = Product.objects.filter(barcode=barcode).first()
        if not product:
            extra_barcode_entry = ProductExtraBarcode.objects.filter(barcode=barcode).select_related('product').first()
            if extra_barcode_entry:
                product = extra_barcode_entry.product

        if not product:
            return JsonResponse({'success': False, 'error': f'Produk dengan barcode "{barcode}" tidak ditemukan.'}, status=404)

        # Dapatkan stok produk di rak yang discan
        inventory_rak_stock = InventoryRakStock.objects.filter(product=product, rak=rak).first()
        stock_in_rak = inventory_rak_stock.quantity if inventory_rak_stock else 0

        if stock_in_rak <= 0:
            return JsonResponse({'success': False, 'error': f'Produk {product.nama_produk} ({product.sku}) tidak ada stok yang perlu di-pick (Qty Putaway: 0).'}, status=400)

        return JsonResponse({
            'success': True,
            'product': {
                'id': product.id,
                'sku': product.sku,
                'barcode': product.barcode,
                'nama_produk': product.nama_produk,
                'variant_produk': product.variant_produk or '',
                'brand': product.brand or '',
            },
            'stock_in_rak': stock_in_rak, # Stok yang tersedia di rak ini
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Terjadi kesalahan: {str(e)}'}, status=500)

@require_POST
@csrf_exempt
def api_save_picking_transaction(request):
    """
    Menyimpan transaksi picking untuk multiple produk dari satu rak:
    - Mengurangi Stock.quantity
    - Mengurangi InventoryRakStock.quantity di rak asal
    - Membuat log di InventoryRakStockLog
    - Membuat log di StockCardEntry
    """
    try:
        data = json.loads(request.body)
        rak_id = data.get('rak_id')
        items_to_pick = data.get('items', [])

        if not rak_id or not items_to_pick:
            return JsonResponse({'success': False, 'error': 'Data tidak lengkap atau daftar produk kosong.'}, status=400)

        with transaction.atomic():
            rak_asal = get_object_or_404(Rak, id=rak_id)
            
            for item_data in items_to_pick:
                product_id = item_data.get('product_id')
                qty_picked = int(item_data.get('qty_picked'))

                if not product_id or qty_picked <= 0:
                    raise ValueError('Data produk atau kuantitas yang diambil tidak valid.')

                product = get_object_or_404(Product, id=product_id)
                stock_product = get_object_or_404(Stock, product=product)

                # 1. Cek & Kurangi stok di InventoryRakStock (rak asal)
                inventory_rak_stock = get_object_or_404(InventoryRakStock, product=product, rak=rak_asal)
                if inventory_rak_stock.quantity < qty_picked:
                    raise ValueError(f'Stok produk {product.sku} tidak cukup di Rak {rak_asal.kode_rak}. Tersedia: {inventory_rak_stock.quantity}.')
                
                inventory_rak_stock.quantity -= qty_picked
                inventory_rak_stock.save()

                # 2. Cek & Kurangi stok total di Stock
                if stock_product.quantity < qty_picked:
                    raise ValueError(f'Stok produk {product.sku} di gudang tidak cukup ({stock_product.quantity}).')
                
                qty_awal_stock_card = stock_product.quantity # Qty sebelum dikurangi
                stock_product.quantity -= qty_picked
                stock_product.save()
                qty_akhir_stock_card = stock_product.quantity # Qty setelah dikurangi

                # 3. Buat log di InventoryRakStockLog
                InventoryRakStockLog.objects.create(
                    produk=product,
                    rak=rak_asal,
                    tipe_pergerakan='picking_keluar',
                    qty=-qty_picked, # Negatif karena keluar
                    qty_awal=inventory_rak_stock.quantity + qty_picked, # Saldo awal rak
                    qty_akhir=inventory_rak_stock.quantity, # Saldo akhir rak
                    user=request.user,
                    waktu_buat=timezone.now(),
                    catatan=f'Picking {qty_picked} unit {product.nama_produk} dari rak {rak_asal.kode_rak}',
                )

                # 4. Buat entri di StockCardEntry
                # Perlu import ContentType jika belum ada di atas
                product_content_type = ContentType.objects.get_for_model(product)
                StockCardEntry.objects.create(
                    product=product,
                    qty=-qty_picked, # Negatif karena stok keluar
                    tipe_pergerakan='outbound', # Atau 'picking_keluar' jika ingin spesifik
                    user=request.user,
                    waktu=timezone.now(),
                    notes=f'Picking {qty_picked} unit {product.nama_produk} dari Rak {rak_asal.kode_rak}.',
                    content_type=product_content_type,
                    object_id=product.id,
                    qty_awal=qty_awal_stock_card,
                    qty_akhir=qty_akhir_stock_card,
                )
            
            return JsonResponse({'success': True, 'message': 'Transaksi picking berhasil disimpan!'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Format data JSON tidak valid.'}, status=400)
    except ValueError as ve:
        return JsonResponse({'success': False, 'error': f'Validasi data gagal: {str(ve)}'}, status=400)
    except (Rak.DoesNotExist, Product.DoesNotExist, Stock.DoesNotExist, InventoryRakStock.DoesNotExist) as e:
        return JsonResponse({'success': False, 'error': f'Item tidak ditemukan: {str(e)}'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Terjadi kesalahan server: {str(e)}'}, status=500)

def rakopname_session_list(request):
    """
    Menampilkan daftar sesi opname rak
    """
    sessions = RakOpnameSession.objects.all().order_by('-tanggal_mulai')
    
    context = {
        'sessions': sessions,
    }
    return render(request, 'inventory/rakopname_session_list.html', context)

@login_required
@permission_required('inventory.view_putaway', raise_exception=True)
def putaway_history(request):
    """
    Menampilkan history putaway dari StockCardEntry dan InventoryRakStockLog
    """
    import re
    
    # Mobile detection
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    
    template_name = 'inventory/putaway_history_mobile.html' if is_mobile else 'inventory/putaway_history.html'
    
    from django.utils import timezone
    
    context = {
        'is_mobile': is_mobile,
        'timestamp': int(timezone.now().timestamp()),  # For cache busting
    }
    
    return render(request, template_name, context)

@require_GET
@login_required
@permission_required('inventory.view_putaway', raise_exception=True)
def putaway_history_data(request):
    """
    API untuk DataTables - history putaway dari 2 tabel
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        logger.info("Starting putaway_history_data request")
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 25))
        search_value = request.GET.get('search[value]', '')
        
        # Filter tanggal
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # Filter tipe (putaway biasa atau transfer putaway)
        tipe_filter = request.GET.get('tipe_filter', 'all')
        
        # Ambil data dari StockCardEntry (putaway)
        # Catatan: StockCardEntry tidak memiliki tipe 'putaway', jadi kita ambil semua entry yang terkait putaway
        stock_card_entries = StockCardEntry.objects.filter(
            notes__icontains='putaway'
        ).select_related('product', 'user').order_by('-waktu')
        
        logger.info(f"Found {stock_card_entries.count()} StockCardEntry records")
        
        # Ambil data dari InventoryRakStockLog (putaway masuk)
        rak_stock_logs = InventoryRakStockLog.objects.filter(
            tipe_pergerakan='putaway_masuk'
        ).select_related('produk', 'rak', 'user').order_by('-waktu_buat')
        
        logger.info(f"Found {rak_stock_logs.count()} InventoryRakStockLog records")
        
        # Filter berdasarkan tanggal
        if date_from:
            stock_card_entries = stock_card_entries.filter(waktu__date__gte=date_from)
            rak_stock_logs = rak_stock_logs.filter(waktu_buat__date__gte=date_from)
        
        if date_to:
            stock_card_entries = stock_card_entries.filter(waktu__date__lte=date_to)
            rak_stock_logs = rak_stock_logs.filter(waktu_buat__date__lte=date_to)
        
        # Filter berdasarkan tipe
        if tipe_filter == 'transfer':
            stock_card_entries = stock_card_entries.filter(notes__icontains='Transfer')
            rak_stock_logs = rak_stock_logs.filter(catatan__icontains='Transfer')
        elif tipe_filter == 'regular':
            stock_card_entries = stock_card_entries.exclude(notes__icontains='Transfer')
            rak_stock_logs = rak_stock_logs.exclude(catatan__icontains='Transfer')
        
        # Search functionality
        if search_value:
            stock_card_entries = stock_card_entries.filter(
                Q(product__sku__icontains=search_value) |
                Q(product__nama_produk__icontains=search_value) |
                Q(product__barcode__icontains=search_value) |
                Q(notes__icontains=search_value) |
                Q(user__username__icontains=search_value)
            )
            rak_stock_logs = rak_stock_logs.filter(
                Q(produk__sku__icontains=search_value) |
                Q(produk__nama_produk__icontains=search_value) |
                Q(produk__barcode__icontains=search_value) |
                Q(rak__kode_rak__icontains=search_value) |
                Q(catatan__icontains=search_value) |
                Q(user__username__icontains=search_value)
            )
        
        # Gabungkan data dari kedua tabel
        combined_data = []
        
        # Data dari StockCardEntry
        try:
            for entry in stock_card_entries:
                combined_data.append({
                    'timestamp': entry.waktu,  # Field yang benar adalah 'waktu'
                    'source': 'StockCardEntry',
                    'product_sku': entry.product.sku if entry.product else '-',
                    'product_name': entry.product.nama_produk if entry.product else '-',
                    'product_barcode': entry.product.barcode if entry.product else '-',
                    'product_variant': entry.product.variant_produk if entry.product else '-',
                    'quantity': entry.qty,  # Field yang benar adalah 'qty'
                    'rak_kode': '-',  # StockCardEntry tidak menyimpan rak
                    'rak_nama': '-',
                    'user': entry.user.username if entry.user else 'System',
                    'notes': entry.notes,
                    'tipe': 'Transfer Putaway' if 'Transfer' in entry.notes else 'Regular Putaway'
                })
        except Exception as e:
            logger.error(f"Error processing StockCardEntry: {str(e)}", exc_info=True)
            raise
        
        # Data dari InventoryRakStockLog
        try:
            for log in rak_stock_logs:
                combined_data.append({
                    'timestamp': log.waktu_buat,
                    'source': 'InventoryRakStockLog',
                    'product_sku': log.produk.sku if log.produk else '-',
                    'product_name': log.produk.nama_produk if log.produk else '-',
                    'product_barcode': log.produk.barcode if log.produk else '-',
                    'product_variant': log.produk.variant_produk if log.produk else '-',
                    'quantity': log.qty,
                    'rak_kode': log.rak.kode_rak if log.rak else '-',
                    'rak_nama': log.rak.nama_rak if log.rak else '-',
                    'user': log.user.username if log.user else 'System',
                    'notes': log.catatan,
                    'tipe': 'Transfer Putaway' if 'Transfer' in log.catatan else 'Regular Putaway'
                })
        except Exception as e:
            logger.error(f"Error processing InventoryRakStockLog: {str(e)}", exc_info=True)
            raise
        
        # Sort berdasarkan timestamp (terbaru dulu)
        combined_data.sort(key=lambda x: x['timestamp'], reverse=True)
        
        logger.info(f"Total combined data: {len(combined_data)}")
        
        # Jika tidak ada data, buat data dummy untuk testing
        if len(combined_data) == 0:
            logger.info("No data found, creating dummy data for testing")
            from datetime import datetime
            combined_data = [{
                'timestamp': datetime.now(),
                'source': 'Test',
                'product_sku': 'TEST-SKU',
                'product_name': 'Test Product',
                'product_barcode': '123456789',
                'quantity': 10,
                'rak_kode': 'TEST-RAK',
                'rak_nama': 'Test Rak',
                'user': 'Test User',
                'notes': 'Test data',
                'tipe': 'Regular Putaway'
            }]
        
        # Pagination
        total_records = len(combined_data)
        filtered_records = len(combined_data)
        
        # Apply pagination
        paginated_data = combined_data[start:start + length]
        
        logger.info(f"Paginated data: {len(paginated_data)} records")
        
        # Format data untuk DataTables
        data = []
        try:
            for item in paginated_data:
                # Format timestamp
                if isinstance(item['timestamp'], str):
                    timestamp_str = item['timestamp']
                else:
                    timestamp_str = item['timestamp'].strftime('%d/%m/%Y %H:%M:%S')
                
                data.append({
                    'timestamp': timestamp_str,
                    'source': item['source'],
                    'product_sku': item['product_sku'],
                    'product_name': item['product_name'],
                    'product_barcode': item['product_barcode'],
                    'product_variant': item['product_variant'],
                    'quantity': item['quantity'],
                    'rak_kode': item['rak_kode'],
                    'rak_nama': item['rak_nama'],
                    'user': item['user'],
                    'notes': item['notes'],
                    'tipe': item['tipe']
                })
        except Exception as e:
            logger.error(f"Error formatting data: {str(e)}", exc_info=True)
            # Return empty data if formatting fails
            data = []
        
        logger.info(f"Returning {len(data)} records for putaway history")
        
        # Debug: Log sample data
        if data:
            logger.info(f"Sample data: {data[0]}")
        
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })
        
    except Exception as e:
        logger.error(f"Error in putaway_history_data: {str(e)}", exc_info=True)
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': str(e)
        }, status=500)

@login_required
@require_GET
def putaway_list_data_api(request):
    """
    API endpoint untuk mendapatkan data putaway list dalam format JSON
    """
    try:
        # Ambil produk yang memiliki quantity_putaway > 0
        stock_qs = Stock.objects.filter(product=OuterRef('pk'))
        # Hanya ambil slotting log yang belum di-putaway (putaway_by is null)
        slotting_qs = PutawaySlottingLog.objects.filter(
            product=OuterRef('pk'),
            putaway_by__isnull=True  # Hanya slotting yang belum di-putaway
        ).order_by('-created_at')
        
        # Ambil timestamp terbaru dari InventoryRakStockLog untuk putaway_masuk
        from .models import InventoryRakStockLog
        putaway_log_qs = InventoryRakStockLog.objects.filter(
            produk=OuterRef('pk'),
            tipe_pergerakan='putaway_masuk'
        ).order_by('-waktu_buat')

        products = Product.objects.annotate(
            quantity=Coalesce(Subquery(stock_qs.values('quantity')[:1]), Value(0)),
            quantity_locked=Coalesce(Subquery(stock_qs.values('quantity_locked')[:1]), Value(0)),
            quantity_putaway=Coalesce(Subquery(stock_qs.values('quantity_putaway')[:1]), Value(0)),
            quantity_ready_v2=F('quantity') - F('quantity_locked'),
            slotting_created_at=Coalesce(Subquery(slotting_qs.values('created_at')[:1]), Value(None)),
            # Timestamp terbaru dari InventoryRakStockLog untuk putaway
            putaway_created_at=Coalesce(Subquery(putaway_log_qs.values('waktu_buat')[:1]), Value(None))
        ).filter(
            quantity_putaway__gt=0  # Hanya produk dengan quantity_putaway > 0
        ).order_by('-putaway_created_at', '-quantity_putaway', 'nama_produk')
        
        # Format data untuk response
        putaway_items = []
        for product in products:
            photo_url = product.photo.url if product.photo else ''
            
            putaway_items.append({
                'sku': product.sku,
                'barcode': product.barcode or '',
                'nama_produk': product.nama_produk,
                'variant_produk': product.variant_produk or '',
                'brand': product.brand or '',
                'quantity_putaway': product.quantity_putaway,
                'photo': photo_url
            })
        
        return JsonResponse({
            'success': True,
            'data': putaway_items,
            'total_items': len(putaway_items),
            'total_quantity': sum(p['quantity_putaway'] for p in putaway_items)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def get_product_id_by_sku(request):
    """
    API endpoint untuk mendapatkan product_id berdasarkan SKU
    """
    try:
        sku = request.GET.get('sku')
        if not sku:
            return JsonResponse({
                'success': False,
                'error': 'SKU parameter is required'
            }, status=400)
        
        try:
            product = Product.objects.get(sku=sku)
            return JsonResponse({
                'success': True,
                'product_id': product.id,
                'sku': product.sku
            })
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'Product with SKU {sku} not found'
            }, status=404)
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_GET
def rak_items_data(request, rak_id):
    """
    AJAX endpoint untuk mendapatkan data item yang ada di rak tertentu
    """
    try:
        rak = get_object_or_404(Rak, id=rak_id)
        
        # Ambil semua item yang ada di rak ini (quantity > 0)
        rak_stocks = InventoryRakStock.objects.filter(
            rak=rak,
            quantity__gt=0
        ).select_related('product').order_by('product__nama_produk')
        
        items = []
        for stock in rak_stocks:
            # Get photo URL
            photo_url = ''
            if stock.product.photo:
                photo_url = stock.product.photo.url
            
            items.append({
                'sku': stock.product.sku,
                'nama_produk': stock.product.nama_produk,
                'variant_produk': stock.product.variant_produk or '',
                'brand': stock.product.brand or '',
                'quantity': stock.quantity,
                'barcode': stock.product.barcode or '',
                'photo': photo_url
            })
        
        return JsonResponse({
            'success': True,
            'rak': {
                'id': rak.id,
                'kode_rak': rak.kode_rak,
                'nama_rak': rak.nama_rak,
                'lokasi': rak.lokasi or ''
            },
            'items': items,
            'total_items': len(items),
            'total_quantity': sum(item['quantity'] for item in items)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Terjadi kesalahan: {str(e)}'
        }, status=500)

# ===== SLOTTING PUTAWAY VIEWS =====

@require_GET
def slotting_options(request):
    """
    AJAX endpoint untuk mendapatkan opsi rak yang sesuai untuk slotting
    """
    sku = request.GET.get('sku')
    quantity = request.GET.get('quantity')
    
    if not sku:
        return JsonResponse({'error': 'SKU tidak boleh kosong'}, status=400)
    
    try:
        product = Product.objects.get(sku=sku)
        
        # Convert quantity to int if provided
        quantity_int = None
        if quantity:
            try:
                quantity_int = int(quantity)
            except ValueError:
                return JsonResponse({'error': 'Quantity harus berupa angka'}, status=400)
        
        # Gunakan centralized SlottingService untuk mendapatkan opsi rak
        from .putaway import SlottingService
        
        rak_options_result = SlottingService.get_rak_options(product, quantity_int)
        
        if rak_options_result['success']:
            rak_options = rak_options_result['rak_options']
        else:
            rak_options = []
        
        # Get product photo URL
        photo_url = ''
        try:
            if product.photo:
                photo_url = product.photo.url
        except ValueError:
            photo_url = ''  # Handle missing file safely
        
        return JsonResponse({
            'success': True,
            'product': {
                'sku': product.sku,
                'nama_produk': product.nama_produk,
                'variant_produk': product.variant_produk or '',
                'brand': product.brand or '',
                'barcode': product.barcode or '',
                'photo_url': photo_url,
                'panjang_cm': product.panjang_cm,
                'lebar_cm': product.lebar_cm,
                'tinggi_cm': product.tinggi_cm,
                'berat_gram': product.berat_gram,
            },
            'rak_options': rak_options
        })
        
    except Product.DoesNotExist:
        return JsonResponse({'error': f'Produk dengan SKU "{sku}" tidak ditemukan'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Terjadi kesalahan: {str(e)}'}, status=500)


@require_POST
@csrf_exempt
def update_slotting(request):
    """
    AJAX endpoint untuk menyimpan hasil slotting (rak_tujuan dan putaway_by)
    HANYA menentukan lokasi rak_tujuan, TIDAK melakukan putaway
    """
    try:
        data = json.loads(request.body)
        sku = data.get('sku')
        rak_id = data.get('rak_id')
        notes = data.get('notes', '')
        
        if not sku:
            return JsonResponse({'success': False, 'error': 'SKU tidak boleh kosong'}, status=400)
        
        if not rak_id:
            return JsonResponse({'success': False, 'error': 'Rak tujuan tidak boleh kosong'}, status=400)
        
        with transaction.atomic():
            # Ambil product dan rak
            product = get_object_or_404(Product, sku=sku)
            rak = get_object_or_404(Rak, id=rak_id)
            user = request.user
            
            # Validasi slotting menggunakan centralized SlottingService
            stock = get_object_or_404(Stock, product=product)
            from .putaway import SlottingService
            
            validation = SlottingService.validate_slotting(product, rak, stock.quantity_putaway)
            
            if not validation['valid']:
                return JsonResponse({'success': False, 'error': validation['message']}, status=400)
            
            # Eksekusi slotting
            result = SlottingService.execute_slotting(
                product=product,
                rak=rak,
                user=user,
                quantity=stock.quantity_putaway,
                notes=notes,
                is_auto=False
            )
            
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'message': f'Slotting berhasil! {product.sku} → {rak.kode_rak} (Quantity putaway: {stock.quantity_putaway} pcs)',
                    'slotting_id': result['slotting_log'].id
                })
            else:
                return JsonResponse({'success': False, 'error': result['error']}, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Data JSON tidak valid'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Terjadi kesalahan: {str(e)}'}, status=500)


def slotting_history(request):
    """
    Menampilkan history slotting putaway
    """
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    
    template_name = 'inventory/slotting_history_mobile.html' if is_mobile else 'inventory/slotting_history.html'
    
    context = {
        'is_mobile': is_mobile,
    }
    
    return render(request, template_name, context)


@require_GET
def slotting_history_data(request):
    """
    API untuk DataTables - history slotting putaway
    """
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 25))
        search_value = request.GET.get('search[value]', '')
        
        # Filter tanggal
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # Query slotting log
        queryset = PutawaySlottingLog.objects.select_related(
            'product', 'suggested_rak', 'putaway_by'
        ).order_by('-created_at')
        
        # Apply search
        if search_value:
            queryset = queryset.filter(
                Q(product__sku__icontains=search_value) |
                Q(product__nama_produk__icontains=search_value) |
                Q(suggested_rak__kode_rak__icontains=search_value) |
                Q(putaway_by__username__icontains=search_value)
            )
        
        # Filter tanggal
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        
        # Get total count before pagination
        total_records = queryset.count()
        
        # Apply pagination
        queryset = queryset[start:start + length]
        
        # Prepare data for DataTables
        data = []
        for log in queryset:
            data.append({
                'id': log.id,
                'product_sku': log.product.sku,
                'product_name': log.product.nama_produk,
                'quantity': log.quantity,
                'rak_kode': log.suggested_rak.kode_rak if log.suggested_rak else '-',
                'rak_nama': log.suggested_rak.nama_rak if log.suggested_rak else '-',
                'putaway_by': log.putaway_by.username if log.putaway_by else 'Belum di-scan',
                'putaway_time': log.putaway_time.strftime('%Y-%m-%d %H:%M') if log.putaway_time else '-',
                'status': 'Selesai' if log.putaway_by else 'Pending',
                'created_at': log.created_at.strftime('%Y-%m-%d %H:%M')
            })
        
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': total_records,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({
            'draw': draw,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': str(e)
        }, status=500)

def auto_slotting(request):
    """
    AJAX endpoint untuk auto slotting - otomatis memilih rak tujuan
    HANYA menentukan lokasi rak_tujuan, TIDAK melakukan putaway
    """
    try:
        data = json.loads(request.body)
        sku = data.get('sku')
        notes = data.get('notes', '')
        
        if not sku:
            return JsonResponse({'success': False, 'error': 'SKU tidak boleh kosong'}, status=400)
        
        with transaction.atomic():
            # Ambil product dan stock
            product = get_object_or_404(Product, sku=sku)
            stock = get_object_or_404(Stock, product=product)
            user = request.user
            
            # Gunakan centralized SlottingService untuk auto slotting
            from .putaway import SlottingService
            
            auto_slotting_result = SlottingService.auto_slotting(product, user, notes)
            
            if not auto_slotting_result['success']:
                return JsonResponse({
                    'success': False, 
                    'error': auto_slotting_result['error']
                }, status=400)
            
            return JsonResponse({
                'success': True,
                'message': auto_slotting_result['message'],
                'slotting_id': auto_slotting_result['slotting_log'].id,
                'selected_rak': auto_slotting_result['rak_info']
            })
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Data JSON tidak valid'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Terjadi kesalahan: {str(e)}'}, status=500)


def rak_capacity_view(request):
    """
    View untuk menampilkan kapasitas rak berdasarkan available_front
    """
    # Ambil semua rak dengan capacity
    rak_capacities = RakCapacity.objects.select_related('rak').all().order_by('rak__kode_rak')
    
    # Hitung summary
    total_raks = rak_capacities.count()
    total_available_front = sum(capacity.available_front for capacity in rak_capacities)
    total_used_front = sum(capacity.used_front for capacity in rak_capacities)
    total_width = sum(capacity.rak.lebar_cm or 0 for capacity in rak_capacities)
    
    # Rata-rata utilization
    avg_utilization = (total_used_front / total_width * 100) if total_width > 0 else 0
    
    context = {
        'rak_capacities': rak_capacities,
        'total_raks': total_raks,
        'total_available_front': total_available_front,
        'total_used_front': total_used_front,
        'total_width': total_width,
        'avg_utilization': avg_utilization,
    }
    
    return render(request, 'inventory/rak_capacity.html', context)


def update_rak_capacity(request):
    """
    AJAX endpoint untuk update rak capacity berdasarkan current stock
    """
    try:
        # Update semua rak capacity
        updated_count = 0
        for capacity in RakCapacity.objects.select_related('rak').all():
            capacity.update_available_front()
            updated_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Berhasil update {updated_count} rak capacity',
            'updated_count': updated_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Terjadi kesalahan: {str(e)}'
        }, status=500)

@require_GET
def rak_detail_data(request):
    """
    API endpoint untuk mendapatkan detail item di rak tertentu
    """
    try:
        rak_code = request.GET.get('rak_code')
        if not rak_code:
            return JsonResponse({
                'success': False,
                'error': 'Kode rak tidak boleh kosong'
            }, status=400)
        
        # Ambil rak
        rak = get_object_or_404(Rak, kode_rak=rak_code)
        
        # Ambil semua stock di rak tersebut
        stocks = InventoryRakStock.objects.filter(
            rak=rak,
            quantity__gt=0  # Hanya yang ada quantity
        ).select_related('product').order_by('-quantity')
        
        items = []
        for stock in stocks:
            # Format dimensi produk
            product_dimensions = []
            if stock.product.lebar_cm:
                product_dimensions.append(f"{float(stock.product.lebar_cm)}cm")
            if stock.product.panjang_cm:
                product_dimensions.append(f"{float(stock.product.panjang_cm)}cm")
            if stock.product.tinggi_cm:
                product_dimensions.append(f"{float(stock.product.tinggi_cm)}cm")
            
            if product_dimensions:
                product_dimensions_str = " × ".join(product_dimensions)
                has_dimensions = True
            else:
                product_dimensions_str = "tolong update dimensi"
                has_dimensions = False
            
            # Hitung berdasarkan 3 dimensi jika ada dimensi produk
            if stock.product.lebar_cm and stock.product.panjang_cm and stock.product.tinggi_cm:
                width_slots_needed = _calculate_width_slots_needed_for_product(rak, stock.product, stock.quantity)
                used_width = width_slots_needed * float(stock.product.lebar_cm)
                percentage = (used_width / float(rak.lebar_cm)) * 100 if rak.lebar_cm else 0
                width_cm = float(stock.product.lebar_cm)
                used_width_str = f"{used_width:.1f}"
                width_slots = width_slots_needed
                products_per_slot = _calculate_products_per_slot(rak, stock.product)
            else:
                # Jika tidak ada dimensi produk, gunakan default values
                width_slots_needed = 1
                used_width = stock.quantity  # Default: 1 unit = 1 cm
                percentage = (used_width / float(rak.lebar_cm)) * 100 if rak.lebar_cm else 0
                width_cm = 0
                used_width_str = f"{used_width:.1f} (estimasi)"
                width_slots = 1
                products_per_slot = 1
            
            items.append({
                'sku': stock.product.sku,
                'nama_produk': stock.product.nama_produk,
                'variant_produk': stock.product.variant_produk,
                'brand': stock.product.brand,
                'photo': stock.product.photo.url if stock.product.photo else None,
                'width_cm': width_cm,
                'product_dimensions': product_dimensions_str,
                'has_dimensions': has_dimensions,
                'product_id': stock.product.id,
                'quantity': stock.quantity,
                'used_width': used_width_str,
                'percentage': percentage,
                'width_slots': width_slots,
                'products_per_slot': products_per_slot
            })
        
        # Format dimensi rak
        rak_dimensions = []
        if rak.lebar_cm:
            rak_dimensions.append(f"{float(rak.lebar_cm)}cm")
        if rak.panjang_cm:
            rak_dimensions.append(f"{float(rak.panjang_cm)}cm")
        if rak.tinggi_cm:
            rak_dimensions.append(f"{float(rak.tinggi_cm)}cm")
        
        rak_dimensions_str = " × ".join(rak_dimensions) if rak_dimensions else "Tidak ada data"
        
        return JsonResponse({
            'success': True,
            'rak_code': rak_code,
            'rak_name': rak.nama_rak,
            'rak_width': float(rak.lebar_cm) if rak.lebar_cm else 0,
            'rak_dimensions': rak_dimensions_str,
            'items': items
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Terjadi kesalahan: {str(e)}'
        }, status=500)


def update_single_rak_capacity(request):
    """
    AJAX endpoint untuk update capacity satu rak tertentu
    """
    try:
        data = json.loads(request.body)
        rak_code = data.get('rak_code')
        
        if not rak_code:
            return JsonResponse({
                'success': False,
                'error': 'Kode rak tidak boleh kosong'
            }, status=400)
        
        # Update capacity untuk rak ini saja
        success, message = update_rak_capacity_for_rak(rak_code)
        
        if success:
            return JsonResponse({
                'success': True,
                'message': message,
                'rak_code': rak_code
            })
        else:
            return JsonResponse({
                'success': False,
                'error': message
            }, status=500)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Data JSON tidak valid'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Terjadi kesalahan: {str(e)}'
        }, status=500)

def update_rak_capacity_for_rak(rak_code):
    """
    Utility function untuk update capacity satu rak tertentu
    """
    try:
        rak = get_object_or_404(Rak, kode_rak=rak_code)
        
        # Update capacity untuk rak ini saja
        try:
            capacity = RakCapacity.objects.get(rak=rak)
        except RakCapacity.DoesNotExist:
            capacity = RakCapacity.objects.create(rak=rak)
        
        # Hitung ulang capacity
        stocks = InventoryRakStock.objects.filter(rak=rak, quantity__gt=0).select_related('product')
        
        total_used_width = 0
        for stock in stocks:
            if (stock.product.lebar_cm and stock.product.panjang_cm and 
                stock.product.tinggi_cm and rak.lebar_cm and 
                rak.panjang_cm and rak.tinggi_cm):
                
                # Hitung berdasarkan 3 dimensi
                width_slots_needed = _calculate_width_slots_needed_for_product(
                    rak, stock.product, stock.quantity
                )
                total_used_width += width_slots_needed * float(stock.product.lebar_cm)
            elif stock.product.lebar_cm:
                # Fallback ke perhitungan lama
                total_used_width += float(stock.product.lebar_cm) * stock.quantity
        
        # Update available_front
        rak_width_float = float(rak.lebar_cm) if rak.lebar_cm else 0
        capacity.available_front = max(0, rak_width_float - total_used_width)
        capacity.save()
        
        return True, f'Capacity rak {rak_code} berhasil diupdate'
        
    except Exception as e:
        return False, f'Error update capacity rak {rak_code}: {str(e)}'


@require_GET
def get_product_dimensions(request, sku):
    """
    AJAX endpoint untuk mengambil dimensi produk
    """
    try:
        product = get_object_or_404(Product, sku=sku)
        
        # Debug: log nilai yang akan dikembalikan
        print(f"DEBUG: Loading dimensions for SKU {sku}")
        print(f"DEBUG: lebar_cm = {product.lebar_cm} (type: {type(product.lebar_cm)})")
        print(f"DEBUG: panjang_cm = {product.panjang_cm} (type: {type(product.panjang_cm)})")
        print(f"DEBUG: tinggi_cm = {product.tinggi_cm} (type: {type(product.tinggi_cm)})")
        print(f"DEBUG: posisi_tidur = {product.posisi_tidur} (type: {type(product.posisi_tidur)})")
        
        # Convert Decimal to float untuk memastikan JSON serialization yang benar
        response_data = {
            'success': True,
            'dimensions': {
                'lebar_cm': float(product.lebar_cm) if product.lebar_cm is not None else None,
                'panjang_cm': float(product.panjang_cm) if product.panjang_cm is not None else None,
                'tinggi_cm': float(product.tinggi_cm) if product.tinggi_cm is not None else None,
                'posisi_tidur': bool(product.posisi_tidur) if product.posisi_tidur is not None else False
            }
        }
        
        print(f"DEBUG: Response data = {response_data}")
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"DEBUG: Error loading dimensions for SKU {sku}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_POST
@csrf_exempt
def update_product_dimensions(request):
    """
    AJAX endpoint untuk mengupdate dimensi produk
    """
    try:
        data = json.loads(request.body)
        sku = data.get('sku')
        lebar_cm = data.get('lebar_cm')
        panjang_cm = data.get('panjang_cm')
        tinggi_cm = data.get('tinggi_cm')
        posisi_tidur = data.get('posisi_tidur', False)
        
        if not sku:
            return JsonResponse({
                'success': False,
                'error': 'SKU tidak boleh kosong'
            }, status=400)
        
        if not all([lebar_cm, panjang_cm, tinggi_cm]):
            return JsonResponse({
                'success': False,
                'error': 'Semua dimensi harus diisi'
            }, status=400)
        
        product = get_object_or_404(Product, sku=sku)
        
        # Update dimensi dan posisi tidur
        product.lebar_cm = lebar_cm
        product.panjang_cm = panjang_cm
        product.tinggi_cm = tinggi_cm
        product.posisi_tidur = posisi_tidur
        product.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Dimensi produk {sku} berhasil diperbarui',
            'dimensions': {
                'lebar_cm': float(product.lebar_cm) if product.lebar_cm is not None else None,
                'panjang_cm': float(product.panjang_cm) if product.panjang_cm is not None else None,
                'tinggi_cm': float(product.tinggi_cm) if product.tinggi_cm is not None else None,
                'posisi_tidur': bool(product.posisi_tidur) if product.posisi_tidur is not None else False
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Data JSON tidak valid'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def get_product_id_by_sku(request):
    """
    API endpoint untuk mendapatkan product_id berdasarkan SKU
    """
    try:
        sku = request.GET.get('sku')
        if not sku:
            return JsonResponse({
                'success': False,
                'error': 'SKU parameter is required'
            }, status=400)
        
        try:
            product = Product.objects.get(sku=sku)
            return JsonResponse({
                'success': True,
                'product_id': product.id,
                'sku': product.sku
            })
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'Product with SKU {sku} not found'
            }, status=404)
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_GET
def get_product_data_for_putaway(request):
    """
    AJAX endpoint untuk mengambil data produk berdasarkan SKU untuk putaway
    """
    try:
        sku = request.GET.get('sku')
        if not sku:
            return JsonResponse({
                'success': False,
                'error': 'SKU tidak boleh kosong'
            }, status=400)
        
        # Get product data
        product = get_object_or_404(Product, sku=sku)
        
        # Get stock data for putaway quantity
        stock = Stock.objects.filter(product=product).first()
        quantity_putaway = stock.quantity_putaway if stock else 0
        
        # Get suggested_rak from PutawaySlottingLog with select_related for better performance
        from .models import PutawaySlottingLog
        slotting_log = PutawaySlottingLog.objects.select_related('suggested_rak').filter(
            product=product,
            putaway_by__isnull=True  # Belum di-scan putaway
        ).order_by('-created_at').first()
        
        suggested_rak = None
        if slotting_log and slotting_log.suggested_rak:
            suggested_rak = slotting_log.suggested_rak.kode_rak
        
        response_data = {
            'success': True,
            'product': {
                'sku': product.sku,
                'nama_produk': product.nama_produk,
                'variant_produk': product.variant_produk,
                'brand': product.brand,
                'barcode': product.barcode,
                'photo': product.photo.url if product.photo else None,
                'quantity_putaway': quantity_putaway,
                'suggested_rak': suggested_rak,  # Add suggested_rak data
            }
        }
        
        response = JsonResponse(response_data)
        
        # Add cache headers to improve performance
        response['Cache-Control'] = 'public, max-age=60'  # Cache for 1 minute
        
        return response
        
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Produk dengan SKU {sku} tidak ditemukan'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def get_product_id_by_sku(request):
    """
    API endpoint untuk mendapatkan product_id berdasarkan SKU
    """
    try:
        sku = request.GET.get('sku')
        if not sku:
            return JsonResponse({
                'success': False,
                'error': 'SKU parameter is required'
            }, status=400)
        
        try:
            product = Product.objects.get(sku=sku)
            return JsonResponse({
                'success': True,
                'product_id': product.id,
                'sku': product.sku
            })
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'Product with SKU {sku} not found'
            }, status=404)
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
