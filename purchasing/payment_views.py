from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Q, Sum
from django.db import transaction
from datetime import timedelta
from purchasing.models import PurchasePayment, PurchaseTaxInvoice, Bank, Purchase


# ==================== PURCHASE PAYMENT ====================

@login_required
def purchase_payment_list(request):
    """List purchase payments"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_finance'):
        raise PermissionDenied("You don't have permission to view purchase payment.")
    
    from inventory.models import Supplier
    
    banks = Bank.objects.filter(is_active=True).order_by('nama_bank')
    # Get unique suppliers from purchase payments
    suppliers = Supplier.objects.filter(
        id__in=PurchasePayment.objects.values_list('supplier_id', flat=True).distinct()
    ).order_by('nama_supplier')
    
    return render(request, 'purchasing/purchase_payment_list.html', {
        'banks': banks,
        'suppliers': suppliers
    })


@login_required
def purchase_payment_api(request):
    """API endpoint for purchase payments with pagination and filtering"""
    
    try:
        # Get pagination parameters
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 25))
        
        # Get filter parameters
        status_filter = request.GET.get('status', '').strip()
        supplier_filter = request.GET.get('supplier', '').strip()
        purchase_filter = request.GET.get('purchase', '').strip()
        transaction_type_filter = request.GET.get('transaction_type', '').strip()
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        
        # Build query
        queryset = PurchasePayment.objects.select_related(
            'purchase', 
            'supplier'
        ).prefetch_related('allocations__transfer_from').order_by('-due_date')
        
        # Apply filters
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if supplier_filter:
            queryset = queryset.filter(supplier_id=supplier_filter)
        
        if purchase_filter:
            queryset = queryset.filter(purchase__nomor_purchase__icontains=purchase_filter)
        
        if transaction_type_filter:
            queryset = queryset.filter(transaction_type=transaction_type_filter)
        
        if date_from:
            try:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(due_date__gte=date_from_obj)
            except:
                pass
        
        if date_to:
            try:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                end_of_day = date_to_obj + timedelta(days=1)
                queryset = queryset.filter(due_date__lt=end_of_day)
            except:
                pass
        
        # Get total count
        total_count = queryset.count()
        
        # Pagination
        start = (page - 1) * page_size
        end = start + page_size
        payments = queryset[start:end]
        
        # Calculate total pages
        total_pages = (total_count + page_size - 1) // page_size
        
        # Prepare data - show each allocation as separate row
        data = []
        for payment in payments:
            allocations = payment.allocations.order_by('-allocation_date')
            
            if allocations.exists():
                # Show each allocation as separate row
                for allocation in allocations:
                    payment_method_display = allocation.get_payment_method_display() if allocation.payment_method else '-'
                    transfer_from_display = '-'
                    
                    if allocation.transfer_from:
                        transfer_from_display = f"{allocation.transfer_from.nama_bank} - {allocation.transfer_from.nomor_rekening}"
                    
                    # Calculate due date status
                    from datetime import timedelta
                    today = timezone.now().date()
                    due_date_obj = payment.due_date.date()
                    days_until_due = (due_date_obj - today).days
                    
                    if days_until_due < 0:
                        due_date_status = 'overdue'
                        due_date_status_display = 'Overdue'
                    elif days_until_due <= 2:
                        due_date_status = 'warning'
                        due_date_status_display = f'H-{days_until_due}'
                    else:
                        due_date_status = 'safe'
                        due_date_status_display = f'{days_until_due} days'
                    
                    data.append({
                        'id': payment.id,
                        'allocation_id': allocation.id,
                        'purchase_number': payment.purchase.nomor_purchase,
                        'supplier': payment.supplier.nama_supplier,
                        'total_amount': float(payment.total_amount),
                        'total_amount_formatted': f"{payment.total_amount:,.0f}".replace(',', '.'),
                        'discount': float(payment.discount),
                        'discount_formatted': f"{payment.discount:,.0f}".replace(',', '.'),
                        'paid_amount': float(allocation.amount),
                        'paid_amount_formatted': f"{allocation.amount:,.0f}".replace(',', '.'),
                        'remaining_amount': float(payment.remaining_amount),
                        'remaining_amount_formatted': f"{payment.remaining_amount:,.0f}".replace(',', '.'),
                        'due_date': payment.due_date.strftime('%d %b %Y'),
                        'due_date_status': due_date_status,
                        'due_date_status_display': due_date_status_display,
                        'days_until_due': days_until_due,
                        'payment_date': allocation.allocation_date.strftime('%d %b %Y'),
                        'status': payment.status,
                        'status_display': payment.get_status_display(),
                        'transaction_type': payment.transaction_type,
                        'transaction_type_display': payment.get_transaction_type_display(),
                        'payment_method': payment_method_display,
                        'transfer_from': transfer_from_display,
                        'notes': allocation.notes or '-',
                    })
            else:
                # No allocations yet - show payment record
                # Calculate due date status
                from datetime import timedelta
                today = timezone.now().date()
                due_date_obj = payment.due_date.date()
                days_until_due = (due_date_obj - today).days
                
                if days_until_due < 0:
                    due_date_status = 'overdue'
                    due_date_status_display = 'Overdue'
                elif days_until_due <= 2:
                    due_date_status = 'warning'
                    due_date_status_display = f'H-{days_until_due}'
                else:
                    due_date_status = 'safe'
                    due_date_status_display = f'{days_until_due} days'
                
                data.append({
                    'id': payment.id,
                    'allocation_id': None,
                    'purchase_number': payment.purchase.nomor_purchase,
                    'supplier': payment.supplier.nama_supplier,
                    'total_amount': float(payment.total_amount),
                    'total_amount_formatted': f"{payment.total_amount:,.0f}".replace(',', '.'),
                    'discount': float(payment.discount),
                    'discount_formatted': f"{payment.discount:,.0f}".replace(',', '.'),
                    'paid_amount': float(payment.paid_amount),
                    'paid_amount_formatted': f"{payment.paid_amount:,.0f}".replace(',', '.'),
                    'remaining_amount': float(payment.remaining_amount),
                    'remaining_amount_formatted': f"{payment.remaining_amount:,.0f}".replace(',', '.'),
                    'due_date': payment.due_date.strftime('%d %b %Y'),
                    'due_date_status': due_date_status,
                    'due_date_status_display': due_date_status_display,
                    'days_until_due': days_until_due,
                    'payment_date': '-',
                    'status': payment.status,
                    'status_display': payment.get_status_display(),
                    'transaction_type': payment.transaction_type,
                    'transaction_type_display': payment.get_transaction_type_display(),
                    'payment_method': '-',
                    'transfer_from': '-',
                    'notes': payment.notes or '-',
                })
        
        # Calculate statistics
        stats = {
            'total_unpaid': queryset.filter(status__in=['unpaid', 'partial', 'overdue']).count(),
            'total_unpaid_amount': float(queryset.filter(status__in=['unpaid', 'partial', 'overdue']).aggregate(total=Sum('remaining_amount'))['total'] or 0),
            'total_overdue': queryset.filter(status='overdue').count(),
            'total_overdue_amount': float(queryset.filter(status='overdue').aggregate(total=Sum('remaining_amount'))['total'] or 0),
        }
        
        return JsonResponse({
            'data': data,
            'total_pages': total_pages,
            'current_page': page,
            'stats': stats
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': str(e),
            'data': [],
            'total_pages': 0,
            'current_page': 1,
            'stats': {
                'total_unpaid': 0,
                'total_unpaid_amount': 0,
                'total_overdue': 0,
                'total_overdue_amount': 0
            }
        }, status=500)


@login_required
@require_POST
def purchase_payment_update(request, payment_id):
    """Update payment (mark as paid) - create allocation"""
    from purchasing.models import PurchasePaymentAllocation
    
    payment = get_object_or_404(PurchasePayment, id=payment_id)
    
    try:
        paid_amount = float(request.POST.get('paid_amount', 0))
        discount = float(request.POST.get('discount', 0))
        payment_method = request.POST.get('payment_method', '')
        payment_date = request.POST.get('payment_date', '')
        reference_number = request.POST.get('reference_number', '')
        transfer_from_id = request.POST.get('transfer_from', '')
        notes = request.POST.get('notes', '')
        
        if paid_amount <= 0:
            messages.error(request, 'Payment amount harus lebih dari 0!')
            return redirect('purchasing:purchase_payment_list')
        
        # Update discount
        if discount != payment.discount:
            payment.discount = int(discount)
            payment.save()
        
        # Check if amount exceeds remaining
        if paid_amount > payment.remaining_amount:
            messages.error(request, f'Payment amount ({paid_amount:,.0f}) melebihi remaining amount ({payment.remaining_amount:,.0f})!')
            return redirect('purchasing:purchase_payment_list')
        
        # Validate transfer_from if payment method is transfer
        transfer_from = None
        if payment_method == 'transfer':
            if not transfer_from_id:
                messages.error(request, 'Transfer From harus dipilih untuk payment method Transfer!')
                return redirect('purchasing:purchase_payment_list')
            transfer_from = get_object_or_404(Bank, id=transfer_from_id, is_active=True)
        
        # Create payment allocation
        allocation = PurchasePaymentAllocation.objects.create(
            payment=payment,
            amount=paid_amount,
            payment_method=payment_method,
            transfer_from=transfer_from,
            reference_number=reference_number,
            notes=notes,
            created_by=request.user
        )
        
        # Update allocation date if provided
        if payment_date:
            from datetime import datetime
            allocation.allocation_date = datetime.strptime(payment_date, '%Y-%m-%d')
            allocation.save()
        
        # Update payment (auto-calculate from allocations)
        payment.save()
        
        messages.success(request, f'Payment allocation untuk {payment.purchase.nomor_purchase} berhasil dibuat!')
        return redirect('purchasing:purchase_payment_list')
    
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('purchasing:purchase_payment_list')


# ==================== PURCHASE TAX INVOICE ====================

@login_required
def purchase_taxinvoice_list(request):
    """List purchase tax invoices"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_finance'):
        raise PermissionDenied("You don't have permission to view tax invoice.")
    
    return render(request, 'purchasing/purchase_taxinvoice_list.html')


@login_required
def purchase_taxinvoice_download_excel(request):
    """Download tax invoices to Excel with filters"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # Get filter parameters
    status_filter = request.GET.get('status', '').strip()
    supplier_filter = request.GET.get('supplier', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    # Build query
    queryset = PurchaseTaxInvoice.objects.select_related('purchase', 'supplier').order_by('-created_at')
    
    # Apply filters
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    
    if supplier_filter:
        queryset = queryset.filter(supplier__nama_supplier__icontains=supplier_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            end_of_day = date_to_obj + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_of_day)
        except:
            pass
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Invoices"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "LAPORAN TAX INVOICE"
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Filter info
    row = 3
    ws[f'A{row}'] = "Filter:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    if status_filter:
        ws[f'A{row}'] = f"Status: {status_filter}"
        row += 1
    if supplier_filter:
        ws[f'A{row}'] = f"Supplier: {supplier_filter}"
        row += 1
    if date_from:
        ws[f'A{row}'] = f"From: {date_from}"
        row += 1
    if date_to:
        ws[f'A{row}'] = f"To: {date_to}"
        row += 1
    
    row += 1
    
    # Headers
    headers = [
        'No', 'Purchase Number', 'Supplier', 'Tax Invoice Number', 'Tax Invoice Date',
        'Subtotal', 'Tax (11%)', 'Discount', 'Total Amount', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Data
    for idx, invoice in enumerate(queryset, 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=invoice.purchase.nomor_purchase).border = border
        ws.cell(row=row, column=3, value=invoice.supplier.nama_supplier).border = border
        ws.cell(row=row, column=4, value=invoice.invoice_number or '-').border = border
        ws.cell(row=row, column=5, value=invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '-').border = border
        ws.cell(row=row, column=6, value=invoice.subtotal).border = border
        ws.cell(row=row, column=7, value=invoice.tax_amount).border = border
        ws.cell(row=row, column=8, value=invoice.discount).border = border
        ws.cell(row=row, column=9, value=invoice.invoice_amount).border = border
        ws.cell(row=row, column=10, value=invoice.get_status_display()).border = border
        
        # Right align numbers
        for col in [6, 7, 8, 9]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Auto-adjust column widths
    column_widths = [5, 20, 25, 25, 15, 15, 15, 15, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with filters
    filename = f"tax_invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    if status_filter:
        filename += f"_{status_filter}"
    if date_from:
        filename += f"_from{date_from}"
    if date_to:
        filename += f"_to{date_to}"
    filename += ".xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def purchase_taxinvoice_download_template(request):
    """Download template Excel for bulk upload"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Upload"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "TEMPLATE UPLOAD TAX INVOICE"
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Instructions
    ws['A3'] = "INSTRUKSI:"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = "1. Isi kolom Purchase Number, Tax Invoice Number, Tax Invoice Date, dan Notes"
    ws['A5'] = "2. Tax Invoice Number harus unik"
    ws['A6'] = "3. Tax Invoice Date format: YYYY-MM-DD (contoh: 2025-01-19)"
    ws['A7'] = "4. Notes bersifat opsional"
    ws['A8'] = "5. Jangan hapus atau ubah header"
    
    row = 10
    
    # Headers
    headers = ['Purchase Number', 'Tax Invoice Number', 'Tax Invoice Date', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Sample data
    sample_data = [
        ['PUR-20251019-0044', 'INV-001-2025', '2025-01-19', 'Sample invoice'],
        ['PUR-20251019-0045', 'INV-002-2025', '2025-01-20', ''],
    ]
    
    for data in sample_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
        row += 1
    
    # Auto-adjust column widths
    column_widths = [20, 20, 15, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"tax_invoice_upload_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def purchase_taxinvoice_upload(request):
    """Upload tax invoices in bulk from Excel"""
    from openpyxl import load_workbook
    from datetime import datetime
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            
            # Load workbook
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Skip header (row 10 based on template)
            start_row = 11
            
            success_count = 0
            error_count = 0
            errors = []
            
            with transaction.atomic():
                for row in range(start_row, ws.max_row + 1):
                    # Get values
                    purchase_number = ws.cell(row=row, column=1).value
                    invoice_number = ws.cell(row=row, column=2).value
                    invoice_date_str = ws.cell(row=row, column=3).value
                    notes = ws.cell(row=row, column=4).value or ''
                    
                    # Skip empty rows
                    if not purchase_number or not invoice_number:
                        continue
                    
                    try:
                        # Find purchase
                        purchase = Purchase.objects.get(nomor_purchase=str(purchase_number))
                        
                        # Get or create tax invoice
                        tax_invoice = PurchaseTaxInvoice.objects.get(purchase=purchase)
                        
                        # Check if invoice number already exists
                        if PurchaseTaxInvoice.objects.filter(invoice_number=str(invoice_number)).exclude(id=tax_invoice.id).exists():
                            errors.append(f"Row {row}: Invoice number '{invoice_number}' already exists")
                            error_count += 1
                            continue
                        
                        # Parse invoice date
                        invoice_date = None
                        if invoice_date_str:
                            if isinstance(invoice_date_str, str):
                                invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d')
                            else:
                                invoice_date = invoice_date_str
                        
                        # Update tax invoice
                        tax_invoice.invoice_number = str(invoice_number)
                        tax_invoice.invoice_date = invoice_date
                        tax_invoice.notes = notes
                        tax_invoice.status = 'received'
                        tax_invoice.save()
                        
                        success_count += 1
                        
                    except Purchase.DoesNotExist:
                        errors.append(f"Row {row}: Purchase '{purchase_number}' not found")
                        error_count += 1
                    except PurchaseTaxInvoice.DoesNotExist:
                        errors.append(f"Row {row}: Tax invoice not found for purchase '{purchase_number}'")
                        error_count += 1
                    except Exception as e:
                        errors.append(f"Row {row}: {str(e)}")
                        error_count += 1
            
            # Return result
            if error_count == 0:
                messages.success(request, f'Berhasil upload {success_count} tax invoice!')
            else:
                messages.warning(request, f'Berhasil upload {success_count} tax invoice. {error_count} error: ' + '; '.join(errors[:5]))
            
            return redirect('purchasing:purchase_taxinvoice_list')
            
        except Exception as e:
            messages.error(request, f'Error uploading file: {str(e)}')
            return redirect('purchasing:purchase_taxinvoice_list')
    
    messages.error(request, 'No file uploaded')
    return redirect('purchasing:purchase_taxinvoice_list')


@login_required
def purchase_taxinvoice_api(request):
    """API endpoint for purchase tax invoices with pagination and filtering"""
    
    # Get pagination parameters
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 25))
    
    # Get filter parameters
    status_filter = request.GET.get('status', '').strip()
    supplier_filter = request.GET.get('supplier', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    # Build query
    queryset = PurchaseTaxInvoice.objects.select_related(
        'purchase', 
        'supplier'
    ).order_by('-created_at')
    
    # Apply filters
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    
    if supplier_filter:
        queryset = queryset.filter(supplier__nama_supplier__icontains=supplier_filter)
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            end_of_day = date_to_obj + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_of_day)
        except:
            pass
    
    # Get total count
    total_count = queryset.count()
    
    # Pagination
    start = (page - 1) * page_size
    end = start + page_size
    invoices = queryset[start:end]
    
    # Calculate total pages
    total_pages = (total_count + page_size - 1) // page_size
    
    # Prepare data
    data = []
    for invoice in invoices:
        data.append({
            'id': invoice.id,
            'purchase_number': invoice.purchase.nomor_purchase,
            'supplier': invoice.supplier.nama_supplier,
            'invoice_number': invoice.invoice_number or '-',
            'invoice_date': invoice.invoice_date.strftime('%d %b %Y') if invoice.invoice_date else '-',
            'invoice_amount': float(invoice.invoice_amount),
            'invoice_amount_formatted': f"{invoice.invoice_amount:,.0f}".replace(',', '.'),
            'tax_amount': float(invoice.tax_amount),
            'tax_amount_formatted': f"{invoice.tax_amount:,.0f}".replace(',', '.'),
            'subtotal': float(invoice.subtotal),
            'subtotal_formatted': f"{invoice.subtotal:,.0f}".replace(',', '.'),
            'tax_rate': float(invoice.tax_rate),
            'status': invoice.status,
            'status_display': invoice.get_status_display(),
            'received_at': invoice.received_at.strftime('%d %b %Y') if invoice.received_at else '-',
            'notes': invoice.notes or '-',
        })
    
    # Calculate statistics
    stats = {
        # Overall Summary
        'total_taxinvoice': queryset.count(),
        'total_ppn': float(queryset.aggregate(total=Sum('tax_amount'))['total'] or 0),
        'total_invoice_amount': float(queryset.aggregate(total=Sum('invoice_amount'))['total'] or 0),
        'total_subtotal': float(queryset.aggregate(total=Sum('subtotal'))['total'] or 0),
        
        # Pending
        'total_pending_invoice': queryset.filter(status='pending').count(),
        'total_pending_ppn': float(queryset.filter(status='pending').aggregate(total=Sum('tax_amount'))['total'] or 0),
        'total_pending_amount': float(queryset.filter(status='pending').aggregate(total=Sum('invoice_amount'))['total'] or 0),
        
        # Received
        'total_received_invoice': queryset.filter(status='received').count(),
        'total_received_ppn': float(queryset.filter(status='received').aggregate(total=Sum('tax_amount'))['total'] or 0),
        'total_received_amount': float(queryset.filter(status='received').aggregate(total=Sum('invoice_amount'))['total'] or 0),
        
        # Verified
        'total_verified_invoice': queryset.filter(status='verified').count(),
        'total_verified_ppn': float(queryset.filter(status='verified').aggregate(total=Sum('tax_amount'))['total'] or 0),
        
        # Legacy (for backward compatibility)
        'total_pending': queryset.filter(status='pending').count(),
        'total_received': queryset.filter(status='received').count(),
        'total_verified': queryset.filter(status='verified').count(),
    }
    
    return JsonResponse({
        'data': data,
        'total_pages': total_pages,
        'current_page': page,
        'stats': stats
    })


@login_required
@require_POST
def purchase_taxinvoice_update(request, invoice_id):
    """Update tax invoice (add invoice number)"""
    
    invoice = get_object_or_404(PurchaseTaxInvoice, id=invoice_id)
    
    try:
        invoice_number = request.POST.get('invoice_number', '').strip()
        invoice_date = request.POST.get('invoice_date', '').strip()
        total_amount = request.POST.get('total_amount', '').strip()
        notes = request.POST.get('notes', '')
        
        if not invoice_number:
            messages.error(request, 'Nomor faktur pajak wajib diisi!')
            return redirect('purchasing:purchase_taxinvoice_list')
        
        if not total_amount:
            messages.error(request, 'Total amount wajib diisi!')
            return redirect('purchasing:purchase_taxinvoice_list')
        
        # Parse total amount (remove thousand separator)
        import re
        total_amount_clean = re.sub(r'[^\d]', '', total_amount)
        if not total_amount_clean:
            messages.error(request, 'Total amount tidak valid!')
            return redirect('purchasing:purchase_taxinvoice_list')
        
        total_amount_int = int(total_amount_clean)
        
        # Update invoice
        invoice.invoice_number = invoice_number
        if invoice_date:
            from datetime import datetime
            invoice.invoice_date = datetime.strptime(invoice_date, '%Y-%m-%d')
        invoice.invoice_amount = total_amount_int
        invoice.status = 'received'
        invoice.received_at = timezone.now()
        if notes:
            invoice.notes = notes
        invoice.save()
        
        messages.success(request, f'Tax invoice untuk {invoice.purchase.nomor_purchase} berhasil diupdate!')
        return redirect('purchasing:purchase_taxinvoice_list')
    
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('purchasing:purchase_taxinvoice_list')

