from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.db import models
from django.template.loader import render_to_string
from django.conf import settings
import re

from .models import PurchaseOrder, PurchaseOrderItem, PurchaseOrderHistory, PriceHistory, Purchase, PurchaseItem
from inventory.models import Inbound, InboundItem, Supplier
from products.models import Product

# Import payment views, bank views, and report views
from . import payment_views, bank_views, report_views


def create_po_history(po, action, user, notes=None, old_values=None, new_values=None):
    """Helper function to create PO history"""
    PurchaseOrderHistory.objects.create(
        po=po,
        action=action,
        user=user,
        notes=notes,
        old_values=old_values,
        new_values=new_values
    )


def get_next_po_id():
    """Generate next PO ID"""
    last = PurchaseOrder.objects.order_by('-id').first()
    if last:
        return last.id + 1
    return 1


@login_required
def po_list(request):
    """List all Purchase Orders"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.po_view'):
        raise PermissionDenied("You don't have permission to view Purchase Orders.")
    
    return render(request, 'purchasing/po_list.html')


@login_required
def po_list_api(request):
    """API endpoint for PO list with pagination and filtering"""
    
    try:
        # Get pagination parameters
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 25))
        
        # Get filter parameters
        status_filter = request.GET.get('status', '').strip()
        supplier_filter = request.GET.get('supplier', '').strip()
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        
        # Build query
        queryset = PurchaseOrder.objects.select_related('supplier', 'created_by', 'received_by').order_by('-tanggal_po', '-id')
        
        # Apply filters
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if supplier_filter:
            queryset = queryset.filter(supplier__nama_supplier__icontains=supplier_filter)
        
        if date_from:
            queryset = queryset.filter(tanggal_po__gte=date_from)
        
        if date_to:
            # Add one day to include the end date
            from datetime import datetime, timedelta
            end_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            queryset = queryset.filter(tanggal_po__lt=end_date.strftime('%Y-%m-%d'))
        
        # Get total count before pagination
        total_count = queryset.count()
        
        # Paginate
        paginator = Paginator(queryset, page_size)
        page_obj = paginator.get_page(page)
        
        # Serialize data
        data = []
        for po in page_obj:
            data.append({
                'id': po.id,
                'nomor_po': po.nomor_po,
                'tanggal_po': po.tanggal_po.strftime('%Y-%m-%d') if po.tanggal_po else '',
                'supplier': po.supplier.nama_supplier if po.supplier else '-',
                'status': po.get_status_display(),
                'status_value': po.status,
                'total_amount': po.total_amount,
                'total_amount_formatted': f'{po.total_amount:,}',
                'created_by': po.created_by.username if po.created_by else '-',
                'created_at': po.created_at.strftime('%Y-%m-%d %H:%M') if po.created_at else '',
            })
        
        # Calculate statistics
        all_pos = PurchaseOrder.objects.all()
        stats = {
            'total_pos': all_pos.count(),
            'draft_pos': all_pos.filter(status='draft').count(),
            'pending_pos': all_pos.filter(status='pending').count(),
            'received_pos': all_pos.filter(status='received').count(),
            'cancelled_pos': all_pos.filter(status='cancelled').count(),
            'total_amount': all_pos.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        }
        
        return JsonResponse({
            'success': True,
            'data': data,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_count': total_count,
                'total_pages': paginator.num_pages,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
            },
            'stats': stats,
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def po_create(request):
    """Create new Purchase Order"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.po_marketing'):
        raise PermissionDenied("You don't have permission to create Purchase Orders.")
    
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    template_name = 'purchasing/po_create_mobile.html' if is_mobile else 'purchasing/po_create.html'
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get po_id from form (draft yang sudah dibuat di GET request)
                po_id = request.POST.get('po_id')
                if not po_id:
                    messages.error(request, 'PO ID tidak ditemukan!')
                    return redirect('purchasing:po_list')
                
                # Get existing draft PO
                po = get_object_or_404(PurchaseOrder, id=po_id, status='draft')
                
                # Get form data
                nomor_po = request.POST.get('nomor_po')
                supplier_id = request.POST.get('supplier_id')
                tanggal_po = request.POST.get('tanggal_po')
                notes = request.POST.get('notes', '')
                produk_sku = request.POST.getlist('produk_sku[]')
                produk_qty = request.POST.getlist('produk_qty[]')
                produk_harga = request.POST.getlist('produk_harga_beli[]')
                
                if not (nomor_po and supplier_id and tanggal_po and produk_sku):
                    messages.error(request, "Nomor PO, Supplier, Tanggal, dan minimal 1 produk wajib diisi.")
                    return render(request, template_name, {
                        'po': po,
                        'default_nomor_po': nomor_po,
                        'default_tanggal_po': tanggal_po,
                        'notes': notes,
                    })
                
                # Update PO
                po.nomor_po = nomor_po
                po.supplier_id = supplier_id
                po.tanggal_po = tanggal_po
                po.notes = notes
                po.status = 'pending'  # Change status to 'pending'
                po.save()
                
                # Delete old items
                po.items.all().delete()
                
                # Create PO Items
                for sku, qty_str, harga_str in zip(produk_sku, produk_qty, produk_harga):
                    if not (sku and qty_str and harga_str):
                        continue
                    
                    qty = int(qty_str)
                    harga = float(harga_str)
                    product = Product.objects.filter(sku=sku).first()
                    
                    if not product:
                        messages.warning(request, f"SKU '{sku}' tidak ditemukan dan dilewati.")
                        continue
                    
                    PurchaseOrderItem.objects.create(
                        po=po,
                        product=product,
                        quantity=qty,
                        harga_beli=harga
                    )
                
                # Create history
                create_po_history(po, 'create', request.user, f'PO {nomor_po} dibuat')
                
                messages.success(request, f"Purchase Order '{nomor_po}' berhasil disimpan.")
                return redirect('purchasing:po_detail', pk=po.pk)
                
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
            return render(request, template_name, {
                'default_nomor_po': request.POST.get('nomor_po'),
                'default_tanggal_po': request.POST.get('tanggal_po'),
                'notes': request.POST.get('notes'),
            })
    
    # GET request - Create new draft PO and redirect to edit
    # Create draft PO (without supplier - can be null for draft)
    # User must select supplier before saving
    # Nomor PO akan di-generate otomatis saat status berubah dari draft
    draft_po = PurchaseOrder.objects.create(
        nomor_po=None,  # No nomor PO for draft
        supplier=None,  # No supplier for draft
        created_by=request.user,
        status='draft',
    )
    
    # Redirect to edit page (not render create template)
    return redirect('purchasing:po_edit', pk=draft_po.id)


@login_required
def po_detail(request, pk):
    """View Purchase Order detail"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.po_view'):
        raise PermissionDenied("You don't have permission to view Purchase Orders.")
    
    po = get_object_or_404(PurchaseOrder.objects.select_related('supplier', 'created_by', 'received_by'), pk=pk)
    items = PurchaseOrderItem.objects.filter(po=po).select_related('product')
    history = PurchaseOrderHistory.objects.filter(po=po).select_related('user').order_by('-timestamp')
    
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    template_name = 'purchasing/po_detail_mobile.html' if is_mobile else 'purchasing/po_detail.html'
    
    context = {
        'po': po,
        'items': items,
        'history': history,
    }
    return render(request, template_name, context)


@login_required
def po_edit(request, pk):
    """Edit Purchase Order"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.po_marketing'):
        raise PermissionDenied("You don't have permission to edit Purchase Orders.")
    
    po = get_object_or_404(PurchaseOrder, pk=pk)
    
    # Check if PO already received or cancelled
    # Allow edit for 'draft' and 'pending' status
    if po.status in ['received', 'cancelled']:
        messages.error(request, f"PO dengan status '{po.status}' tidak dapat diedit.")
        return redirect('purchasing:po_detail', pk=pk)
    
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    template_name = 'purchasing/po_edit_mobile.html' if is_mobile else 'purchasing/po_edit.html'
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Store old values for history
                old_values = {
                    'supplier': str(po.supplier),
                    'tanggal_po': str(po.tanggal_po),
                    'notes': po.notes,
                    'items_count': po.items.count(),
                }
                
                # Update PO header
                supplier_id = request.POST.get('supplier_id')
                tanggal_po = request.POST.get('tanggal_po')
                notes = request.POST.get('notes', '')
                
                # Validate supplier is required before saving
                if not supplier_id:
                    messages.error(request, "Supplier wajib diisi sebelum menyimpan PO!")
                    items = po.items.select_related('product').all()
                    context = {
                        'po': po,
                        'items': items,
                    }
                    return render(request, template_name, context)
                
                po.supplier_id = supplier_id
                po.tanggal_po = tanggal_po
                po.notes = notes
                
                # If status is 'draft', change to 'pending' after save
                if po.status == 'draft':
                    po.status = 'pending'
                
                po.save(update_fields=['supplier', 'tanggal_po', 'notes', 'status'])
                
                # Get form data
                produk_sku = request.POST.getlist('produk_sku[]')
                produk_qty = request.POST.getlist('produk_qty[]')
                produk_harga = request.POST.getlist('produk_harga_beli[]')
                
                # Create dictionary of old items
                old_items = {item.product.sku: item for item in po.items.all()}
                current_skus = set()
                
                # Process items from form
                for sku, qty_str, harga_str in zip(produk_sku, produk_qty, produk_harga):
                    if not (sku and qty_str):
                        continue
                    
                    qty = int(qty_str)
                    harga = float(harga_str) if harga_str else 0
                    current_skus.add(sku)
                    product = Product.objects.filter(sku=sku).first()
                    
                    if not product:
                        continue
                    
                    if sku in old_items:
                        # Update existing item
                        old_item = old_items[sku]
                        old_item.quantity = qty
                        
                        # If user provided harga (> 0), use it. Otherwise keep existing harga_beli
                        if harga > 0:
                            old_item.harga_beli = harga
                        # else: keep existing harga_beli from database
                        
                        # Always save with harga_beli to ensure subtotal is calculated correctly
                        old_item.save(update_fields=['quantity', 'harga_beli', 'subtotal'])
                    else:
                        # Create new item
                        PurchaseOrderItem.objects.create(
                            po=po,
                            product=product,
                            quantity=qty,
                            harga_beli=harga
                        )
                
                # Delete items that are not in form anymore
                skus_to_delete = set(old_items.keys()) - current_skus
                for sku in skus_to_delete:
                    old_items[sku].delete()
                
                # Create history
                new_values = {
                    'supplier': str(po.supplier),
                    'tanggal_po': str(po.tanggal_po),
                    'notes': po.notes,
                    'items_count': po.items.count(),
                }
                create_po_history(po, 'update', request.user, f'PO {po.nomor_po} diedit', old_values, new_values)
                
                if po.status == 'pending':
                    messages.success(request, f"PO {po.nomor_po} berhasil disimpan!")
                else:
                    messages.success(request, f"PO {po.nomor_po} berhasil diperbarui.")
                return redirect('purchasing:po_detail', pk=pk)
                
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
    
    # GET request
    items = po.items.select_related('product').all()
    context = {
        'po': po,
        'items': items,
    }
    return render(request, template_name, context)


@login_required
def po_auto_save(request, pk):
    """Auto-save PO changes via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    
    try:
        po = get_object_or_404(PurchaseOrder, pk=pk)
        
        # Only allow auto-save for draft and pending status
        if po.status not in ['draft', 'pending']:
            return JsonResponse({'success': False, 'error': 'Cannot auto-save received or cancelled PO'})
        
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Auto-save PO {pk}: supplier={request.POST.get('supplier_id')}, items={len(request.POST.getlist('produk_sku[]'))}")
        
        with transaction.atomic():
            # Update PO fields
            po.nomor_po = request.POST.get('nomor_po', po.nomor_po)
            po.supplier_id = request.POST.get('supplier_id', po.supplier_id)
            po.tanggal_po = request.POST.get('tanggal_po', po.tanggal_po)
            po.notes = request.POST.get('notes', po.notes)
            po.save()
            
            # Update items
            skus = request.POST.getlist('produk_sku[]')
            qtys = request.POST.getlist('produk_qty[]')
            hargas = request.POST.getlist('produk_harga_beli[]')
            
            logger.info(f"Auto-save items: skus={skus}, qtys={qtys}, hargas={hargas}")
            
            # Get current items
            old_items = {item.product.sku: item for item in po.items.all()}
            current_skus = set()
            
            for sku, qty_str, harga_str in zip(skus, qtys, hargas):
                logger.info(f"Processing item: sku={sku}, qty={qty_str}, harga={harga_str}")
                if not (sku and qty_str and harga_str):
                    logger.warning(f"Skipping item: sku={sku}, qty={qty_str}, harga={harga_str} (empty)")
                    continue
                
                qty = int(qty_str)
                # Clean harga: remove thousand separators (.)
                harga_clean = str(harga_str).replace('.', '').replace(',', '')
                harga = float(harga_clean) if harga_clean else 0
                current_skus.add(sku)
                
                product = Product.objects.filter(sku=sku).first()
                if not product:
                    continue
                
                if sku in old_items:
                    # Update existing item
                    old_item = old_items[sku]
                    old_item.quantity = qty
                    old_item.harga_beli = harga
                    old_item.save(update_fields=['quantity', 'harga_beli', 'subtotal'])
                    logger.info(f"Updated item: {sku}, qty={qty}, harga={harga}")
                else:
                    # Create new item
                    PurchaseOrderItem.objects.create(
                        po=po,
                        product=product,
                        quantity=qty,
                        harga_beli=harga
                    )
                    logger.info(f"Created new item: {sku}, qty={qty}, harga={harga}")
            
            # Delete items that are not in form anymore
            skus_to_delete = set(old_items.keys()) - current_skus
            for sku in skus_to_delete:
                old_items[sku].delete()
            
            logger.info(f"Auto-save PO {pk} completed: {len(current_skus)} items saved")
            return JsonResponse({'success': True, 'message': 'PO berhasil disimpan otomatis'})
    
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Auto-save PO {pk} failed: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def po_cancel(request, pk):
    """Delete Purchase Order (Draft or Pending)"""
    po = get_object_or_404(PurchaseOrder, pk=pk)
    
    # Only allow delete for 'draft' and 'pending' status
    if po.status in ['draft', 'pending']:
        try:
            po_nomor = po.nomor_po
            po.delete()
            messages.success(request, f'PO {po_nomor} berhasil dihapus!')
        except Exception as e:
            messages.error(request, f"Gagal menghapus PO: {str(e)}")
    else:
        messages.error(request, f"PO dengan status '{po.status}' tidak dapat dihapus.")
    
    return redirect('purchasing:po_list')


@login_required
def po_delete(request, pk):
    """Delete Purchase Order - Only Draft PO can be deleted"""
    po = get_object_or_404(PurchaseOrder, pk=pk)
    
    # Prevent deletion of received or cancelled PO
    if po.status != 'draft':
        messages.error(request, f"PO {po.nomor_po} tidak dapat dihapus karena status sudah '{po.get_status_display()}'.")
        return redirect('purchasing:po_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            nomor_po = po.nomor_po
            po.delete()
            messages.success(request, f'PO {nomor_po} berhasil dihapus.')
            return redirect('purchasing:po_list')
        except Exception as e:
            messages.error(request, f"Gagal menghapus PO: {str(e)}")
            return redirect('purchasing:po_detail', pk=pk)
    
    context = {
        'po': po,
    }
    return render(request, 'purchasing/po_confirm_delete.html', context)


@login_required
def po_receive(request, pk):
    """Receive PO - Auto create Inbound"""
    po = get_object_or_404(PurchaseOrder, pk=pk)
    
    if po.status == 'received':
        messages.error(request, "PO ini sudah di-receive sebelumnya.")
        return redirect('purchasing:po_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Generate Inbound number
                next_id = Inbound.objects.count() + 1
                now = timezone.localtime()
                nomor_inbound = f"INB/{now.strftime('%d-%m-%Y')}/{next_id}"
                
                # Create Inbound
                inbound = Inbound.objects.create(
                    nomor_inbound=nomor_inbound,
                    tanggal=timezone.now(),
                    supplier=po.supplier,
                    po=po,
                    keterangan=f"Auto dari PO {po.nomor_po}",
                    created_by=request.user,
                    received_by=request.user,
                    received_at=timezone.now(),
                )
                
                # Create InboundItems and update Product harga_beli
                for po_item in po.items.all():
                    InboundItem.objects.create(
                        inbound=inbound,
                        product=po_item.product,
                        quantity=po_item.quantity
                    )
                    
                    # Update Product harga_beli
                    po_item.product.harga_beli = po_item.harga_beli
                    po_item.product.last_purchase_price = po_item.harga_beli
                    po_item.product.last_purchase_date = timezone.now()
                    po_item.product.save()
                    
                    # Create Price History
                    PriceHistory.objects.create(
                        product=po_item.product,
                        purchase_order=po,
                        purchase_order_item=po_item,
                        price=po_item.harga_beli,
                        quantity=po_item.quantity,
                        subtotal=po_item.subtotal,
                        supplier=po.supplier,
                        purchase_date=po.tanggal_po,
                    )
                
                # Update PO status
                po.status = 'received'
                po.received_by = request.user
                po.received_at = timezone.now()
                po.save()
                
                # Create history
                create_po_history(po, 'receive', request.user, f'PO {po.nomor_po} di-receive, Inbound: {nomor_inbound}')
                
                messages.success(request, f"PO {po.nomor_po} berhasil di-receive. Inbound: {nomor_inbound}")
                return redirect('inventory:inbound_detail', pk=inbound.id)
                
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
            return redirect('purchasing:po_detail', pk=pk)
    
    # GET request - Show confirmation
    context = {
        'po': po,
    }
    return render(request, 'purchasing/po_receive_confirm.html', context)


@login_required
def po_print(request, pk):
    """Print Purchase Order"""
    po = get_object_or_404(PurchaseOrder.objects.select_related('supplier', 'created_by'), pk=pk)
    items = PurchaseOrderItem.objects.filter(po=po).select_related('product')
    
    context = {
        'po': po,
        'items': items,
    }
    
    html = render_to_string('purchasing/po_print.html', context)
    
    response = HttpResponse(html)
    response['Content-Type'] = 'text/html'
    return response


@login_required
def po_download_pdf(request, pk):
    """Download PO as PDF"""
    po = get_object_or_404(PurchaseOrder.objects.select_related('supplier', 'created_by'), pk=pk)
    items = PurchaseOrderItem.objects.filter(po=po).select_related('product')
    
    context = {
        'po': po,
        'items': items,
    }
    
    # Use simplified PDF template (without gradients and complex CSS)
    html = render_to_string('purchasing/po_print_pdf.html', context)
    
    # Convert HTML to PDF using xhtml2pdf
    from xhtml2pdf import pisa
    from io import BytesIO
    
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="PO_{po.nomor_po}.pdf"'
        return response
    
    return HttpResponse(f'Error generating PDF: {pdf.err}', status=500)


@login_required
def po_download_excel(request, pk):
    """Download PO as Excel"""
    po = get_object_or_404(PurchaseOrder.objects.select_related('supplier', 'created_by'), pk=pk)
    items = PurchaseOrderItem.objects.filter(po=po).select_related('product')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"
    
    # Header
    ws['C1'] = 'PURCHASE ORDER'
    ws['C1'].font = Font(size=16, bold=True)
    ws.merge_cells('C1:F1')
    ws['C1'].alignment = Alignment(horizontal='center')
    
    ws['C2'] = 'PT. ALFA ERP'
    ws['C2'].font = Font(size=12)
    ws.merge_cells('C2:F2')
    ws['C2'].alignment = Alignment(horizontal='center')
    
    # PO Information
    row = 4
    ws[f'A{row}'] = 'Dari:'
    ws[f'B{row}'] = 'PT. ALFA ERP'
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Ke:'
    supplier_info = po.supplier.nama_supplier
    if po.supplier.alamat:
        supplier_info += f'\n{po.supplier.alamat}'
    if po.supplier.kota:
        supplier_info += f'\n{po.supplier.kota}'
    ws[f'B{row}'] = supplier_info
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].alignment = Alignment(vertical='top')
    row += 1
    ws[f'A{row}'] = ''
    ws[f'B{row}'] = ''
    row += 1
    ws[f'A{row}'] = 'Nomor PO:'
    ws[f'B{row}'] = po.nomor_po
    row += 1
    ws[f'A{row}'] = 'Tanggal:'
    ws[f'B{row}'] = po.tanggal_po.strftime('%d %b %Y')
    row += 1
    ws[f'A{row}'] = 'Status:'
    ws[f'B{row}'] = po.get_status_display()
    
    # Table Header
    row += 2
    headers = ['No', 'Barcode', 'Nama Produk', 'Variant', 'Qty', 'Harga Beli', 'Subtotal']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Table Data
    row += 1
    for idx, item in enumerate(items, start=1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = item.product.barcode or item.product.sku
        ws[f'C{row}'] = item.product.nama_produk
        ws[f'D{row}'] = item.product.variant_produk or '-'
        ws[f'E{row}'] = item.quantity
        ws[f'F{row}'] = item.harga_beli
        ws[f'G{row}'] = item.subtotal
        row += 1
    
    # Total
    ws[f'F{row}'] = 'TOTAL:'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    ws[f'G{row}'] = po.total_amount
    ws[f'G{row}'].font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="PO_{po.nomor_po}.xlsx"'
    wb.save(response)
    return response


@login_required
def po_history(request, pk=None):
    """View PO History - Global or per PO"""
    if pk:
        # Per PO history
        po = get_object_or_404(PurchaseOrder, pk=pk)
        history = po.history.select_related('user').order_by('-timestamp')
        
        context = {
            'po': po,
            'history': history,
            'is_global': False,
        }
        return render(request, 'purchasing/po_history.html', context)
    else:
        # Global PO history - all changes
        history = PurchaseOrderHistory.objects.select_related('po', 'user').order_by('-timestamp')[:500]
        
        context = {
            'history': history,
            'is_global': True,
        }
        return render(request, 'purchasing/po_history.html', context)


@login_required
def price_history(request):
    """View Price History for a product"""
    product_sku = request.GET.get('sku', '')
    
    if not product_sku:
        messages.error(request, "SKU produk tidak ditemukan.")
        return redirect('purchasing:po_list')
    
    try:
        product = Product.objects.get(sku=product_sku)
    except Product.DoesNotExist:
        messages.error(request, f"Produk dengan SKU {product_sku} tidak ditemukan.")
        return redirect('purchasing:po_list')
    
    # Get price history
    price_histories = PriceHistory.objects.filter(product=product).select_related(
        'purchase_order', 'purchase_order__supplier'
    ).order_by('-purchase_date')
    
    context = {
        'product': product,
        'price_histories': price_histories,
    }
    return render(request, 'purchasing/price_history.html', context)


@login_required
def po_dashboard(request):
    """Dashboard Purchase Orders"""
    # Statistics
    total_po = PurchaseOrder.objects.count()
    draft_po = PurchaseOrder.objects.filter(status='draft').count()
    pending_po = PurchaseOrder.objects.filter(status='pending').count()
    received_po = PurchaseOrder.objects.filter(status='received').count()
    
    # Total amount
    total_amount = PurchaseOrder.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Recent POs
    recent_pos = PurchaseOrder.objects.select_related('supplier').order_by('-tanggal_po')[:10]
    
    # Top suppliers
    top_suppliers = PurchaseOrder.objects.values('supplier__nama_supplier').annotate(
        count=Count('id'),
        total=Sum('total_amount')
    ).order_by('-count')[:5]
    
    context = {
        'total_po': total_po,
        'draft_po': draft_po,
        'pending_po': pending_po,
        'received_po': received_po,
        'total_amount': total_amount,
        'recent_pos': recent_pos,
        'top_suppliers': top_suppliers,
    }
    
    return render(request, 'purchasing/po_dashboard.html', context)


@login_required
def po_data(request):
    """API for DataTable"""
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    search_type = request.GET.get('search_type', 'supplier')  # 'supplier' or 'nomor_po'
    supplier_id = request.GET.get('supplier_id', '')
    status_filter = request.GET.get('status', '')
    
    columns = ['id', 'nomor_po', 'tanggal_po', 'supplier', 'status', 'total_amount', 'created_by', 'received_by']
    
    queryset = PurchaseOrder.objects.select_related('supplier', 'created_by', 'received_by').all()
    
    # Filter by supplier_id if provided
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)
    
    # Filter by status if provided
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    
    total_count = queryset.count()
    
    # Search based on search_type
    if search_value:
        if search_type == 'nomor_po':
            # Search by nomor PO only
            queryset = queryset.filter(nomor_po__icontains=search_value)
        else:
            # Search by supplier name (default)
            queryset = queryset.filter(supplier__nama_supplier__icontains=search_value)
    
    filtered_count = queryset.count()
    queryset = queryset.order_by('-tanggal_po', '-id')[start:start+length]
    
    data = []
    for po in queryset:
        # Build action buttons based on status
        aksi_html = f'<a href="/purchasing/{po.id}/" class="btn btn-sm btn-info"><i class="bi bi-eye"></i> Detail</a> '
        
        if po.status == 'draft':
            # Draft: Edit, Hapus
            aksi_html += f'<a href="/purchasing/{po.id}/edit/" class="btn btn-sm btn-warning"><i class="bi bi-pencil"></i> Edit</a> '
            aksi_html += f'<a href="/purchasing/{po.id}/cancel/" class="btn btn-sm btn-danger" onclick="return confirm(\'Yakin ingin menghapus draft ini?\')"><i class="bi bi-trash"></i> Hapus</a> '
        elif po.status == 'pending':
            # Pending: Edit, Hapus
            aksi_html += f'<a href="/purchasing/{po.id}/edit/" class="btn btn-sm btn-warning"><i class="bi bi-pencil"></i> Edit</a> '
            aksi_html += f'<a href="/purchasing/{po.id}/cancel/" class="btn btn-sm btn-danger" onclick="return confirm(\'Yakin ingin menghapus PO ini?\')"><i class="bi bi-trash"></i> Hapus</a> '
        
        # Dropdown for print options
        aksi_html += f'''
        <div class="btn-group">
            <button type="button" class="btn btn-sm btn-secondary dropdown-toggle" data-bs-toggle="dropdown" aria-expanded="false">
                <i class="bi bi-three-dots-vertical"></i>
            </button>
            <ul class="dropdown-menu">
                <li><a class="dropdown-item" href="/purchasing/{po.id}/print/" target="_blank"><i class="bi bi-printer"></i> Print</a></li>
                <li><a class="dropdown-item" href="/purchasing/{po.id}/download-pdf/" target="_blank"><i class="bi bi-file-pdf"></i> Download PDF</a></li>
                <li><a class="dropdown-item" href="/purchasing/{po.id}/download-excel/" target="_blank"><i class="bi bi-file-excel"></i> Download Excel</a></li>
            </ul>
        </div>
        '''
        
        data.append({
            'id': po.id,
            'nomor_po': po.nomor_po,
            'tanggal_po': timezone.localtime(po.tanggal_po).strftime('%d-%m-%Y %H:%M'),
            'supplier': po.supplier.nama_supplier if po.supplier else '-',
            'status': po.get_status_display(),
            'total_amount': f"Rp {po.total_amount:,.0f}",
            'created_by': po.created_by.username if po.created_by else '-',
            'received_by': po.received_by.username if po.received_by else '-',
            'aksi': aksi_html
        })
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_count,
        'recordsFiltered': filtered_count,
        'data': data,
    })


# Helper views for product lookup
@login_required
def search_product(request):
    """Search product for PO"""
    q = request.GET.get('q', '').strip()
    
    products = Product.objects.filter(
        Q(sku__icontains=q) |
        Q(barcode__icontains=q) |
        Q(nama_produk__icontains=q)
    ).order_by('nama_produk')[:20]
    
    data = []
    for p in products:
        # Get stock quantity
        stock_qty = 0
        if hasattr(p, 'stock'):
            stock_qty = p.stock.quantity_ready_virtual
        
        # Get photo URL - return empty if photo not exists
        photo_url = ''
        if p.photo:
            try:
                # Check if file exists
                if p.photo.storage.exists(p.photo.name):
                    photo_url = p.photo.url
            except:
                pass
        
        data.append({
            'sku': p.sku,
            'barcode': p.barcode,
            'nama_produk': p.nama_produk,
            'variant_produk': p.variant_produk or '',
            'brand': p.brand or '',
            'harga_beli': float(p.harga_beli) if p.harga_beli else 0,
            'photo_url': photo_url,
            'stock_qty': stock_qty,
        })
    
    return JsonResponse(data, safe=False)


@login_required
def search_supplier(request):
    """Search supplier"""
    q = request.GET.get('q', '').strip()
    
    suppliers = Supplier.objects.filter(
        Q(nama_supplier__icontains=q) |
        Q(kota__icontains=q)
    ).order_by('nama_supplier')[:20]
    
    data = []
    for s in suppliers:
        data.append({
            'id': s.id,
            'nama_supplier': s.nama_supplier,
            'alamat': s.alamat or '',
            'kota': s.kota or '',
        })
    
    return JsonResponse(data, safe=False)


@login_required
def add_supplier(request):
    """Add new supplier"""
    if request.method == 'POST':
        nama = request.POST.get('nama_supplier', '').strip()
        if not nama:
            return JsonResponse({'error': 'Nama supplier wajib diisi'}, status=400)
        
        supplier = Supplier.objects.create(
            nama_supplier=nama,
            alamat=request.POST.get('alamat_supplier', ''),
            kota=request.POST.get('kota_supplier', ''),
        )
        
        return JsonResponse({
            'id': supplier.id,
            'nama_supplier': supplier.nama_supplier,
        })
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ==========================================
# PURCHASE LIST VIEWS (Goods Receipt)
# ==========================================

def get_next_purchase_id():
    """Generate next Purchase ID"""
    last = Purchase.objects.order_by('-id').first()
    if last:
        return last.id + 1
    return 1


@login_required
def po_detail_json(request, pk):
    """API to get PO detail in JSON format"""
    try:
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"po_detail_json called for PO ID: {pk}")
        
        po = get_object_or_404(
            PurchaseOrder.objects.select_related('supplier')
            .prefetch_related('items__product__stock'),
            id=pk
        )
        
        # Serialize PO data
        data = {
            'id': po.id,
            'nomor_po': po.nomor_po,
            'tanggal_po': po.tanggal_po.isoformat(),
            'supplier': po.supplier.nama_supplier if po.supplier else '',
            'status': po.status,
            'total_amount': float(po.total_amount),
            'items': []
        }
        
        # Serialize items
        for item in po.items.all():
            # Check if photo exists
            photo_url = ''
            if item.product.photo:
                try:
                    if item.product.photo.storage.exists(item.product.photo.name):
                        photo_url = item.product.photo.url
                except Exception as e:
                    logger.warning(f"Photo error for product {item.product.sku}: {str(e)}")
                    pass
            
            item_data = {
                'id': item.id,
                'product': {
                    'sku': item.product.sku,
                    'barcode': item.product.barcode,
                    'nama_produk': item.product.nama_produk,
                    'variant_produk': item.product.variant_produk,
                    'brand': item.product.brand,
                    'photo_url': photo_url,
                    'stock': {
                        'quantity_ready': item.product.stock.quantity_ready_virtual if item.product.stock else 0
                    }
                },
                'quantity': item.quantity,
                'harga_beli': float(item.harga_beli),
                'subtotal': float(item.subtotal)
            }
            data['items'].append(item_data)
        
        logger.info(f"po_detail_json returning {len(data['items'])} items")
        return JsonResponse(data)
    
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"po_detail_json error for PO ID {pk}: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def purchase_list(request):
    """List all Purchases"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_view'):
        raise PermissionDenied("You don't have permission to view Purchases.")
    
    # Check if mobile user agent
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = any(mobile in user_agent for mobile in ['mobile', 'android', 'iphone', 'ipad', 'ipod'])
    
    template = 'purchasing/mobile_purchase_list.html' if is_mobile else 'purchasing/purchase_list.html'
    return render(request, template)


@login_required
def purchase_list_api(request):
    """API endpoint for Purchase list with pagination and filtering"""
    
    try:
        # Get pagination parameters
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 25))
        
        # Get filter parameters
        purchase_number_filter = request.GET.get('purchase_number', '').strip()
        status_filter = request.GET.get('status', '').strip()
        supplier_filter = request.GET.get('supplier', '').strip()
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        
        # Build query
        queryset = Purchase.objects.select_related('supplier', 'po', 'created_by', 'received_by', 'verified_by').prefetch_related('items').order_by('-id')
        
        # Apply filters
        if purchase_number_filter:
            queryset = queryset.filter(nomor_purchase__icontains=purchase_number_filter)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if supplier_filter:
            queryset = queryset.filter(supplier__nama_supplier__icontains=supplier_filter)
        
        if date_from:
            queryset = queryset.filter(tanggal_purchase__gte=date_from)
        
        if date_to:
            # Add one day to include the end date
            from datetime import datetime, timedelta
            end_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            queryset = queryset.filter(tanggal_purchase__lt=end_date.strftime('%Y-%m-%d'))
        
        # Get total count before pagination
        total_count = queryset.count()
        
        # Paginate
        paginator = Paginator(queryset, page_size)
        page_obj = paginator.get_page(page)
        
        # Serialize data
        data = []
        for purchase in page_obj:
            # Calculate total items and total quantity
            items = purchase.items.all()
            total_items = items.count()  # Jumlah produk berbeda
            total_qty = sum(item.quantity for item in items)  # Total quantity semua produk
            
            data.append({
                'id': purchase.id,
                'nomor_purchase': purchase.nomor_purchase,
                'tanggal_purchase': purchase.tanggal_purchase.strftime('%Y-%m-%d') if purchase.tanggal_purchase else '',
                'supplier': purchase.supplier.nama_supplier if purchase.supplier else '-',
                'nomor_po': purchase.po.nomor_po if purchase.po else '-',
                'status': purchase.get_status_display(),
                'status_value': purchase.status,
                'total_amount': purchase.total_amount,
                'total_amount_formatted': f'{purchase.total_amount:,}',
                'total_items': total_items,
                'total_qty': total_qty,
                'created_by': purchase.created_by.username if purchase.created_by else '-',
                'created_at': purchase.created_at.strftime('%Y-%m-%d %H:%M') if purchase.created_at else '',
                'verified_by': purchase.verified_by.username if purchase.verified_by else None,
                'verified_at': purchase.verified_at.strftime('%Y-%m-%d %H:%M') if purchase.verified_at else None,
            })
        
        # Calculate statistics from filtered queryset
        stats = {
            'total_purchases': queryset.count(),
            'draft_purchases': queryset.filter(status='draft').count(),
            'pending_purchases': queryset.filter(status='pending').count(),
            'received_purchases': queryset.filter(status='received').count(),
            'verified_purchases': queryset.filter(status='verified').count(),
            'cancelled_purchases': queryset.filter(status='cancelled').count(),
            'total_amount': queryset.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        }
        
        return JsonResponse({
            'success': True,
            'data': data,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_count': total_count,
                'total_pages': paginator.num_pages,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
            },
            'stats': stats,
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def purchase_create(request, po_id=None):
    """Create new Purchase (with or without PO)"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_warehouse'):
        raise PermissionDenied("You don't have permission to create purchase.")
    
    from datetime import datetime
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get purchase_id from form (draft yang sudah dibuat di GET request)
                purchase_id = request.POST.get('purchase_id')
                if not purchase_id:
                    messages.error(request, 'Purchase ID tidak ditemukan!')
                    return redirect('purchasing:purchase_list')
                
                # Get existing draft purchase
                purchase = get_object_or_404(Purchase, id=purchase_id, status='draft')
                
                # Get form data
                supplier_id = request.POST.get('supplier_id')
                tanggal_purchase = request.POST.get('tanggal_purchase', '').strip()
                notes = request.POST.get('notes', '')
                po_id_from_form = request.POST.get('po_id', po_id)
                
                # Validate required fields
                if not supplier_id:
                    return JsonResponse({
                        'success': False,
                        'error': 'Supplier wajib diisi!'
                    })
                
                if not tanggal_purchase:
                    return JsonResponse({
                        'success': False,
                        'error': 'Tanggal purchase (invoice) wajib diisi!'
                    })
                
                # Check if there are any products
                skus = request.POST.getlist('produk_sku[]')
                if not skus or len(skus) == 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'Minimal 1 produk harus ditambahkan!'
                    })
                
                # Update purchase
                purchase.supplier_id = supplier_id
                purchase.tanggal_purchase = tanggal_purchase
                purchase.notes = notes
                purchase.po_id = po_id_from_form
                purchase.status = 'pending'  # Change status to 'pending'
                purchase.save()
                
                # Delete old items
                purchase.items.all().delete()
                
                # Add items from form
                skus = request.POST.getlist('produk_sku[]')
                qtys = request.POST.getlist('produk_qty[]')
                hargas = request.POST.getlist('produk_harga_beli[]')
                
                for sku, qty, harga in zip(skus, qtys, hargas):
                    if sku and qty and harga:
                        product = get_object_or_404(Product, sku=sku)
                        # Parse harga: remove dots (thousand separators) and convert to int
                        harga_clean = str(harga).replace('.', '').replace(',', '')
                        PurchaseItem.objects.create(
                            purchase=purchase,
                            product=product,
                            quantity=int(qty),
                            harga_beli=int(harga_clean),
                        )
                
                # Update PO status to 'received' if purchase was created from PO
                if po_id_from_form:
                    po = get_object_or_404(PurchaseOrder, id=po_id_from_form)
                    po.status = 'received'
                    po.received_by = request.user
                    po.received_at = timezone.now()
                    po.save()
                
                # Return JSON response for AJAX submission
                return JsonResponse({
                    'success': True,
                    'purchase_id': purchase.id,
                    'message': f'Purchase {purchase.nomor_purchase} berhasil dibuat!'
                })
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    # GET request - Create new draft purchase and redirect to edit
    suppliers = Supplier.objects.all().order_by('nama_supplier')
    
    # Get po_id from URL parameter or query string
    po_id_from_query = request.GET.get('po_id')
    if po_id_from_query:
        po_id = int(po_id_from_query)
    
    # Create new draft purchase (without supplier - can be null for draft)
    # User must select supplier before saving
    # Nomor purchase akan di-generate otomatis saat status berubah dari draft
    draft_purchase = Purchase.objects.create(
        nomor_purchase=None,  # No nomor purchase for draft
        supplier=None,  # No supplier for draft
        created_by=request.user,
        status='draft',
    )
    
    # If creating from PO, get PO data and populate draft
    if po_id:
        po = get_object_or_404(PurchaseOrder.objects.select_related('supplier').prefetch_related('items__product'), id=po_id)
        draft_purchase.supplier = po.supplier
        draft_purchase.po = po
        # Tanggal purchase - Leave blank, user must fill manually (tanggal invoice)
        # draft_purchase.tanggal_purchase = po.tanggal_po
        draft_purchase.save()
        
        # Copy items from PO to draft
        for item in po.items.all():
            PurchaseItem.objects.create(
                purchase=draft_purchase,
                product=item.product,
                quantity=item.quantity,
                harga_beli=item.harga_beli,
            )
    
    # Redirect to edit page (not render create template)
    return redirect('purchasing:purchase_edit', purchase_id=draft_purchase.id)


@login_required
def purchase_detail(request, purchase_id):
    """View Purchase detail"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_view'):
        raise PermissionDenied("You don't have permission to view Purchases.")
    
    purchase = get_object_or_404(
        Purchase.objects.select_related('supplier', 'po', 'created_by', 'received_by')
        .prefetch_related('items__product'),
        id=purchase_id
    )
    
    items = purchase.items.all()
    
    context = {
        'purchase': purchase,
        'items': items,
    }
    
    # Check if mobile user agent
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = any(mobile in user_agent for mobile in ['mobile', 'android', 'iphone', 'ipad', 'ipod'])
    
    template = 'purchasing/mobile_purchase_detail.html' if is_mobile else 'purchasing/purchase_detail.html'
    return render(request, template, context)


@login_required
def purchase_print(request, purchase_id):
    """Print Purchase"""
    purchase = get_object_or_404(Purchase.objects.select_related('supplier', 'created_by', 'received_by'), id=purchase_id)
    items = PurchaseItem.objects.filter(purchase=purchase).select_related('product')
    
    context = {
        'purchase': purchase,
        'items': items,
    }
    
    html = render_to_string('purchasing/purchase_print.html', context)
    
    response = HttpResponse(html)
    response['Content-Type'] = 'text/html'
    return response


@login_required
def purchase_download_pdf(request, purchase_id):
    """Download Purchase as PDF"""
    purchase = get_object_or_404(Purchase.objects.select_related('supplier', 'created_by', 'received_by'), id=purchase_id)
    items = PurchaseItem.objects.filter(purchase=purchase).select_related('product')
    
    context = {
        'purchase': purchase,
        'items': items,
    }
    
    # Use simplified PDF template (without gradients and complex CSS)
    html = render_to_string('purchasing/purchase_print_pdf.html', context)
    
    # Convert HTML to PDF using xhtml2pdf
    from xhtml2pdf import pisa
    from io import BytesIO
    
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Purchase_{purchase.nomor_purchase}.pdf"'
        return response
    
    return HttpResponse(f'Error generating PDF: {pdf.err}', status=500)


@login_required
def purchase_download_excel(request, purchase_id):
    """Download Purchase as Excel"""
    purchase = get_object_or_404(Purchase.objects.select_related('supplier', 'created_by', 'received_by'), id=purchase_id)
    items = PurchaseItem.objects.filter(purchase=purchase).select_related('product')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase"
    
    # Header
    ws['C1'] = 'PURCHASE'
    ws['C1'].font = Font(size=16, bold=True)
    ws.merge_cells('C1:F1')
    ws['C1'].alignment = Alignment(horizontal='center')
    
    ws['C2'] = 'PT. ALFA ERP'
    ws['C2'].font = Font(size=12)
    ws.merge_cells('C2:F2')
    ws['C2'].alignment = Alignment(horizontal='center')
    
    # Purchase Information
    row = 4
    ws[f'A{row}'] = 'Dari:'
    ws[f'B{row}'] = 'PT. ALFA ERP'
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Ke:'
    supplier_info = purchase.supplier.nama_supplier
    if purchase.supplier.alamat:
        supplier_info += f'\n{purchase.supplier.alamat}'
    if purchase.supplier.kota:
        supplier_info += f'\n{purchase.supplier.kota}'
    ws[f'B{row}'] = supplier_info
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].alignment = Alignment(vertical='top')
    row += 1
    ws[f'A{row}'] = ''
    ws[f'B{row}'] = ''
    row += 1
    ws[f'A{row}'] = 'Nomor Purchase:'
    ws[f'B{row}'] = purchase.nomor_purchase
    row += 1
    ws[f'A{row}'] = 'Tanggal:'
    ws[f'B{row}'] = purchase.tanggal_purchase.strftime('%d %b %Y')
    row += 1
    ws[f'A{row}'] = 'Status:'
    ws[f'B{row}'] = purchase.get_status_display()
    if purchase.po:
        row += 1
        ws[f'A{row}'] = 'PO Reference:'
        ws[f'B{row}'] = purchase.po.nomor_po
    
    # Table Header
    row += 2
    headers = ['No', 'Barcode', 'Nama Produk', 'Variant', 'Qty', 'Harga Beli', 'Subtotal']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Table Data
    row += 1
    for idx, item in enumerate(items, start=1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = item.product.barcode or item.product.sku
        ws[f'C{row}'] = item.product.nama_produk
        ws[f'D{row}'] = item.product.variant_produk or '-'
        ws[f'E{row}'] = item.quantity
        ws[f'F{row}'] = item.harga_beli
        ws[f'G{row}'] = item.subtotal
        row += 1
    
    # Total
    ws[f'F{row}'] = 'TOTAL:'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    ws[f'G{row}'] = purchase.total_amount
    ws[f'G{row}'].font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Purchase_{purchase.nomor_purchase}.xlsx"'
    wb.save(response)
    return response


@login_required
def purchase_auto_save(request, purchase_id):
    """Auto-save Purchase changes via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    
    try:
        purchase = get_object_or_404(Purchase, id=purchase_id)
        
        # Only allow auto-save for draft, pending, and received status (for verification mode)
        if purchase.status not in ['draft', 'pending', 'received']:
            return JsonResponse({'success': False, 'error': 'Cannot auto-save cancelled or verified purchase'})
        
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Auto-save purchase {purchase_id}: supplier={request.POST.get('supplier_id')}, items={len(request.POST.getlist('produk_sku[]'))}")
        
        with transaction.atomic():
            # Update Purchase fields
            purchase.supplier_id = request.POST.get('supplier_id', purchase.supplier_id)
            
            # Tanggal purchase - Abaikan jika kosong (biarkan user isi manual saat save)
            tanggal_purchase = request.POST.get('tanggal_purchase', '').strip()
            if tanggal_purchase:
                purchase.tanggal_purchase = tanggal_purchase
            # If empty, leave as is (don't update)
            
            purchase.notes = request.POST.get('notes', purchase.notes)
            purchase.po_id = request.POST.get('po_id', purchase.po_id)
            purchase.save()
            
            # Update items
            skus = request.POST.getlist('produk_sku[]')
            qtys = request.POST.getlist('produk_qty[]')
            hargas = request.POST.getlist('produk_harga_beli[]')
            
            logger.info(f"Auto-save items: skus={skus}, qtys={qtys}, hargas={hargas}")
            
            # Get current items
            old_items = {item.product.sku: item for item in purchase.items.all()}
            current_skus = set()
            
            for sku, qty_str, harga_str in zip(skus, qtys, hargas):
                logger.info(f"Processing item: sku={sku}, qty={qty_str}, harga={harga_str}")
                if not (sku and qty_str and harga_str):
                    logger.warning(f"Skipping item: sku={sku}, qty={qty_str}, harga={harga_str} (empty)")
                    continue
                
                qty = int(qty_str)
                # Clean harga: remove thousand separators (.) and convert to int
                harga_clean = str(harga_str).replace('.', '').replace(',', '')
                harga = int(harga_clean) if harga_clean else 0
                current_skus.add(sku)
                
                product = Product.objects.filter(sku=sku).first()
                if not product:
                    continue
                
                if sku in old_items:
                    # Update existing item
                    old_item = old_items[sku]
                    old_item.quantity = qty
                    old_item.harga_beli = harga
                    old_item.save(update_fields=['quantity', 'harga_beli', 'subtotal'])
                    logger.info(f"Updated item: {sku}, qty={qty}, harga={harga}")
                else:
                    # Create new item
                    PurchaseItem.objects.create(
                        purchase=purchase,
                        product=product,
                        quantity=qty,
                        harga_beli=harga
                    )
                    logger.info(f"Created new item: {sku}, qty={qty}, harga={harga}")
            
            # Delete items that are not in form anymore
            skus_to_delete = set(old_items.keys()) - current_skus
            for sku in skus_to_delete:
                old_items[sku].delete()
            
            logger.info(f"Auto-save purchase {purchase_id} completed: {len(current_skus)} items saved")
            return JsonResponse({'success': True, 'message': 'Purchase berhasil disimpan otomatis'})
    
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Auto-save purchase {purchase_id} failed: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def purchase_edit(request, purchase_id):
    """Edit Purchase"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_warehouse'):
        raise PermissionDenied("You don't have permission to edit purchase.")
    
    purchase = get_object_or_404(
        Purchase.objects.select_related('supplier', 'po')
        .prefetch_related('items__product'),
        id=purchase_id
    )
    
    # Check if purchase is already verified (cannot edit verified purchases)
    if purchase.status == 'verified':
        messages.error(request, "Purchase yang sudah verified tidak dapat di-edit lagi. Silakan hubungi Finance untuk perubahan.")
        return redirect('purchasing:purchase_detail', purchase_id=purchase.id)
    
    # Check if this is verification mode
    is_verification_mode = request.GET.get('mode') == 'verification'
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Update Purchase
                supplier_id = request.POST.get('supplier_id')
                tanggal_purchase = request.POST.get('tanggal_purchase', '').strip()
                notes = request.POST.get('notes', '')
                
                # Validate supplier is required before saving
                if not supplier_id:
                    messages.error(request, "Supplier wajib diisi sebelum menyimpan Purchase!")
                    items = purchase.items.all()
                    suppliers = Supplier.objects.all().order_by('nama_supplier')
                    context = {
                        'purchase': purchase,
                        'items': items,
                        'suppliers': suppliers,
                        'is_verification_mode': is_verification_mode,
                    }
                    return render(request, 'purchasing/purchase_edit.html', context)
                
                # Validate tanggal purchase is required before saving
                if not tanggal_purchase:
                    messages.error(request, "Tanggal purchase (invoice) wajib diisi sebelum menyimpan Purchase!")
                    items = purchase.items.all()
                    suppliers = Supplier.objects.all().order_by('nama_supplier')
                    context = {
                        'purchase': purchase,
                        'items': items,
                        'suppliers': suppliers,
                        'is_verification_mode': is_verification_mode,
                    }
                    return render(request, 'purchasing/purchase_edit.html', context)
                
                # Check if there are any products
                skus = request.POST.getlist('produk_sku[]')
                if not skus or len(skus) == 0:
                    messages.error(request, "Minimal 1 produk harus ditambahkan sebelum menyimpan Purchase!")
                    items = purchase.items.all()
                    suppliers = Supplier.objects.all().order_by('nama_supplier')
                    context = {
                        'purchase': purchase,
                        'items': items,
                        'suppliers': suppliers,
                        'is_verification_mode': is_verification_mode,
                    }
                    return render(request, 'purchasing/purchase_edit.html', context)
                
                purchase.supplier_id = supplier_id
                purchase.tanggal_purchase = tanggal_purchase
                purchase.notes = notes
                
                # If status is 'draft', change to 'pending' after save
                if purchase.status == 'draft':
                    purchase.status = 'pending'
                
                purchase.save()
                
                # Update items
                skus = request.POST.getlist('produk_sku[]')
                qtys = request.POST.getlist('produk_qty[]')
                hargas = request.POST.getlist('produk_harga_beli[]')
                
                # Delete old items
                purchase.items.all().delete()
                
                # Create new items
                for sku, qty, harga in zip(skus, qtys, hargas):
                    if sku and qty and harga:
                        product = get_object_or_404(Product, sku=sku)
                        # Parse harga: remove dots (thousand separators) and convert to int
                        harga_clean = str(harga).replace('.', '').replace(',', '')
                        PurchaseItem.objects.create(
                            purchase=purchase,
                            product=product,
                            quantity=int(qty),
                            harga_beli=int(harga_clean),
                        )
                
                if purchase.status == 'pending':
                    messages.success(request, f'Purchase {purchase.nomor_purchase} berhasil disimpan!')
                else:
                    messages.success(request, f'Purchase {purchase.nomor_purchase} berhasil diupdate!')
                
                # If verification mode, redirect back to list
                if is_verification_mode:
                    return redirect('purchasing:purchase_list')
                else:
                    return redirect('purchasing:purchase_detail', purchase_id=purchase.id)
        
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('purchasing:purchase_list')
    
    # GET request
    suppliers = Supplier.objects.all().order_by('nama_supplier')
    items = purchase.items.all()
    
    context = {
        'purchase': purchase,
        'suppliers': suppliers,
        'items': items,
        'is_verification_mode': is_verification_mode,
    }
    
    # Check if mobile user agent
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = any(mobile in user_agent for mobile in ['mobile', 'android', 'iphone', 'ipad', 'ipod'])
    
    template = 'purchasing/mobile_purchase_edit.html' if is_mobile else 'purchasing/purchase_edit.html'
    return render(request, template, context)


@login_required
def purchase_verify_list(request):
    """Purchase Verify List - List purchases yang sudah received dan perlu di-verify"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_finance'):
        raise PermissionDenied("You don't have permission to view purchase verify list.")
    
    # Base queryset
    base_qs = Purchase.objects.select_related('supplier', 'po').prefetch_related('items__product').order_by('-tanggal_purchase')

    # Tab filter: default 'pending' shows received (not yet verified); 'verified' shows verified
    tab = request.GET.get('tab', 'pending').strip().lower()
    if tab == 'verified':
        purchases = base_qs.filter(status='verified')
    else:
        # default
        purchases = base_qs.filter(status='received')
    
    # Get filter parameters
    supplier_filter = request.GET.get('supplier', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search = request.GET.get('search', '').strip()
    
    # Apply filters
    if supplier_filter:
        purchases = purchases.filter(supplier__id=supplier_filter)
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            purchases = purchases.filter(tanggal_purchase__gte=date_from_obj.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            purchases = purchases.filter(tanggal_purchase__lte=date_to_obj.date())
        except ValueError:
            pass
    
    if search:
        purchases = purchases.filter(
            models.Q(nomor_purchase__icontains=search) |
            models.Q(supplier__nama_supplier__icontains=search) |
            models.Q(notes__icontains=search)
        )
    
    # Get summary statistics AFTER all filters are applied
    total_count = purchases.count()
    total_amount = purchases.aggregate(total=models.Sum('total_amount'))['total'] or 0
    total_received = total_count  # For calculating average
    
    # Get unique suppliers for filter dropdown (from base queryset, not filtered)
    from inventory.models import Supplier
    if tab == 'verified':
        supplier_base_qs = base_qs.filter(status='verified')
    else:
        supplier_base_qs = base_qs.filter(status='received')
    
    suppliers = Supplier.objects.filter(
        id__in=supplier_base_qs.values_list('supplier_id', flat=True).distinct()
    ).order_by('nama_supplier')
    
    context = {
        'purchases': purchases,
        'suppliers': suppliers,
        'total_count': total_count,
        'total_amount': total_amount,
        'total_received': total_received,
        'stats_title_1': 'Total Verified' if tab == 'verified' else 'Total Received',
        'stats_title_4': 'Verified' if tab == 'verified' else 'Pending Verify',
        'filter_data': {
            'supplier': supplier_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
            'tab': tab,
        }
    }
    
    return render(request, 'purchasing/purchase_verify_list.html', context)


@login_required
def purchase_verify(request, purchase_id):
    """Verify Purchase (mark as verified)"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_finance'):
        raise PermissionDenied("You don't have permission to verify purchase.")
    
    import logging
    logger = logging.getLogger(__name__)
    
    logger.info(f"[VERIFY] Request: {request.method} for purchase_id={purchase_id}")
    
    purchase = get_object_or_404(
        Purchase.objects.select_related('supplier', 'po')
        .prefetch_related('items__product'),
        id=purchase_id
    )
    
    logger.info(f"[VERIFY] Purchase found: {purchase.nomor_purchase}, status={purchase.status}")
    
    # Only received purchases can be verified
    if purchase.status != 'received':
        logger.warning(f"[VERIFY] Invalid status: {purchase.status}, expected 'received'")
        messages.error(request, 'Hanya purchase dengan status "Received" yang bisa di-verify!')
        return redirect('purchasing:purchase_list')
    
    if request.method == 'POST':
        # POST: Process verification
        logger.info(f"[VERIFY] Processing POST verification for {purchase.nomor_purchase}")
        try:
            # Get form data
            due_date_str = request.POST.get('due_date', '').strip()
            transaction_type = request.POST.get('transaction_type', '').strip()
            has_fp_value = request.POST.get('has_fp', 'Y').strip().upper()
            has_fp = (has_fp_value == 'Y')
            
            # Validate due_date
            if not due_date_str:
                messages.error(request, 'Tanggal jatuh tempo wajib diisi!')
                return redirect('purchasing:purchase_verify', purchase_id=purchase.id)
            
            # Validate transaction_type
            if not transaction_type:
                messages.error(request, 'Tipe transaksi wajib dipilih!')
                return redirect('purchasing:purchase_verify', purchase_id=purchase.id)
            
            # Parse due_date
            from datetime import datetime
            try:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
                due_date = timezone.make_aware(due_date)
            except ValueError:
                messages.error(request, 'Format tanggal tidak valid!')
                return redirect('purchasing:purchase_verify', purchase_id=purchase.id)
            
            with transaction.atomic():
                # Update status to verified
                old_status = purchase.status
                purchase.status = 'verified'
                purchase.verified_at = timezone.now()
                purchase.verified_by = request.user
                purchase.has_tax_invoice = has_fp
                purchase.save()
                logger.info(f"[VERIFY] Status updated: {old_status} -> {purchase.status}")
                
                # Create journal entry for purchase verify
                try:
                    from .accounting import create_purchase_verify_journal_entry
                    journal_entry = create_purchase_verify_journal_entry(purchase, request.user)
                    logger.info(f"[VERIFY] Journal entry created: {journal_entry.entry_number}")
                    messages.success(request, f'Purchase berhasil di-verify! Journal entry {journal_entry.entry_number} telah dibuat.')
                except Exception as e:
                    logger.error(f"[VERIFY] Error creating journal entry: {str(e)}")
                    messages.warning(request, f'Purchase berhasil di-verify, tapi ada error saat membuat journal entry: {str(e)}')
                
                # Update product harga_beli (last purchase price) and calculate HPP for each item
                from inventory.models import Stock
                
                for item in purchase.items.all():
                    product = item.product
                    product.harga_beli = item.harga_beli
                    product.last_purchase_price = item.harga_beli
                    product.last_purchase_date = timezone.now()
                    
                    # Create PriceHistory record
                    PriceHistory.objects.create(
                        product=product,
                        purchase=purchase,
                        purchase_item=item,
                        price=item.harga_beli,
                        quantity=item.quantity,
                        subtotal=item.subtotal,
                        supplier=purchase.supplier,
                        purchase_date=purchase.tanggal_purchase,
                    )
                    
                    # Calculate HPP (Weighted Average) - Simple Method
                    # Get current stock quantity
                    stock = Stock.objects.filter(product=product).first()
                    current_stock_qty = stock.quantity if stock else 0
                    
                    # Current HPP and purchase quantity
                    hpp_lama = float(product.hpp or 0)
                    qty_beli = item.quantity
                    harga_beli_baru = float(item.harga_beli)
                    
                    # Calculate new HPP
                    if hpp_lama == 0:
                        # If HPP is 0, use current purchase price
                        hpp_baru = harga_beli_baru
                        logger.info(f"[VERIFY] HPP for {product.sku} is 0, set to purchase price: Rp {hpp_baru:,.0f}")
                    else:
                        # Weighted average: (HPP_lama × Stock_Qty + Harga_Beli_Baru × Qty_Beli) / (Stock_Qty + Qty_Beli)
                        # Note: We use stock qty BEFORE the purchase was received
                        stock_qty_before = current_stock_qty - qty_beli
                        if stock_qty_before < 0:
                            stock_qty_before = 0
                        
                        total_cost = (hpp_lama * stock_qty_before) + (harga_beli_baru * qty_beli)
                        total_qty = stock_qty_before + qty_beli
                        
                        if total_qty > 0:
                            hpp_baru = total_cost / total_qty
                            logger.info(f"[VERIFY] Calculated HPP for {product.sku}: "
                                      f"(Rp {hpp_lama:,.0f} × {stock_qty_before} + Rp {harga_beli_baru:,.0f} × {qty_beli}) / {total_qty} "
                                      f"= Rp {hpp_baru:,.0f}")
                        else:
                            hpp_baru = harga_beli_baru
                            logger.info(f"[VERIFY] Total qty is 0, set HPP to purchase price: Rp {hpp_baru:,.0f}")
                    
                    product.hpp = round(hpp_baru, 2)
                    product.save()
                    logger.info(f"[VERIFY] Updated {product.sku}: Harga Beli = Rp {item.harga_beli:,.0f}, HPP = Rp {product.hpp:,.0f}")
                    
                logger.info(f"[VERIFY] All product prices and HPP updated")
            
            # Create PurchasePayment with user input
            from purchasing.models import PurchasePayment, PurchaseTaxInvoice
            
            payment = PurchasePayment.objects.create(
                purchase=purchase,
                supplier=purchase.supplier,
                total_amount=purchase.total_amount,
                discount=0,  # Default no discount
                paid_amount=0,
                remaining_amount=purchase.total_amount,
                due_date=due_date,
                transaction_type=transaction_type,
                status='unpaid',
            )
            logger.info(f"[VERIFY] PurchasePayment created: ID={payment.id}, amount={payment.total_amount}, due_date={due_date}, transaction_type={transaction_type}")
            
            # Create PurchaseTaxInvoice only if has_fp is True
            if has_fp:
                # Calculate tax: invoice_amount = subtotal + tax
                tax_rate = 11  # 11% PPN
                subtotal = int(purchase.total_amount / (1 + tax_rate / 100))
                tax_amount = purchase.total_amount - subtotal
                
                tax_invoice, _created = PurchaseTaxInvoice.objects.get_or_create(
                    purchase=purchase,
                    supplier=purchase.supplier,
                    defaults={
                        'invoice_number': None,
                        'invoice_amount': purchase.total_amount,
                        'discount': 0,
                        'tax_rate': tax_rate,
                        'subtotal': subtotal,
                        'tax_amount': tax_amount,
                        'status': 'pending',
                    }
                )
                logger.info(f"[VERIFY] PurchaseTaxInvoice ensured: ID={tax_invoice.id}, amount={tax_invoice.invoice_amount}")
            
            logger.info(f"[VERIFY] Verification completed successfully for {purchase.nomor_purchase}")
            if has_fp:
                messages.success(request, f'Purchase {purchase.nomor_purchase} berhasil di-verify! Payment dibuat (Due: {due_date.strftime("%d %b %Y")}, Type: {transaction_type}) dan Tax Invoice disiapkan.')
            else:
                messages.success(request, f'Purchase {purchase.nomor_purchase} berhasil di-verify! Payment dibuat (Due: {due_date.strftime("%d %b %Y")}, Type: {transaction_type}). Tidak ada Tax Invoice (FP = Tidak).')
            return redirect('purchasing:purchase_list')
        
        except Exception as e:
            import traceback
            logger.error(f"[VERIFY] Error: {str(e)}")
            logger.error(f"[VERIFY] Traceback: {traceback.format_exc()}")
            messages.error(request, f'Error: {str(e)}')
            return redirect('purchasing:purchase_list')
    
    # GET: Render verification page
    logger.info(f"[VERIFY] Rendering GET page for {purchase.nomor_purchase}")
    suppliers = Supplier.objects.all().order_by('nama_supplier')
    items = purchase.items.all()
    
    context = {
        'purchase': purchase,
        'items': items,
        'suppliers': suppliers,
        'has_tax_invoice': purchase.has_tax_invoice if hasattr(purchase, 'has_tax_invoice') else True,
    }
    
    return render(request, 'purchasing/purchase_verify.html', context)


@login_required
def purchase_receive(request, purchase_id):
    """Receive Purchase (update stock)"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_warehouse'):
        raise PermissionDenied("You don't have permission to receive purchase.")
    
    purchase = get_object_or_404(Purchase, id=purchase_id)
    
    # Draft cannot be received
    if purchase.status == 'draft':
        messages.error(request, 'Draft tidak bisa di-receive! Silakan simpan purchase terlebih dahulu.')
        return redirect('purchasing:purchase_list')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Update Purchase status
                purchase.status = 'received'
                purchase.tanggal_received = timezone.now()
                purchase.received_by = request.user
                purchase.save()
                
                # Update stock for each item
                for item in purchase.items.all():
                    product = item.product
                    
                    # Get or create stock with row locking
                    from inventory.models import Stock, StockCardEntry
                    stock, stock_created = Stock.objects.select_for_update().get_or_create(
                        product=product,
                        defaults={'sku': product.sku, 'quantity': 0, 'quantity_locked': 0, 'quantity_putaway': 0}
                    )
                    
                    qty_awal = stock.quantity
                    
                    # Add quantity to stock (ready for putaway)
                    stock.quantity += item.quantity  # Langsung ke inventory stock
                    stock.quantity_putaway += item.quantity  # Stock belum di-putaway
                    stock.save(update_fields=['quantity', 'quantity_putaway'])
                    
                    # Create stock card entry for audit trail
                    StockCardEntry.objects.create(
                        product=product,
                        tipe_pergerakan='purchase',
                        qty=item.quantity,
                        qty_awal=qty_awal,
                        qty_akhir=stock.quantity,
                        notes=f'Purchase {purchase.nomor_purchase}' + (' (Stok Baru)' if stock_created else ''),
                        user=request.user,
                        reference=purchase
                    )
                    
                    # Note: Harga beli akan diupdate saat purchase di-verify (purchase_verify)
                
                messages.success(request, f'Purchase {purchase.nomor_purchase} berhasil diterima! Stock telah diupdate dan siap untuk putaway. Silakan verify purchase untuk melanjutkan ke payment.')
                return redirect('purchasing:purchase_detail', purchase_id=purchase.id)
        
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('purchasing:purchase_list')
    
    return redirect('purchasing:purchase_list')


@login_required
def purchase_delete(request, purchase_id):
    """Delete Purchase (AJAX)"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_warehouse'):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    if request.method == 'POST':
        try:
            purchase = get_object_or_404(Purchase, id=purchase_id)
            
            # Only allow delete for draft and pending status
            if purchase.status not in ['draft', 'pending']:
                return JsonResponse({'success': False, 'error': 'Hanya draft dan pending yang bisa dihapus'})
            
            purchase.delete()
            return JsonResponse({'success': True, 'message': 'Purchase berhasil dihapus'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@login_required
def purchase_cancel(request, purchase_id):
    """Cancel Purchase (or Delete Draft)"""
    from django.core.exceptions import PermissionDenied
    
    # Check permission
    if not request.user.has_perm('purchasing.purchase_warehouse'):
        raise PermissionDenied("You don't have permission to delete purchase.")
    
    purchase = get_object_or_404(Purchase, id=purchase_id)
    
    # If status is 'draft', delete it instead of cancelling
    if purchase.status == 'draft':
        purchase.delete()
        messages.success(request, f'Draft {purchase.nomor_purchase} berhasil dihapus!')
        return redirect('purchasing:purchase_list')
    
    # For non-draft purchases, require POST method
    if request.method == 'POST':
        try:
            purchase.status = 'cancelled'
            purchase.save()
            messages.success(request, f'Purchase {purchase.nomor_purchase} berhasil dibatalkan!')
            return redirect('purchasing:purchase_list')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('purchasing:purchase_list')
    
    return redirect('purchasing:purchase_list')

