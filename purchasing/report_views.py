from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q, Sum, Count
from django.template.loader import render_to_string
from datetime import datetime, timedelta
from purchasing.models import Purchase, PurchaseItem
from inventory.models import Supplier


@login_required
def purchase_report(request):
    """Purchase Report Page"""
    suppliers = Supplier.objects.all().order_by('nama_supplier')
    
    # Get default date range (last 30 days)
    today = datetime.now().date()
    date_from_default = today - timedelta(days=30)
    date_to_default = today
    
    context = {
        'suppliers': suppliers,
        'date_from_default': date_from_default.strftime('%Y-%m-%d'),
        'date_to_default': date_to_default.strftime('%Y-%m-%d'),
    }
    
    return render(request, 'purchasing/purchase_report.html', context)


@login_required
def purchase_report_preview(request):
    """Preview purchase report data"""
    from django.http import JsonResponse
    from purchasing.models import PurchasePayment, PurchaseTaxInvoice
    
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    supplier_id = request.GET.get('supplier_id', '').strip()
    report_type = request.GET.get('report_type', 'detail')
    
    # Build query
    queryset = Purchase.objects.select_related('supplier', 'created_by', 'received_by').prefetch_related(
        'items__product',
        'payment',
        'tax_invoice'
    ).order_by('-tanggal_purchase')
    
    # Apply filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(tanggal_purchase__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            end_of_day = date_to_obj + timedelta(days=1)
            queryset = queryset.filter(tanggal_purchase__lt=end_of_day)
        except:
            pass
    
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)
    
    purchases = queryset.all()
    
    # Calculate statistics
    total_purchases = purchases.count()
    total_amount = sum(p.total_amount for p in purchases)
    
    # Calculate payment statistics
    total_paid = 0
    total_unpaid = 0
    for p in purchases:
        try:
            payment = p.payment.first()
            if payment:
                total_paid += payment.paid_amount or 0
                total_unpaid += payment.remaining_amount or 0
        except:
            pass
    
    # Calculate tax invoice statistics
    total_invoice_amount = 0
    pending_invoices = 0
    for p in purchases:
        try:
            tax_invoice = p.tax_invoice.first()
            if tax_invoice:
                total_invoice_amount += tax_invoice.invoice_amount or 0
                if tax_invoice.status == 'pending':
                    pending_invoices += 1
        except:
            pass
    
    # Render preview HTML
    html = render_to_string('purchasing/purchase_report_preview.html', {
        'purchases': purchases,
        'report_type': report_type,
        'total_purchases': total_purchases,
        'total_amount': total_amount,
        'total_paid': total_paid,
        'total_unpaid': total_unpaid,
        'total_invoice_amount': total_invoice_amount,
        'pending_invoices': pending_invoices,
    })
    
    return JsonResponse({'html': html})


@login_required
def purchase_report_excel(request):
    """Generate Purchase Report as Excel"""
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    supplier_id = request.GET.get('supplier_id', '').strip()
    report_type = request.GET.get('report_type', 'detail')  # detail or summary
    
    # Build query
    queryset = Purchase.objects.select_related('supplier', 'created_by', 'received_by').order_by('-tanggal_purchase')
    
    # Apply filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(tanggal_purchase__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            end_of_day = date_to_obj + timedelta(days=1)
            queryset = queryset.filter(tanggal_purchase__lt=end_of_day)
        except:
            pass
    
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)
    
    purchases = queryset.all()
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Report"
    
    # Header
    ws['A1'] = 'PURCHASE REPORT'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = 'PT. ALFA ERP'
    ws['A2'].font = Font(size=12)
    ws.merge_cells('A2:E2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Report Info
    row = 4
    ws[f'A{row}'] = f'Periode: {date_from} s/d {date_to}'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = f'Report Type: {report_type.upper()}'
    ws[f'A{row}'].font = Font(bold=True)
    row += 2
    
    if report_type == 'detail':
        # Detail Report
        headers = ['Date', 'Purchase Number', 'Supplier', 'Product', 'Barcode', 'Qty', 'Harga Beli', 'Subtotal', 'Status', 'Received By', 'Payment Status', 'Tax Invoice Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for purchase in purchases:
            items = PurchaseItem.objects.filter(purchase=purchase).select_related('product')
            for item in items:
                ws[f'A{row}'] = purchase.tanggal_purchase.strftime('%d %b %Y')
                ws[f'B{row}'] = purchase.nomor_purchase
                ws[f'C{row}'] = purchase.supplier.nama_supplier
                ws[f'D{row}'] = item.product.nama_produk
                ws[f'E{row}'] = item.product.barcode or item.product.sku
                ws[f'F{row}'] = item.quantity
                ws[f'G{row}'] = item.harga_beli
                ws[f'H{row}'] = item.subtotal
                ws[f'I{row}'] = purchase.get_status_display()
                ws[f'J{row}'] = purchase.received_by.get_full_name() if purchase.received_by else '-'
                
                # Payment Status
                payment = purchase.payment.first()
                if payment:
                    payment_status = payment.get_status_display()
                    payment_amount = f"Rp {payment.paid_amount:,} / Rp {payment.total_amount:,}"
                    ws[f'K{row}'] = f"{payment_status}\n{payment_amount}"
                else:
                    ws[f'K{row}'] = ''
                
                # Tax Invoice Status
                tax_invoice = purchase.tax_invoice.first()
                if tax_invoice:
                    tax_status = tax_invoice.get_status_display()
                    invoice_number = tax_invoice.invoice_number or 'No Invoice'
                    ws[f'L{row}'] = f"{tax_status}\n{invoice_number}"
                else:
                    ws[f'L{row}'] = ''
                
                row += 1
        
        # Total
        ws[f'G{row}'] = 'TOTAL:'
        ws[f'G{row}'].font = Font(bold=True)
        ws[f'G{row}'].alignment = Alignment(horizontal='right')
        ws[f'H{row}'] = sum(p.total_amount for p in purchases)
        ws[f'H{row}'].font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 25
        ws.column_dimensions['L'].width = 25
    
    else:
        # Summary Report
        headers = ['Date', 'Purchase Number', 'Supplier', 'Total Amount', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for purchase in purchases:
            ws[f'A{row}'] = purchase.tanggal_purchase.strftime('%d %b %Y')
            ws[f'B{row}'] = purchase.nomor_purchase
            ws[f'C{row}'] = purchase.supplier.nama_supplier
            ws[f'D{row}'] = purchase.total_amount
            ws[f'E{row}'] = purchase.get_status_display()
            row += 1
        
        # Total
        ws[f'C{row}'] = 'TOTAL:'
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].alignment = Alignment(horizontal='right')
        ws[f'D{row}'] = sum(p.total_amount for p in purchases)
        ws[f'D{row}'].font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
    
    # Generate filename
    filename = f"Purchase_Report_{date_from}_{date_to}_{report_type}.xlsx"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def purchase_report_pdf(request):
    """Generate Purchase Report as PDF"""
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    supplier_id = request.GET.get('supplier_id', '').strip()
    report_type = request.GET.get('report_type', 'detail')
    
    # Build query
    queryset = Purchase.objects.select_related('supplier', 'created_by', 'received_by').order_by('-tanggal_purchase')
    
    # Apply filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(tanggal_purchase__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            end_of_day = date_to_obj + timedelta(days=1)
            queryset = queryset.filter(tanggal_purchase__lt=end_of_day)
        except:
            pass
    
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)
    
    purchases = queryset.all()
    
    # Calculate statistics
    total_purchases = purchases.count()
    total_amount = sum(p.total_amount for p in purchases)
    
    context = {
        'purchases': purchases,
        'date_from': date_from,
        'date_to': date_to,
        'report_type': report_type,
        'total_purchases': total_purchases,
        'total_amount': total_amount,
    }
    
    html = render_to_string('purchasing/purchase_report_pdf.html', context)
    
    # Convert HTML to PDF using xhtml2pdf
    from xhtml2pdf import pisa
    from io import BytesIO
    
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        filename = f"Purchase_Report_{date_from}_{date_to}_{report_type}.pdf"
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    return HttpResponse(f'Error generating PDF: {pdf.err}', status=500)

