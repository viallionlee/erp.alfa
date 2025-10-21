import json
import random
import io
from importlib import import_module
from django.shortcuts import render
from django.http import JsonResponse
from django.db import transaction
from django.db.models import Q, F, Count, Window, Min
from django.db.models.functions import RowNumber
from django.utils import timezone
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_GET
from orders.models import Order, OrderShippingHistory, OrderPackingHistory
import pytz
from django.db.models import CharField, Value as V
from django.db.models.functions import Concat
from collections import defaultdict
from django.core.exceptions import PermissionDenied


COURIER_GROUPS = {
    "all": [
        # Semua kurir digabung
        "LEX", "Shopee Xpress", "Ninja Pusat", "Ninja Agen", 
        "JNE", "J&T", "GTL", "CARGO", "instant"
    ],
    "global": [
        # Semua kurir (sama seperti "all")
        "LEX", "Shopee Xpress", "Ninja Pusat", "Ninja Agen", 
        "JNE", "J&T", "GTL", "CARGO", "instant"
    ],
    "LEX": [
        "Pickup: LEX ID, Delivery: LEX ID",
        "Pickup: LEX ID, Delivery: null",
        "Pickup: LEX ID, Delivery: JNE",
        "Pickup: LEX ID, Delivery: J&T"
    ],
    "Shopee Xpress": [
        "SPX Standard", "SPX Hemat", "SPX Instant", "Agen SPX Express",
        "SPX Sameday", "SPX Express Point"
    ],
    "Ninja Pusat": ["NinjaVan Indonesia"],
    "Ninja Agen": ["Pickup: Ninja", "Pickup: NinjaVanID", "Delivery: NinjaVanID"],
    "JNE": ["JNE JNE Trucking", "JNE Reguler", "JNE Cashless"],
    "J&T": ["J&T-MP", "J&T Reguler", "JNT CASHLESS", "J&T Cargo", "J&T Express"],
    "GTL": ["GoTo Logistics GTL"],
    "CARGO": ["Hemat Kargo", "Kurir Rekomendasi Reguler", "Pickup: J&T CARGO", "Delivery: J&T CARGO"],
    "instant": ["GoSend Instant", "GrabExpress Instant", "Gojek", "GoSend Same Day 8 Jam", "GoSend Instant 3 Jam"]
}

def get_motivational_message_list(rank, total_shippers):
    """
    Mengambil daftar pesan motivasi berdasarkan peringkat user.
    """
    # Logika untuk peringkat top 3
    if rank == 1:
        module_name = 'rank_1'
    elif rank == 2:
        module_name = 'rank_2'
    elif rank == 3:
        module_name = 'rank_3'
    # Logika untuk peringkat "umum" di tengah
    elif rank > 5:
        module_name = 'rank_umum'
    # Fallback untuk kasus yang tidak terduga
    else:
        module_name = 'default'
    
    try:
        module_path = f'fullfilment.motivasi.{module_name}'
        motivation_module = import_module(module_path)
        return motivation_module.MESSAGES
    except (ImportError, AttributeError):
        # Fallback jika file spesifik belum dibuat.
        from fullfilment.motivasi import default
        return default.MESSAGES

def is_order_in_courier_group(order_kurir, selected_courier):
    """
    Validasi kurir order berdasarkan grup yang dipilih (case-insensitive substring).
    """
    k = (order_kurir or '').lower()
    
    # Langsung return True jika 'all', tidak perlu proses lebih lanjut
    if selected_courier == 'all':
        return True
    
    # Mapping grup ke keyword pencarian
    group_keywords = {
        'LEX': ['lex'],
        'J&T': ['j&t', 'jnt'],
        'JNE': ['jne'],
        'Shopee Xpress': ['spx'],
        'Ninja Pusat': ['ninja'], # Asumsi 'ninja' cukup
        'Ninja Agen': ['ninja'], # Asumsi 'ninja' cukup
        'GTL': ['gtl', 'goto'],
        'CARGO': ['kargo', 'cargo'],
        'instant': ['instant', 'sameday', 'gojek', 'grab']
    }

    keywords = group_keywords.get(selected_courier)
    if keywords:
        return any(keyword in k for keyword in keywords)
    
    return False

@login_required
@permission_required('fullfilment.view_desktop_shipping_module', raise_exception=True)
def scanshipping(request):
    """
    Menampilkan halaman Ready to Ship dan memproses update status via AJAX.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            scan_input = data.get('order_id', '').strip()
            selected_courier = data.get('courier', '').strip()

            if not scan_input:
                return JsonResponse({'success': False, 'message': 'Input tidak boleh kosong.'})
            
            if not selected_courier:
                return JsonResponse({'success': False, 'message': 'Kurir belum dipilih.'})

            with transaction.atomic():
                orders = Order.objects.filter(
                    Q(id_pesanan__iexact=scan_input) | Q(awb_no_tracking__iexact=scan_input)
                )

                if not orders.exists():
                    return JsonResponse({'success': False, 'message': f"Order dengan ID atau Resi '{scan_input}' tidak ditemukan."})

                order = orders.first()

                # NEW: Validasi status pembayaran atau status fulfillment adalah 'batal'/'cancel'
                if 'batal' in (order.status or '').lower() or \
                   'cancel' in (order.status or '').lower() or \
                   'batal' in (order.status_order or '').lower() or \
                   'cancel' in (order.status_order or '').lower():
                    from .models import OrderCancelLog # Import here to avoid circular dependency if placed at top
                    # RE-ADDED: Cek apakah OrderCancelLog sudah ada untuk Order ID ini
                    if not OrderCancelLog.objects.filter(order_id_scanned=scan_input).exists():
                        OrderCancelLog.objects.create(
                            order_id_scanned=scan_input,
                            user=request.user if request.user.is_authenticated else None,
                            status_pembayaran_at_scan=order.status,
                            status_fulfillment_at_scan=order.status_order
                        )
                    return JsonResponse({
                        'success': False,
                        'message': f"Halo {request.user.username} , Orderan ini {scan_input} sudah dibatalkan customer , tolong di retur ke tim retur , Terima Kasih , Semangat terus!"
                    })

                # --- VALIDASI STATUS DAN KURIR ---
                if order.status_order == 'shipped':
                    last_ship = OrderShippingHistory.objects.filter(order=order).order_by('-waktu_ship').first()
                    if last_ship:
                        jakarta_tz = pytz.timezone('Asia/Jakarta')
                        waktu_jakarta = last_ship.waktu_ship.astimezone(jakarta_tz)
                        message = f"Pesanan '{scan_input}' sudah di-SHIPPED oleh {last_ship.user.username} pada {waktu_jakarta.strftime('%d-%m-%Y %H:%M')}. Tolong kembalikan kertas ini ke admin."
                    else:
                        message = f"Pesanan '{scan_input}' sudah di-SHIPPED sebelumnya. Tolong kembalikan kertas ini ke admin."
                    return JsonResponse({'success': False, 'message': message})

                if order.status_order != 'packed':
                    return JsonResponse({'success': False, 'message': f"Order ini belum di-pack. Status saat ini: {order.status_order}."})

                if not is_order_in_courier_group(order.kurir, selected_courier):
                    return JsonResponse({'success': False, 'message': f"Order ini bukan milik kurir '{selected_courier}'. Kurir order: {order.kurir}."})

                # --- JIKA SEMUA VALID, LANJUTKAN PROSES ---
                updated_count = orders.update(status_order='shipped')
                # Ubah dari OrderHandoverHistory ke OrderShippingHistory
                history_entries = [
                    OrderShippingHistory(order=o, user=request.user) for o in orders
                ]
                OrderShippingHistory.objects.bulk_create(history_entries) # Ubah nama model di sini
                
                # --- Start of Scan Counter Logic ---
                scan_count = request.session.get('scan_count_for_message', 0) + 1
                request.session['scan_count_for_message'] = scan_count
                
                if scan_count >= 10:
                    request.session.pop('motivational_message', None)
                    request.session.pop('last_rank', None)
                    request.session['scan_count_for_message'] = 0
                # --- End of Scan Counter Logic ---

                return JsonResponse({
                    'success': True,
                    'message': f"Berhasil! {updated_count} item untuk pesanan '{scan_input}' telah diubah menjadi 'Shipped'."
                })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Format request tidak valid.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Terjadi error internal: {str(e)}'}, status=500)
    
    today = timezone.now().date()
    current_user = request.user

    user_scores = OrderShippingHistory.objects.filter(
        waktu_ship__date=today
    ).values('user_id').annotate(
        total_ships=Count('order__id_pesanan', distinct=True)
    ).order_by('-total_ships')

    total_shippers_today = len(user_scores)
    user_ships_today = 0
    user_rank_today = 0
    
    for i, score in enumerate(user_scores):
        if score['user_id'] == current_user.id:
            user_rank_today = i + 1
            user_ships_today = score['total_ships']
            break
            
    if user_rank_today == 0:
        user_rank_today = total_shippers_today + 1

    # --- Motivational Message Logic with 10-Scan Cycle ---
    stored_message = request.session.get('motivational_message')
    stored_rank = request.session.get('last_rank')
    
    new_message_needed = (stored_message is None) or (user_rank_today != stored_rank)

    if new_message_needed:
        message_list = get_motivational_message_list(user_rank_today, total_shippers_today)
        specific_message = random.choice(message_list)
        motivational_message = f"Hi, <strong>{current_user.username}</strong>! {specific_message}"
        
        request.session['motivational_message'] = motivational_message
        request.session['last_rank'] = user_rank_today
    else:
        motivational_message = stored_message

    now = timezone.now()
    start_of_work = now.replace(hour=9, minute=0, second=0, microsecond=0)
    ships_per_hour = 0.0 

    if now > start_of_work:
        duration_seconds = (now - start_of_work).total_seconds()
        if duration_seconds > 60:
            duration_hours = duration_seconds / 3600
            ships_per_hour = user_ships_today / duration_hours
        else:
            ships_per_hour = float(user_ships_today)
            
    last_10_shipping = OrderShippingHistory.objects.annotate(
        row_num=Window(
            expression=RowNumber(),
            partition_by=[F('order__id_pesanan')],
            order_by=F('waktu_ship').desc()
        )
    ).filter(row_num=1).select_related('user', 'order').order_by('-waktu_ship')[:10]
    
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = any(x in user_agent for x in ['android', 'iphone', 'ipad', 'ipod', 'blackberry', 'iemobile', 'opera mini'])
    template_name = 'fullfilment/mobile_scanshipping.html' if is_mobile else 'fullfilment/scanshipping.html'
    
    # --- Tambahkan kalkulasi untuk statistik ---
    total_batched_orders = Order.objects.filter(nama_batch__isnull=False).values('id_pesanan').distinct().count()
    total_shipped_orders = Order.objects.filter(status_order='shipped', nama_batch__isnull=False).values('id_pesanan').distinct().count()

    context = {
        'last_10_shipping': last_10_shipping,
        'user_ships_today': user_ships_today,
        'user_rank_today': user_rank_today,
        'total_shippers_today': total_shippers_today,
        'ships_per_hour': ships_per_hour,
        'motivational_message': motivational_message,
        'total_batched': total_batched_orders,
        'total_shipped': total_shipped_orders,
    }
    return render(request, template_name, context) 

@login_required
def shipping_history_view(request):
    """Menampilkan halaman riwayat pengiriman."""
    import re
    
    # Mobile detection
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = re.search(r'mobile|android|iphone', user_agent)
    
    template_name = 'fullfilment/shipping_history_mobile.html' if is_mobile else 'fullfilment/shippinghistory.html'
    return render(request, template_name)

@login_required
def shipping_history_api(request):
    """API untuk DataTables riwayat pengiriman."""
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 50))
    search_value = request.GET.get('search[value]', '').strip()

    # Mendefinisikan urutan kolom
    order_column_index = int(request.GET.get('order[0][column]', 1))
    order_dir = request.GET.get('order[0][dir]', 'desc')
    
    column_mapping = ['id', 'waktu_ship', 'order__id_pesanan', 'user__username', 'order__kurir']
    order_column_name = column_mapping[order_column_index]
    
    sort_prefix = '-' if order_dir == 'desc' else ''
    order_by = f'{sort_prefix}{order_column_name}'

    # Queryset dasar
    queryset = OrderShippingHistory.objects.select_related('order', 'user').all()

    # Total record sebelum filter
    total_records = queryset.count()

    # Terapkan filter pencarian
    if search_value:
        queryset = queryset.filter(
            Q(order__id_pesanan__icontains=search_value) |
            Q(user__username__icontains=search_value) |
            Q(order__kurir__icontains=search_value)
        )

    # Total record setelah filter
    filtered_records = queryset.count()
    
    # Terapkan pengurutan dan paginasi
    queryset = queryset.order_by(order_by)[start:start + length]

    # Format data untuk JSON response
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    data = []
    for history in queryset:
        waktu_ship_localized = history.waktu_ship.astimezone(jakarta_tz).isoformat() if history.waktu_ship else None
        data.append({
            'id': history.id,
            'waktu_ship': waktu_ship_localized,
            'order_id_pesanan': history.order.id_pesanan if history.order else 'N/A',
            'user': history.user.username if history.user else 'N/A',
            'kurir': history.order.kurir if history.order and history.order.kurir else 'N/A'
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data,
    })

@login_required
@require_GET
def shipping_history_data_api(request):
    """API endpoint untuk mendapatkan data shipping history dalam format JSON untuk mobile"""
    try:
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        # Get all shipping history
        queryset = OrderShippingHistory.objects.select_related('order', 'user').order_by('-waktu_ship')
        
        # Calculate summary statistics
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        total_count = queryset.count()
        today_count = queryset.filter(waktu_ship__date=today).count()
        week_count = queryset.filter(waktu_ship__date__gte=week_ago).count()
        
        # Format data for mobile
        jakarta_tz = pytz.timezone('Asia/Jakarta')
        data = []
        
        for history in queryset:
            waktu_ship_localized = history.waktu_ship.astimezone(jakarta_tz) if history.waktu_ship else None
            data.append({
                'id': history.id,
                'waktu_ship': waktu_ship_localized.isoformat() if waktu_ship_localized else None,
                'order_id': history.order.id_pesanan if history.order else 'N/A',
                'courier': history.order.kurir if history.order and history.order.kurir else 'N/A',
                'user': history.user.username if history.user else 'N/A'
            })
        
        summary = {
            'total': total_count,
            'today': today_count,
            'week': week_count
        }
        
        return JsonResponse({
            'success': True,
            'data': data,
            'summary': summary
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500) 

@login_required
def order_shipping_list_view(request):
    """Menampilkan halaman daftar order shipping dengan pagination server-side."""
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = any(mobile in user_agent for mobile in ['mobile', 'android', 'iphone', 'ipad'])

    if is_mobile:
        if not request.user.has_perm('fullfilment.view_mobile_shipping_module'):
            raise PermissionDenied
    else:
        if not request.user.has_perm('fullfilment.view_desktop_shipping_module'):
            raise PermissionDenied
            
    if is_mobile:
        return render(request, 'fullfilment/mobile_order_shipping_list.html')
    else:
        return render(request, 'fullfilment/order_shipping_list.html')

@login_required
def order_shipping_list_api(request):
    """API untuk DataTables daftar order shipping dengan pagination per 25."""
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))  # Pagination per 25
    search_value = request.GET.get('search[value]', '').strip()

    # Mendefinisikan urutan kolom
    order_column_index = int(request.GET.get('order[0][column]', 1))
    order_dir = request.GET.get('order[0][dir]', 'desc')
    
    column_mapping = ['id', 'waktu_ship', 'order__id_pesanan', 'user__username', 'order__kurir', 'order__nama_toko']
    order_column_name = column_mapping[order_column_index]
    
    sort_prefix = '-' if order_dir == 'desc' else ''
    order_by = f'{sort_prefix}{order_column_name}'

    # Queryset dasar
    queryset = OrderShippingHistory.objects.select_related('order', 'user').all()

    # Total record sebelum filter
    total_records = queryset.count()

    # Terapkan filter pencarian
    if search_value:
        queryset = queryset.filter(
            Q(order__id_pesanan__icontains=search_value) |
            Q(user__username__icontains=search_value) |
            Q(order__kurir__icontains=search_value) |
            Q(order__nama_toko__icontains=search_value)
        )

    # Total record setelah filter
    filtered_records = queryset.count()
    
    # Terapkan pengurutan dan paginasi
    queryset = queryset.order_by(order_by)[start:start + length]

    # Format data untuk JSON response
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    data = []
    for history in queryset:
        waktu_ship_localized = history.waktu_ship.astimezone(jakarta_tz).strftime('%d-%m-%Y %H:%M:%S') if history.waktu_ship else 'N/A'
        data.append({
            'id': history.id,
            'waktu_ship': waktu_ship_localized,
            'order_id_pesanan': history.order.id_pesanan if history.order else 'N/A',
            'nama_toko': history.order.nama_toko if history.order else 'N/A',
            'user': history.user.username if history.user else 'N/A',
            'kurir': history.order.kurir if history.order and history.order.kurir else 'N/A',
            'awb_no_tracking': history.order.awb_no_tracking if history.order else 'N/A'
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data,
    })

@login_required
def order_shipping_report_view(request):
    """Menampilkan halaman laporan shipping per kurir dengan filter tanggal."""
    from datetime import datetime, date
    from collections import defaultdict
    
    # Get selected date from request (default: today)
    selected_date_str = request.GET.get('date', '')
    
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().astimezone(jakarta_tz).date()
    else:
        selected_date = timezone.now().astimezone(jakarta_tz).date()
    
    # Query shipping history untuk tanggal yang dipilih
    shipping_histories = OrderShippingHistory.objects.filter(
        waktu_ship__date=selected_date
    ).select_related('order', 'user')
    
    # Group by courier
    from collections import defaultdict
    courier_data = defaultdict(lambda: {'count': 0, 'orders': []})
    
    for history in shipping_histories:
        courier_name = history.order.kurir if history.order and history.order.kurir else 'Unknown'
        
        # Normalize courier name
        courier_normalized = courier_name.strip()
        if not courier_normalized:
            courier_normalized = 'Unknown'
        
        courier_data[courier_normalized]['count'] += 1
        courier_data[courier_normalized]['orders'].append({
            'id_pesanan': history.order.id_pesanan if history.order else 'N/A',
            'nama_toko': history.order.nama_toko if history.order else 'N/A',
            'user': history.user.username if history.user else 'N/A',
            'waktu_ship': history.waktu_ship.astimezone(jakarta_tz) if history.waktu_ship else None,
            'awb_no_tracking': history.order.awb_no_tracking if history.order else 'N/A'
        })
    
    # Convert to sorted list
    courier_stats = []
    for courier_name, data in courier_data.items():
        courier_stats.append({
            'name': courier_name,
            'count': data['count'],
            'orders': data['orders']
        })
    
    # Sort by count (descending)
    courier_stats.sort(key=lambda x: x['count'], reverse=True)
    
    # Calculate totals
    total_shipped = sum(stat['count'] for stat in courier_stats)
    
    context = {
        'courier_stats': courier_stats,
        'total_shipped': total_shipped,
        'selected_date': selected_date,
        'selected_date_str': selected_date.strftime('%Y-%m-%d'),
        'today': timezone.now().astimezone(jakarta_tz).date()
    }
    
    return render(request, 'fullfilment/ordershippingreport.html', context)

@login_required
def order_shipping_report_download_excel(request):
    """Download Excel report untuk shipping per kurir."""
    try:
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
    except ImportError as e:
        return HttpResponse(f"Error importing openpyxl: {str(e)}", status=500)
    
    # Get selected date from request (default: today)
    selected_date_str = request.GET.get('date', '')
    courier_name = request.GET.get('courier', '')
    
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().astimezone(jakarta_tz).date()
    else:
        selected_date = timezone.now().astimezone(jakarta_tz).date()
    
    # Query shipping history untuk tanggal yang dipilih
    shipping_histories = OrderShippingHistory.objects.filter(
        waktu_ship__date=selected_date
    ).select_related('order', 'user')
    
    # Filter by courier if specified
    if courier_name:
        shipping_histories = shipping_histories.filter(order__kurir=courier_name)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Shipping Report {selected_date.strftime('%d-%m-%Y')}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'No', 'Order ID', 'Nama Toko', 'Kurir', 'User Shipper', 'User Packer',
        'Waktu Ship', 'AWB/Tracking'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    for idx, history in enumerate(shipping_histories, 2):
        # Get user packer from OrderPackingHistory
        user_packer = 'N/A'
        if history.order:
            packing_history = OrderPackingHistory.objects.filter(order=history.order).first()
            if packing_history and packing_history.user:
                user_packer = packing_history.user.username
        
        ws.cell(row=idx, column=1, value=idx-1).border = border
        ws.cell(row=idx, column=2, value=history.order.id_pesanan if history.order else 'N/A').border = border
        ws.cell(row=idx, column=3, value=history.order.nama_toko if history.order else 'N/A').border = border
        ws.cell(row=idx, column=4, value=history.order.kurir if history.order else 'N/A').border = border
        ws.cell(row=idx, column=5, value=history.user.username if history.user else 'N/A').border = border
        ws.cell(row=idx, column=6, value=user_packer).border = border
        ws.cell(row=idx, column=7, value=history.waktu_ship.astimezone(jakarta_tz).strftime('%d-%m-%Y %H:%M:%S') if history.waktu_ship else 'N/A').border = border
        ws.cell(row=idx, column=8, value=history.order.awb_no_tracking if history.order else 'N/A').border = border
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f"{column_letter}{row}"].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    filename = f"shipping_report_{selected_date.strftime('%Y%m%d')}"
    if courier_name:
        filename += f"_{courier_name.replace(' ', '_')}"
    filename += ".xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def order_shipping_detail_view(request):
    """Menampilkan halaman detail shipping per kurir."""
    from datetime import datetime
    
    # Get parameters
    courier_name = request.GET.get('courier', '')
    selected_date_str = request.GET.get('date', '')
    
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().astimezone(jakarta_tz).date()
    else:
        selected_date = timezone.now().astimezone(jakarta_tz).date()
    
    # Count total orders for this courier and date
    total_count = OrderShippingHistory.objects.filter(
        waktu_ship__date=selected_date,
        order__kurir=courier_name
    ).count()
    
    # Get unique toko names and users for dropdowns
    unique_toko = OrderShippingHistory.objects.filter(
        waktu_ship__date=selected_date,
        order__kurir=courier_name
    ).values_list('order__nama_toko', flat=True).distinct().order_by('order__nama_toko')
    
    unique_users = OrderShippingHistory.objects.filter(
        waktu_ship__date=selected_date,
        order__kurir=courier_name
    ).values_list('user__username', flat=True).distinct().order_by('user__username')
    
    context = {
        'courier_name': courier_name,
        'selected_date': selected_date,
        'selected_date_str': selected_date.strftime('%Y-%m-%d'),
        'total_count': total_count,
        'unique_toko': unique_toko,
        'unique_users': unique_users
    }
    
    return render(request, 'fullfilment/order_shipping_detail.html', context)

@login_required
def order_shipping_detail_api(request):
    """API untuk halaman detail shipping dengan pagination, search, dan filter."""
    from datetime import datetime
    
    # Get parameters
    courier_name = request.GET.get('courier', '')
    selected_date_str = request.GET.get('date', '')
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 100))
    
    # Search and filter parameters
    search_order_id = request.GET.get('search_order_id', '').strip()
    search_toko = request.GET.get('search_toko', '').strip()
    search_user = request.GET.get('search_user', '').strip()
    search_awb = request.GET.get('search_awb', '').strip()
    
    try:
        jakarta_tz = pytz.timezone('Asia/Jakarta')
        if selected_date_str:
            try:
                selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
            except ValueError:
                selected_date = timezone.now().astimezone(jakarta_tz).date()
        else:
            selected_date = timezone.now().astimezone(jakarta_tz).date()
        
        # Base query shipping history
        queryset = OrderShippingHistory.objects.filter(
            waktu_ship__date=selected_date,
            order__kurir=courier_name
        ).select_related('order', 'user')
        
        # Apply search filters
        if search_order_id:
            queryset = queryset.filter(order__id_pesanan__icontains=search_order_id)
        
        if search_toko:
            queryset = queryset.filter(order__nama_toko__icontains=search_toko)
        
        if search_user:
            queryset = queryset.filter(user__username__icontains=search_user)
        
        if search_awb:
            queryset = queryset.filter(order__awb_no_tracking__icontains=search_awb)
        
        # Order by waktu_ship descending
        queryset = queryset.order_by('-waktu_ship')
        
        # Count total
        total_count = queryset.count()
        
        # Calculate pagination
        start = (page - 1) * per_page
        end = start + per_page
        queryset = queryset[start:end]
        
        # Format data
        data = []
        for history in queryset:
            waktu_ship_localized = history.waktu_ship.astimezone(jakarta_tz).strftime('%d-%m-%Y %H:%M:%S') if history.waktu_ship else 'N/A'
            data.append({
                'order_id_pesanan': history.order.id_pesanan if history.order else 'N/A',
                'nama_toko': history.order.nama_toko if history.order else 'N/A',
                'user': history.user.username if history.user else 'N/A',
                'waktu_ship': waktu_ship_localized,
                'awb_no_tracking': history.order.awb_no_tracking if history.order else 'N/A'
            })
        
        return JsonResponse({
            'success': True,
            'data': data,
            'total_count': total_count,
            'page': page,
            'per_page': per_page,
            'total_pages': (total_count + per_page - 1) // per_page
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def order_shipping_report_download_pdf(request):
    """Download PDF report untuk shipping per kurir."""
    try:
        from datetime import datetime
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError as e:
        return HttpResponse(f"Error importing PDF libraries: {str(e)}", status=500)
    
    # Get selected date from request (default: today)
    selected_date_str = request.GET.get('date', '')
    courier_name = request.GET.get('courier', '')
    
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().astimezone(jakarta_tz).date()
    else:
        selected_date = timezone.now().astimezone(jakarta_tz).date()
    
    # Query shipping history untuk tanggal yang dipilih
    shipping_histories = OrderShippingHistory.objects.filter(
        waktu_ship__date=selected_date
    ).select_related('order', 'user')
    
    # Filter by courier if specified
    if courier_name:
        shipping_histories = shipping_histories.filter(order__kurir=courier_name)
    
    # Group by courier for summary
    from collections import defaultdict
    courier_data = defaultdict(lambda: {'count': 0, 'orders': []})
    
    for history in shipping_histories:
        courier_name_item = history.order.kurir if history.order and history.order.kurir else 'Unknown'
        courier_normalized = courier_name_item.strip()
        if not courier_normalized:
            courier_normalized = 'Unknown'
        
        courier_data[courier_normalized]['count'] += 1
        courier_data[courier_normalized]['orders'].append({
            'id_pesanan': history.order.id_pesanan if history.order else 'N/A',
            'nama_toko': history.order.nama_toko if history.order else 'N/A',
            'user': history.user.username if history.user else 'N/A',
            'waktu_ship': history.waktu_ship.astimezone(jakarta_tz) if history.waktu_ship else None,
            'awb_no_tracking': history.order.awb_no_tracking if history.order else 'N/A'
        })
    
    # Convert to sorted list
    courier_stats = []
    for courier_name_item, data in courier_data.items():
        courier_stats.append({
            'name': courier_name_item,
            'count': data['count'],
            'orders': data['orders']
        })
    
    # Sort by count (descending)
    courier_stats.sort(key=lambda x: x['count'], reverse=True)
    
    # Calculate totals
    total_shipped = sum(stat['count'] for stat in courier_stats)
    
    context = {
        'courier_stats': courier_stats,
        'total_shipped': total_shipped,
        'selected_date': selected_date,
        'selected_date_str': selected_date.strftime('%Y-%m-%d'),
        'courier_filter': courier_name,
        'shipping_histories': shipping_histories
    }
    
    # Use reportlab for PDF generation
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    title = Paragraph("Laporan Shipping per Kurir", title_style)
    story.append(title)
    
    # Date info
    date_text = f"Tanggal: {selected_date.strftime('%d %B %Y')}"
    if courier_name:
        date_text += f"<br/>Kurir: {courier_name}"
    story.append(Paragraph(date_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Summary
    summary_data = [
        ['Total Dikirim', str(total_shipped)],
        ['Jumlah Kurir', str(len(courier_stats))],
        ['Tanggal', selected_date.strftime('%d %B %Y')]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Courier data
    if courier_name:
        # Single courier
        courier = courier_stats[0] if courier_stats else None
        if courier:
            story.append(Paragraph(f"<b>{courier['name']} - {courier['count']} order</b>", styles['Heading2']))
            
            # Table headers
            table_data = [['No', 'Order ID', 'Waktu Ship', 'AWB/Tracking']]
            
            # Table data
            for idx, order in enumerate(courier['orders'], 1):
                table_data.append([
                    str(idx),
                    order['id_pesanan'],
                    order['waktu_ship'].strftime('%d-%m-%Y %H:%M') if order['waktu_ship'] else 'N/A',
                    order['awb_no_tracking']
                ])
            
            # Create table
            table = Table(table_data, colWidths=[0.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
    else:
        # All couriers
        for courier in courier_stats:
            story.append(Paragraph(f"<b>{courier['name']} - {courier['count']} order</b>", styles['Heading2']))
            
            # Table headers
            table_data = [['No', 'Order ID', 'Waktu Ship', 'AWB/Tracking']]
            
            # Table data
            for idx, order in enumerate(courier['orders'], 1):
                table_data.append([
                    str(idx),
                    order['id_pesanan'],
                    order['waktu_ship'].strftime('%d-%m-%Y %H:%M') if order['waktu_ship'] else 'N/A',
                    order['awb_no_tracking']
                ])
            
            # Create table
            table = Table(table_data, colWidths=[0.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(story)
    pdf_buffer.seek(0)
    
    # Create response
    filename = f"shipping_report_{selected_date.strftime('%Y%m%d')}"
    if courier_name:
        filename += f"_{courier_name.replace(' ', '_')}"
    filename += ".pdf"
    
    response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
